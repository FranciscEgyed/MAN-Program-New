import csv
import os
import re
from tkinter import filedialog, messagebox, Tk, ttk, HORIZONTAL, Label
import pandas as pd
from openpyxl.reader.excel import load_workbook


def apfr():
    save_file = os.path.abspath(os.curdir) + "/MAN/Output/APFR.xlsx"
    array_clasificare_std = [[8001, 'HANDSCHNEIDEPLATZ LEITUNGEN', 'X', None, None, 'CST'],
                             [8002, 'KOMAX O,C/1S/2S/DRUCK', 'X', None, None, 'CST'],
                             [8003, 'KOMAX OC/1S/2S', 'X', None, None, 'CST'],
                             [8004, 'KOMAX OHNE CRIMP ab 6,00 QMM', None, 'X', None, 'SUBBY'],
                             [8009, 'KOMAX 42 TUELLE', 'X', None, None, 'CST'],
                             [8010, 'KOMAX 42 TUELLE/DRUCK', 'X', None, None, 'CST'],
                             [8010, 'KOMAX 42 TUELLE/DRUCK', 'X', None, None, 'CST'],
                             [8021, 'HALBAUTOMAT 21916', None, 'X', None, 'SUBBY'],
                             [8025, 'HALBAUTOMAT 22077', None, 'X', None, 'SUBBY'],
                             [8029, 'HANDLOETPLATZ', 'X', None, None, 'CST'],
                             [8030, 'HANDSCHNEIDEPLATZ SCHLAUCH', None, 'X', None, 'SUBBY'],
                             [8031, 'ARTOS-SCHLAUCHSCHNEIDEMASCHINE', None, 'X', None, 'SUBBY'],
                             [8032, 'KABELMAT-SCHLAUCHSCHNEIDEMASCHINE', 'X', None, None, 'CST'],
                             [8040, 'ABISOLIERBOX', 'X', None, None, 'CST'],
                             [8041, 'ABMANTELPLATZ', None, 'X', None, 'SUBBY'],
                             [8043, 'ANSCHLAGBEREICH', None, 'X', None, 'SUBBY'],
                             [8046, 'SPRITZMASCHINE', 'X', None, None, 'CST'],
                             [8047, 'SCHAUMANLAGE', None, 'X', None, 'SUBBY'],
                             [8048, 'VM-ANSCHLAGBEREICH', 'X', None, None, 'CST'],
                             [8049, 'SCHRUMPFAUTOMAT', 'X', None, None, 'CST'],
                             [8051, 'ZINNBAD', 'X', None, None, 'CST'],
                             [8052, 'STOCKO-PRESSE', None, 'X', None, 'SUBBY'],
                             [8052, 'STOCKO-PRESSE', None, 'X', None, 'SUBBY'],
                             [8060, 'MONTAGE FENDT ALLGEMEIN', None, None, 'X', 'ASSY'],
                             [8066, 'VERDRILLMACHINE', 'X', None, None, 'CST'],
                             [8068, 'FRAESPLATZ', 'X', None, None, 'CST'],
                             [8069, 'SUBBY', None, 'x', None, 'SUBBY'],
                             [8103, '8103', None, 'X', None, 'SUBBY'],
                             [8260, '8260', None, None, 'X', 'ASSY'],
                             [8295, '8295', None, None, 'X', 'ASSY'],
                             [8298, '8298', None, None, 'X', 'ASSY'],
                             [8350, 'SCHWEISSAUTOMAT', 'X', None, None, 'CST'],
                             [8352, 'STOCKO-PRESSE', None, None, None, 'SUBBY'],
                             [8360, 'MONTAGE NFZ LH', None, None, 'X', 'ASSY'],
                             [8365, 'VM-MONTAGE NFZ LH', 'X', None, None, 'CST'],
                             [8370, 'SILBERETIKETT KLEBEN', None, 'X', None, 'SUBBY'],
                             [8370, 'SILBERETIKETT KLEBEN', None, 'X', None, 'SUBBY'],
                             [8395, 'PRUEFTISCH 1 NFZ', None, None, 'X', 'ASSY'],
                             [8396, 'PRUEFTISCH 2 NFZ', None, None, None, 'ASSY'],
                             [8398, 'PACKTISCH', None, None, 'X', 'ASSY'],
                             [10001, 'Cutting general', 'x', None, None, 'CST'],
                             [10002, '#N/A', 'x', None, None, 'CST'],
                             [10003, 'Cutting with print', 'x', None, None, 'CST'],
                             [10004, 'Special cables', None, None, 'x', 'ASSY'],
                             [10005, 'Cutting sample', 'X', None, None, 'CST'],
                             [10006, 'Cutting tubes CST', None, None, None, 'CST'],
                             [10010, 'AUDI Cutting General', 'x', None, None, 'CST'],
                             [10020, 'Convolute tubing', None, 'x', None, 'SUBBY'],
                             [10021, 'Other tubing', None, 'x', None, 'SUBBY'],
                             [10022, 'Insert tube in subassembly', None, 'x', None, 'SUBBY'],
                             [10030, 'Stripping', None, 'x', None, 'SUBBY'],
                             [10031, 'Outer case stripping', None, 'x', None, 'SUBBY'],
                             [10032, 'Crimping only', None, 'x', None, 'SUBBY'],
                             [10033, 'Stripper- crimper', None, 'x', None, 'SUBBY'],
                             [10034, 'Stripper- crimper- seal', None, 'x', None, 'SUBBY'],
                             [10036, 'Busbar crimper', None, 'x', None, 'SUBBY'],
                             [10038, 'Dip soldering', None, 'x', None, 'SUBBY'],
                             [10039, 'Soldering iron', None, 'x', None, 'SUBBY'],
                             [10040, 'Restistance soldering', None, 'X', None, 'SUBBY'],
                             [10041, 'Induction soldering', None, 'x', None, 'SUBBY'],
                             [10042, 'Moulding', None, 'x', None, 'SUBBY'],
                             [10044, 'Hand- work station', None, 'x', None, 'SUBBY'],
                             [10045, 'Twisting machine', None, 'x', None, 'SUBBY'],
                             [10046, 'Raychem heat shrink shuttle', None, 'x', None, 'SUBBY'],
                             [10047, 'Raychem solder sleeve machine', None, 'x', None, 'SUBBY'],
                             [10048, 'Coiling', None, 'x', None, 'SUBBY'],
                             [10049, 'Pre- assembly fitting', None, 'x', None, 'SUBBY'],
                             [10050, 'General subassembly', None, 'x', None, 'SUBBY'],
                             [10051, 'FIT CARDBOARD ARGL', None, None, 'X', 'ASSY'],
                             [10052, 'ASSEMBLY PURCHASED', None, 'X', None, 'SUBBY'],
                             [10070, 'Center strip', 'x', None, None, 'CST'],
                             [10071, 'Ultrasonic splice', None, 'x', None, 'SUBBY'],
                             [10072, 'Crimp splice 2 ton', None, 'x', None, 'SUBBY'],
                             [10073, 'Crimp splice 4 ton', None, 'x', None, 'SUBBY'],
                             [10075, 'Taping machine Ondal', None, 'x', None, 'SUBBY'],
                             [10077, 'Raychem shrinking', None, 'x', None, 'SUBBY'],
                             [10079, 'Diode and resistor splicing', None, 'x', None, 'SUBBY'],
                             [11001, 'CV CUTTING GENERAL', 'x', None, None, 'CST'],
                             [11002, 'CV Wire cutting by hand', None, 'x', None, 'SUBBY'],
                             [11003, 'Daimler cutting general', 'x', None, None, 'CST'],
                             [11004, 'Daimler cutting coiling', 'x', None, None, 'CST'],
                             [11006, 'CV CUTTING BEIUS', 'x', None, None, 'CST'],
                             [11009, 'cut wires section more than 10 mm2', None, 'x', None, 'SUBBY'],
                             [11010, 'CV cutting special cables', 'x', None, None, 'CST'],
                             [11020, 'CV Tube cutting up to 6,5 mm', None, 'x', None, 'TUBES'],
                             [11021, 'CV Tube cutting more than  6,5 mm', None, 'x', None, 'TUBES'],
                             [11022, 'CV cutting Wellrohr tube', None, 'x', None, 'TUBES'],
                             [11023, 'half autom closed convolute tube', None, 'x', None, 'SUBBY'],
                             [11024, 'cut/open corrugate/convolute tube', None, 'x', None, 'SUBBY'],
                             [11029, 'cable cover jaket strip machine', None, 'x', None, 'SUBBY'],
                             [11030, 'CV Strip machine', None, 'x', None, 'SUBBY'],
                             [11031, 'CV Shrink machine', None, 'x', None, 'SUBBY'],
                             [11032, 'CV Ring terminal press', None, 'x', None, 'SUBBY'],
                             [11033, 'CV Crimping area', None, 'x', None, 'SUBBY'],
                             [11034, 'CV Ultrasonic welding', None, 'x', None, 'SUBBY'],
                             [11035, 'soldering 22077 contact by hand', None, 'x', None, 'SUBBY'],
                             [11036, 'soldering by hand', None, 'x', None, 'SUBBY'],
                             [11037, 'crimping contact radiostecker', None, 'x', None, 'SUBBY'],
                             [11038, 'conector moulding machine', None, 'x', None, 'SUBBY'],
                             [11039, 'tin bath', None, 'x', None, 'SUBBY'],
                             [11041, 'HS200 strip seal crimp machine', None, 'x', None, 'SUBBY'],
                             [11042, '#N/A', None, 'x', None, 'SUBBY'],
                             [11044, 'Hand-work station CV', None, 'x', None, 'SUBBY'],
                             [11045, 'CV Wire twisting', None, 'x', None, 'SUBBY'],
                             [11048, 'Coiling CV', None, 'x', None, 'SUBBY'],
                             [11049, 'Pre-assembly fitting CV', None, 'x', None, 'SUBBY'],
                             [11050, 'CV Pre-assembly within the crimping', None, 'x', None, 'SUBBY'],
                             [11051, 'CV Pre-assembly area', None, 'x', None, 'SUBBY'],
                             [11052, 'stick silver etiquette', None, 'x', None, 'SUBBY'],
                             [11055, 'checking area', None, 'x', None, 'SUBBY'],
                             [11060, 'CV Foaming area', None, None, 'x', 'ASSY'],
                             [20020, 'Rover 75', None, None, 'x', 'ASSY'], [20030, 'T5', None, None, 'x', 'ASSY'],
                             [20040, 'Bentley', None, None, 'x', 'ASSY'], [20050, 'AML', None, None, 'x', 'ASSY'],
                             [20060, 'Defender', None, None, 'x', 'ASSY'],
                             [20080, 'Power cable Assembly', None, None, 'x', 'ASSY'],
                             [20090, 'Land Rover', None, None, 'x', 'ASSY'],
                             [21003, 'CV Montage/Assembly DAF area', None, None, 'x', 'ASSY'],
                             [21004, 'CV Montage/Assembly MAN area', None, None, 'x', 'ASSY'],
                             [21005, 'CV Montage/Assembly Claas area', None, None, 'x', 'ASSY'],
                             [21006, 'CV Montage/Assembly Fendt area', None, None, 'x', 'ASSY'],
                             [21007, 'CV Montage/Assembly Volvo area', None, None, 'x', 'ASSY'],
                             [21008, 'CV Montage/Assembly Power Cab area', None, None, 'x', 'ASSY'],
                             [21009, 'CV Montage/Assembly Cummins area', None, None, 'x', 'ASSY'],
                             [21010, 'CV Montage/Assembly KAMAZ area', None, None, 'x', 'ASSY'],
                             [21011, 'CV Montage/Assembly PERKINS area', None, None, 'x', 'ASSY'],
                             [21012, 'Montage/Assembly BMW area', None, None, 'x', 'ASSY'],
                             [21013, 'CV Montage/Assembly DAIMLER area', None, None, 'x', 'ASSY'],
                             [21014, 'Montage/Assembly AUDI area', None, None, 'x', 'ASSY'],
                             [21015, 'CV Montage/Assembly MTU area', None, None, 'x', 'ASSY'],
                             [21016, 'CV Montage/Assembly Daf Beius', None, None, 'x', 'ASSY'],
                             [21017, 'CV Montage/Assembly Leyland Beius', None, None, 'x', 'ASSY'],
                             [21018, 'CV Montage/Assembly Daimler ATEGO', None, None, 'x', 'ASSY'],
                             [21021, 'CV Montage/Assembly Volvo Hybrid', None, None, 'x', 'ASSY'],
                             [21022, 'CV Montage/Assembly Cabin Daf Beius', None, None, 'x', 'ASSY'],
                             [33600, '#N/A', None, None, 'X', 'ASSY'],
                             [90000, 'Electrical Test', None, None, 'x', 'TEST'],
                             [90001, 'CV Test area', None, None, 'x', 'TEST'],
                             [99000, 'Packing', None, None, 'X', 'PACK'],
                             [99001, 'CV Packing area', None, None, 'X', 'PACK'],
                             [21024, 'Assembly Man KSK', None, None, 'x', 'ASSY'],
                             [21025, 'Assembly Scania', None, None, 'x', 'ASSY'],
                             [8044, 'CRIMPMASCHINE RADIOSTECKER', None, 'x', None, 'SUBBY'],
                             [754012, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754013, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754051, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754231, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754291, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [755901, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [755902, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754012, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754013, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754051, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754231, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754291, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [755901, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [755902, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754012, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754013, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754051, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754231, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754291, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [755901, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [755902, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754012, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754013, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754051, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754231, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [754291, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [755901, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [755902, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [41503, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [41510, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [41512, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [43010, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [43110, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [43130, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [43150, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [43190, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [43204, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [43210, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [43211, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [43242, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [43451, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [43460, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [43499, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [43799, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [49004, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [49005, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [49006, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [49008, 'Nicht gefunden!/Not found!', None, None, None, 'Not Found!'],
                             [10085, 'braid machine', None, None, 'x', 'ASSY'],
                             [10009, 'BY636 Cutting general', 'x', None, None, 'CST'],
                             [10090, 'BY Foaming area', None, None, 'x', 'ASSY'],
                             [10076, 'Taping machine Sumitomo', None, 'x', None, 'SUBBY'],
                             [10007, 'Cutting Scania', 'x', None, None, 'CST'],
                             [10043, '6 ton- loose piece', None, 'x', None, 'SUBBY'],
                             [10082, 'Insert silicon', None, None, 'x', 'ASSY'],
                             [21026, 'HR Assembly training resource', None, None, 'x', 'ASSY'],
                             [20010, 'Rover 25', None, None, 'x', 'ASSY'],
                             [10083, 'insert wire in housing', None, None, 'x', 'ASSY'],
                             [11007, 'CV CUTTING ORDER', 'x', None, None, 'CST'],
                             ["11007A", 'CV CUTTING ORDER', 'x', None, None, 'CST'],
                             [11005, 'CV CUTTING BEIUS Engine', 'x', None, None, 'CST'],
                             [10096, 'pre-assembly fit and crimp metal', None, 'x', None, 'SUBBY'],
                             [10044, 'Hand- work station', None, 'x', None, 'SUBBY'],
                             [11008, 'Man CUTTING BEIUS', 'x', None, None, 'CST'],
                             [11030, 'CV Strip machine', None, 'x', None, 'SUBBY'],
                             [11031, 'CV Shrink machine', None, 'x', None, 'SUBBY'],
                             [11033, 'CV Crimping area', None, 'x', None, 'SUBBY'],
                             [11034, 'CV Ultrasonic welding', None, 'x', None, 'SUBBY'],
                             [11045, 'CV Wire twisting', None, 'x', None, 'SUBBY'],
                             [11048, 'Coiling CV', None, 'x', None, 'SUBBY'],
                             [21004, 'CV Montage/Assembly MAN area', None, None, 'x', 'ASSY'],
                             [21029, 'Montage/Assembly MAN Beius', None, None, 'x', 'ASSY'],
                             [11020, 'CUTTING TUBES CST', 'x', None, None, 'CST'],
                             [11044, 'CV HAND- WORK STATION', None, None, 'x', 'ASSY'],
                             ['11001A', 'CV CUTTING GENERAL', None, None, 'X', 'CST'],
                             ['08066', 'VERDRILLMACHINE', None, None, 'X', 'SUBBY'],
                             ['08365', 'VM-MONTAGE NFZ LH', None, None, 'X', 'SUBBY'],
                             ['23U203', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U405', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['11005A', 'CV CUTTING BEIUS Engine', None, None, 'X', 'CST'],
                             ['10007A', 'Cutting Scania', None, None, 'X', 'CST'],
                             ['11006A', 'CV CUTTING BEIUS', None, None, 'X', 'CST'],
                             ['02305', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['023104', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['023105', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U102', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U103', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U104', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U205', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U306', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U602', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U605', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U606', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U701', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U702', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U801', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U803', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U804', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U904', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U201', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U204', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U501', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U918', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['02306', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U916', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U802', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['10019', 'insert grommet on wires', None, None, 'X', 'SUBBY'],
                             ['23U120', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['0754012', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['0754013', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['0754051', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['0754231', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['0754291', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['0755901', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['0755902', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['90002', 'splice water/electrical Test', None, None, 'X', 'SUBBY'],
                             ['08804', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['08805', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['08808', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23R104', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23R201', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23R204', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23R301', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23R501', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23R503', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23R601', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23R803', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23R904', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23R999', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['11002A', 'cutting MAN CICOR', None, None, 'X', 'CST'],
                             ['11008A', 'Man CUTTING BEIUS Alb', None, None, 'X', 'CST'],
                             ['23R103', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23R107', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['21035', 'Assembly CNH', None, None, 'X', 'SUBBY'],
                             ['23R704', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23R802', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U919', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U917', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['02307', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23R109', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['086617', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['086614', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['086616', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['086105', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['10024', 'Raychem heat shrink with additional', None, None, 'X', 'SUBBY'],
                             ['08398', 'PACKTISCH', None, None, 'X', 'SUBBY'],
                             ['08003', 'KOMAX OC/1S/2S', None, None, 'X', 'SUBBY'],
                             ['08048', 'VM-ANSCHLAGBEREICH', None, None, 'X', 'SUBBY'],
                             ['08360', 'MONTAGE NFZ LH', None, None, 'X', 'SUBBY'],
                             ['08395', 'PRUEFTISCH 1 NFZ', None, None, 'X', 'SUBBY'],
                             ['08002', 'KOMAX O,C/1S/2S/DRUCK', None, None, 'X', 'SUBBY'],
                             ['08029', 'HANDLOETPLATZ', None, None, 'X', 'SUBBY'],
                             ['08032', 'KABELMAT-SCHLAUCHSCHNEIDEMASCHINE', None, None, 'X', 'SUBBY'],
                             ['08043', 'ANSCHLAGBEREICH', None, None, 'X', 'SUBBY'],
                             ['08044', 'CRIMPMASCHINE RADIOSTECKER', None, None, 'X', 'SUBBY'],
                             ['08052', 'STOCKO-PRESSE', None, None, 'X', 'SUBBY'],
                             ['08030', 'HANDSCHNEIDEPLATZ SCHLAUCH', None, None, 'X', 'SUBBY'],
                             ['08370', 'SILBERETIKETT KLEBEN', None, None, 'X', 'SUBBY'],
                             ['08031', 'ARTOS-SCHLAUCHSCHNEIDEMASCHINE', None, None, 'X', 'SUBBY'],
                             ['08051', 'ZINNBAD', None, None, 'X', 'SUBBY'],
                             ['08352', 'STOCKO-PRESSE', None, None, 'X', 'SUBBY'],
                             ['08069', '8069', None, None, 'X', 'Not Found!'],
                             ['08010', 'KOMAX 42 TUELLE/DRUCK', None, None, 'X', 'SUBBY'],
                             ['08049', 'SCHRUMPFAUTOMAT', None, None, 'X', 'SUBBY'],
                             ['08004', 'KOMAX OHNE CRIMP ab 6,00 QMM', None, None, 'X', 'SUBBY'],
                             ['08046', 'SPRITZMASCHINE', None, None, 'X', 'SUBBY'],
                             ['08040', 'ABISOLIERBOX', None, None, 'X', 'SUBBY'],
                             ['08041', 'ABMANTELPLATZ', None, None, 'X', 'SUBBY'],
                             ['08001', 'HANDSCHNEIDEPLATZ LEITUNGEN', None, None, 'X', 'SUBBY'],
                             ['08021', 'HALBAUTOMAT 21916', None, None, 'X', 'SUBBY'],
                             ['08060', 'MONTAGE FENDT ALLGEMEIN', None, None, 'X', 'SUBBY'],
                             ['08068', 'FRAESPLATZ', None, None, 'X', 'SUBBY'],
                             ['08025', 'HALBAUTOMAT 22077', None, None, 'X', 'SUBBY'],
                             ['08047', 'SCHAUMANLAGE', None, None, 'X', 'SUBBY'],
                             ['08396', 'PRUEFTISCH 2 NFZ', None, None, 'X', 'SUBBY'],
                             ['08260', '8260', None, None, 'X', 'Not Found!'],
                             ['08295', '8295', None, None, 'X', 'Not Found!'],
                             ['23U402', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U601', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U603', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['23U200', 'Nicht gefunden!/Not found!', None, None, 'X', 'Not Found!'],
                             ['10080', 'FIT TAPE ON WIRE', None, None, 'X', 'SUBBY'],
                             ['10084', 'fit sleeve and crimp terminal', None, None, 'X', 'SUBBY']
                             ]

    file_load = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                           title="Incarcati fisierul TXT APFR din FORS:")
    file_load2 = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                           title="Incarcati fisierul EXCEL cu clasificarile Resource group-urilor:")
    pbargui = Tk()
    pbargui.title("APFR")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    timelabel = Label(pbargui, text="Time . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    array_clasificare = []
    if file_load2:
        wb = load_workbook(file_load2)
        ws = wb.active
        for row in ws['A']:
            if row.value is not None:
                array_clasificare.append([ws.cell(row=row.row, column=1).value, ws.cell(row=row.row, column=2).value,
                                          ws.cell(row=row.row, column=3).value, ws.cell(row=row.row, column=4).value,
                                          ws.cell(row=row.row, column=5).value, ws.cell(row=row.row, column=6).value])
        if array_clasificare[0] != ['code', 'details', 'CST', 'SUBBY', 'ASSY', 'Group']:
            array_clasificare = array_clasificare_std
    else:
        array_clasificare = array_clasificare_std

    try:
        with open(file_load, newline='') as csvfile:
            array_apfr = list(csv.reader(csvfile, delimiter='\t'))
    except FileNotFoundError:
        messagebox.showerror("No file", "Nu ati incarcat nimic")
        return
    array_delete = [
    ['\x0c+===================================================================================================================================+'],
    ['+===================================================================================================================================+'],
    ['|===================================================================================================================================|'],
    ['| product number    :    0000000000000     - 9999999999999                                                                          |'],
    ['|===================================================================================================================================|'],
    ['|external number               product number                      designation                                                      |'],
    ['|ressource group          designation ressource group                    total   unit                                               |'],
    ['|===================================================================================================================================|'],
    ['| product number    :    0                 - 9999999999999                                                                          |'],
    []
    ]
    array_output = [["Client PN", "Leoni PN", "Designation", "Resource group", "Resignation resource group", "Total", "Unit", "Group"]]
    array_curatat = [line for line in array_apfr if line not in array_delete]
    array_curatat2 = [line for line in array_curatat if "" not in line]
    array_curatat3 = [line for line in array_curatat2 if 'FAVLS' not in line[0]]
    array_breakers = []
    for i in range(len(array_curatat3)):
        if array_curatat3[i] == ['+-----------------------------------------------------------------------------------------------------------------------------------+']:
            array_breakers.append(i)
    array_breakers.insert(0,0)
    for i in range(0, len(array_breakers) - 1):
        statuslabel["text"] = "Harness numarul " + str(i) + " din " + str(len(array_breakers))
        pbar['value'] += 1
        pbargui.update_idletasks()
        array_output_temp = []
        if array_breakers[i] == 0:
            for x in range(array_breakers[i], array_breakers[i + 1]):
                array_output_temp.append(array_curatat3[x])
            clientpn = [s.strip() for s in array_output_temp[0][0].split("   ") if s][0]
            leonipn = [s.strip() for s in array_output_temp[0][0].split("   ") if s][1]
            designation = [s.strip() for s in array_output_temp[0][0].split("   ") if s][2]
            for lista in array_output_temp[1:]:
                resgroup = [s.strip() for s in lista[0].split("   ") if s][0]
                resign = [s.strip() for s in lista[0].split("   ") if s][1]
                total = [s.strip() for s in lista[0].split("   ") if s][2]
                unit = [s.strip() for s in lista[0].split("   ") if s][3]
                array_output.append([clientpn, leonipn, designation, resgroup, resign,
                                     float(total.replace(",", ".")), unit])
        else:
            for x in range(array_breakers[i] + 1, array_breakers[i + 1]):
                array_output_temp.append(array_curatat3[x])
            clientpn = [s.strip() for s in array_output_temp[0][0].split("   ") if s][0]
            leonipn = [s.strip() for s in array_output_temp[0][0].split("   ") if s][1]
            designation = [s.strip() for s in array_output_temp[0][0].split("   ") if s][2]
            for lista in array_output_temp[1:]:
                resgroup = [s.strip() for s in lista[0].split("   ") if s][0]
                resign = [s.strip() for s in lista[0].split("   ") if s][1]
                total = [s.strip() for s in lista[0].split("   ") if s][2]
                unit = [s.strip() for s in lista[0].split("   ") if s][3]
                array_output.append([clientpn, leonipn, designation, resgroup, resign,
                                     float(total.replace(",", ".")), unit])
    for i in range(len(array_output)):
        for code in array_clasificare:
            if array_output[i][3] == str(code[0]):
                array_output[i].append(code[5])
                break
    for i in range(len(array_output)):
        if len(array_output[i]) < 7:
            array_output[i].append("Not Found")
    statuslabel["text"] = "Pivotare informatii . . . "
    pbar['value'] += 1
    pbargui.update_idletasks()
    dataframe = pd.DataFrame(array_output)
    dataframe.columns = dataframe.iloc[0]
    dataframe = dataframe[1:]
    pivot_dataframe = dataframe.pivot_table(index="Leoni PN", columns="Group", values="Total",
                                            fill_value=0, aggfunc='sum')
    #pivot_dataframe = dataframe.pivot_table(index="Leoni PN", columns="Group", values="Total",
    #                                        fill_value=0, aggfunc='sum', )

    # Calculate the sum of "ASSY," "PACK," and "TEST" columns and assign it to a new column
    pivot_dataframe = pivot_dataframe.assign(
        Asembly_Total=pivot_dataframe["ASSY"] + pivot_dataframe["PACK"] + pivot_dataframe["TEST"])
    pivot_dataframe = pivot_dataframe.assign(
        Total=pivot_dataframe["ASSY"] + pivot_dataframe["PACK"] + pivot_dataframe["TEST"] + pivot_dataframe["CST"] +
              pivot_dataframe["SUBBY"] + pivot_dataframe["TUBES"]  + pivot_dataframe["Not Found!"])
    statuslabel["text"] = "Salvare informatii in EXCEL"
    pbar['value'] += 1
    pbargui.update_idletasks()
    pivot_dataframe.to_excel(save_file, sheet_name="APFR", index=True)
    msg_box = messagebox.askquestion('Salvare detalii APFR', 'Doriti salvarea detaliilor?',
                                        icon='warning')
    if msg_box == 'yes':
        with pd.ExcelWriter(save_file, engine='openpyxl', mode='a') as writer:
            # Write the DataFrame to the Excel file (you can specify the sheet name)
            dataframe.to_excel(writer, sheet_name='APFR detalii', index=False)
    messagebox.showinfo('Finalizat', "!!!!!!!!!!!!!!")
    pbargui.destroy()


def breaklargefiles():
    output_excel_path = os.path.abspath(os.curdir) + "/MAN/Output/"
    input_file_path = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                           title="Incarcati fisierul TXT:")
    pbargui = Tk()
    pbargui.title("Big files break up")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    timelabel = Label(pbargui, text="Time . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    timelabel.grid(row=2, column=2)
    chunksize = 500000

    # Open the input file in read mode
    with open(input_file_path, 'r', encoding='latin-1') as file:
        # Read the content of the file
        content = file.read()

    # Use regular expression to replace NUL values with the desired replacement string
    content_without_null = re.sub(r'\x00', '', content)
    content_without_null2 = re.sub(r'\x82', '', content_without_null)
    content_without_null3 = re.sub(r'\x90', '', content_without_null2)
    content_without_null4 = re.sub(r'\x83', '', content_without_null3)
    content_without_null5 = re.sub(r'\x89', '', content_without_null4)
    content_without_null6 = re.sub(r'\x88', '', content_without_null5)
    lines = content_without_null6.splitlines()
    # Create a list of dictionaries where each dictionary represents a row in the Excel file
    data = [{'Content': line} for line in lines]
    df = pd.DataFrame(data)
    # Calculate the number of chunks
    num_chunks = len(df) // chunksize + 1
    # Save DataFrame in chunks to text files
    for i in range(num_chunks):
        start_idx = i * chunksize
        end_idx = (i + 1) * chunksize
        chunk_df = df.iloc[start_idx:end_idx]
        chunk_df.to_csv(output_excel_path + f'output_{i}.txt', sep='\t', index=False)  # Tab-separated text file
    messagebox.showinfo('Finalizat', "!!!!!!!!!!!!!!")
    pbargui.destroy()
