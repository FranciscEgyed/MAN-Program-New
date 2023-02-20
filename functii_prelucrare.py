import csv
import os
import time
from tkinter import messagebox, filedialog, Tk, ttk, HORIZONTAL, Label
from typing import Any
from openpyxl import load_workbook
from diverse import log_file, error_file
from functii_print import prn_excel_separare_ksk, prn_excel_wires_complete_leoni, prn_excel_ksk_neprelucrate, \
    prn_excel_wires_complete, prn_excel_bom_complete, prn_excel_boomcumulat, prn_excel_wirelistsallinone
import sqlite3


def sortare_jit():
    pbargui = Tk()
    pbargui.title("Sortare JIT")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    timelabel = Label(pbargui, text="Time . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2)
    timelabel.grid(row=2, column=2)
    c8000 = 0
    c8011 = 0
    c8023 = 0
    cnec = 0
    tip = ""
    # Import data files
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Sortare Module.txt", newline='') as csvfile:
        array_sortare_module = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/KSKLight.txt", newline='') as csvfile:
        array_sortare_light = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Module Active.txt", newline='') as csvfile:
        array_module_active = list(csv.reader(csvfile, delimiter=';'))
    normal = ["04.37161-9100", "81.25484-5259", "81.25484-5263", "81.25484-5260", "81.25484-5264", "81.25484-5273",
              "81.25484-5272", "81.25484-5267", "81.25484-5268"]
    ADR = ["04.37161-9000", "81.25484-5261", "81.25484-5265", "81.25484-5262", "81.25484-5266", "81.25484-5275",
           "81.25484-5274", "81.25484-5271", "81.25484-5270"]
    # Open JIT file
    fisier_calloff = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                                title="Incarcati fisierul care necesita sortare")
    if len(fisier_calloff) == 0:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showinfo("Nu ati selectat nimic")
        return None
    statuslabel["text"] = "Incarcare fisier excel"
    timelabel["text"] = "0.15 secunde / KSK"
    pbar['value'] += 2
    pbargui.update_idletasks()
    data_download = os.path.basename(fisier_calloff)[11:21]
    try:
        wb = load_workbook(fisier_calloff)
    except:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showinfo("Fisier invalid", fisier_calloff + " extensie incompatibila!")
        return None
    try:
        conn = sqlite3.connect("//SVRO8FILE01/Groups/General/EFI/DBMAN/database.db")
    except sqlite3.OperationalError:
        conn = sqlite3.connect(os.path.abspath(os.curdir) + "/MAN/Input/Others/database.db")
    statuslabel["text"] = "Sortare fisier JIT"
    timelabel["text"] = "0.15 secunde / KSK"
    pbar['value'] += 2
    pbargui.update_idletasks()
    ws = wb.worksheets[0]
    lista_calloff_total = []
    is_error = False
    for row in ws['A']:
        if row.value is not None and row.value != "ext.Abrufnummer" and row.value != "PRODN" and \
                row.value != "Ext.JIT Call No" and (len(ws.cell(row=row.row, column=2).value) == 13 or
                                                    len(ws.cell(row=row.row, column=2).value) == 14):
            lista_calloff_total.append(row.value)
        elif row.value is not None and row.value != "ext.Abrufnummer" and row.value != "PRODN" and \
                row.value != "Ext.JIT Call No":
            is_error = True
            error_file("Eroare in " + os.path.basename(fisier_calloff) + " pe randul " + str(row.row) + " valoarea " +
                       ws.cell(row=row.row, column=2).value + " nu este corecta")

    lista_calloff_unice = list(dict.fromkeys(lista_calloff_total))
    file_counter = len(lista_calloff_unice)
    file_progres = 0
    start = time.time()
    # Extract each KSK from JIT file
    for element in lista_calloff_unice:
        is_light = "NO"
        array_temporar: list[list[str | Any]] = []
        array_temporar_module = []
        array_database = []
        for row in ws['A']:
            if row.value == element:
                if len(str(ws.cell(row=row.row, column=4).value)) > 4:
                    qty = "0"
                elif ws.cell(row=row.row, column=4).value == 1000:
                    qty = "1"
                elif ws.cell(row=row.row, column=4).value == 2000:
                    qty = "2"
                elif ws.cell(row=row.row, column=4).value == 3000:
                    qty = "3"
                else:
                    qty = ws.cell(row=row.row, column=4).value
                array_temporar.append(
                    [ws.cell(row=row.row, column=1).value[1:], ws.cell(row=row.row, column=2).value,
                     ws.cell(row=row.row, column=3).value, qty, ws.cell(row=row.row, column=5).value,
                     ws.cell(row=row.row, column=12).value, ws.cell(row=row.row, column=13).value,
                     ws.cell(row=row.row, column=16).value, ws.cell(row=row.row, column=14).value])
                array_temporar_module.append(
                    ws.cell(row=row.row, column=2).value.replace('PM.', '81.').replace('VM.', '81.'))
        for i in range(len(array_temporar)):
            if "PM." in array_temporar[i][1]:
                array_temporar[i][1] = array_temporar[i][1].replace('PM.', '81.')
        for i in range(len(array_temporar)):
            if "VM." in array_temporar[i][1]:
                array_temporar[i][1] = array_temporar[i][1].replace('VM.', '81.')

        for item in array_temporar:
            if item[1] in array_sortare_module[0]:
                tip = "8000"
                c8000 = c8000 + 1
                break
            elif item[1] in array_sortare_module[1]:
                tip = "8011"
                c8011 = c8011 + 1
                break
            elif item[1] in array_sortare_module[2]:
                tip = "8023"
                c8023 = c8023 + 1
                break
            else:
                tip = "Necunoscut"
        if tip == "Necunoscut":
            cnec = cnec + 1
        # Chech KSK if part of KSK Light project
        if set(array_temporar_module).issubset(array_sortare_light[0]):
            prn_excel_separare_ksk(array_temporar_module, element[1:])
            is_light = "YES"
        # Check harness type
        harnesstype = []
        for m in range(len(array_temporar_module)):
            for n in range(len(array_module_active)):
                if array_temporar_module[m] == array_module_active[n][0] and array_module_active[n][3] != "XXXX":
                    harnesstype.append(array_module_active[n][3].replace(' LHD', '').replace(' RHD', ''))
        harnesstype = list(set(harnesstype))
        # Check ss type
        sstype = "None"
        for x in range(len(array_temporar_module)):
            if array_temporar_module[x] in normal:
                sstype = "NON ADR"
                break
            elif array_temporar_module[x] in ADR:
                sstype = "ADR"
                break
        # Write to database
        primarykey = os.path.basename(fisier_calloff) + element[1:]
        array_database.append([primarykey, os.path.basename(fisier_calloff), ';'.join(harnesstype), is_light,
                               array_temporar[1][8], data_download, element[1:], array_temporar[1][7], sstype,
                               ';'.join(array_temporar_module)])
        conn = sqlite3.connect(os.path.abspath(os.curdir) + "/MAN/Input/Others/database.db")
        cursor = conn.cursor()
        # create a table
        cursor.execute("""CREATE TABLE IF NOT EXISTS KSKDatabase
                          (primarykey text UNIQUE, numejit text, TipHarness text, Light text, DataLivrare text, 
                          DataJIT text, KSKNo text, TrailerNO text, SSType text, Module text) """)
        # insert multiple records using the more secure "?" method
        cursor.executemany("INSERT OR IGNORE INTO KSKDatabase VALUES (?,?,?,?,?,?,?,?,?,?)", array_database)
        conn.commit()
        conn.close()

        # Write to disk
        with open(os.path.abspath(os.curdir) + "/MAN/Input/Module Files/" + tip + "/"
                  + element[1:] + ".csv", 'w', newline='') as myfile:
            wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
            wr.writerow(['Harness', 'Module', 'Side', 'Quantity', tip, "Date", "Time", "Trailer No"])
            wr.writerows(array_temporar)
        del array_temporar
        end = time.time()

        file_progres = file_progres + 1
        statuslabel["text"] = str(file_progres) + "/" + str(file_counter) + "                "
        timelabel["text"] = "Estimated time to complete : " + \
                            str(((file_counter * 0.15) - (end - start)) / 60)[:5] + " minutes."
        pbar['value'] += 2
        pbargui.update_idletasks()
    conn.close()
    pbar.destroy()
    pbargui.destroy()
    log_file("Sortate  8000 = " + str(c8000) + ", 8011 = " + str(c8011) + ", 8023 = " + str(c8023) +
             "Necunoscute = " + str(cnec))
    messagebox.showinfo("Finalizat", "Sortate  8000 = " + str(c8000) + ", 8011 = " + str(c8011) + ", 8023 = " +
                        str(c8023) + ", Necunoscute = " + str(cnec))
    if is_error:
        messagebox.showinfo("Erori in fisiere", "Verificati fisierul === Error file.txt === !")
        os.startfile(os.path.abspath(os.curdir) + "/MAN/Error file.txt")
    return None


def sortare_jit_dir():
    pbargui = Tk()
    pbargui.title("Sortare JIT din director")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    statuslabel2 = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2)
    statuslabel2.grid(row=2, column=2)
    # Import data files
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Sortare Module.txt", newline='') as csvfile:
        array_sortare_module = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/KSKLight.txt", newline='') as csvfile:
        array_sortare_light = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Module Active.txt", newline='') as csvfile:
        array_module_active = list(csv.reader(csvfile, delimiter=';'))
    normal = ["04.37161-9100", "81.25484-5259", "81.25484-5263", "81.25484-5260", "81.25484-5264", "81.25484-5273",
              "81.25484-5272", "81.25484-5267", "81.25484-5268"]
    ADR = ["04.37161-9000", "81.25484-5261", "81.25484-5265", "81.25484-5262", "81.25484-5266", "81.25484-5275",
           "81.25484-5274", "81.25484-5271", "81.25484-5270"]
    dir_Jit = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                      title="Selectati directorul cu fisiere JIT:")
    try:
        conn = sqlite3.connect("//SVRO8FILE01/Groups/General/EFI/DBMAN/database.db")
    except sqlite3.OperationalError:
        conn = sqlite3.connect(os.path.abspath(os.curdir) + "/MAN/Input/Others/database.db")
    cursor = conn.cursor()
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_Jit):
        if file_all.endswith(".xlsx"):
            file_counter = file_counter + 1
    if file_counter == 0:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showinfo("Fisier invalid", "Nu am gasit fisiere de prelucrat!")
        return None
    for file_all in os.listdir(dir_Jit):
        if file_all.endswith(".xlsx") and file_all.startswith("JIT"):
            c8000 = 0
            c8011 = 0
            c8023 = 0
            cnec = 0
            tip = ""
            data_download = os.path.basename(file_all)[11:21]
            fisier_calloff = os.path.join(dir_Jit, file_all)
            try:
                wb = load_workbook(fisier_calloff)
            except:
                pbar.destroy()
                pbargui.destroy()
                messagebox.showinfo("Fisier invalid", fisier_calloff + " extensie incompatibila!")
                return None

            ws = wb.worksheets[0]
            lista_calloff_total = []
            for row in ws['A']:
                if row.value is not None and row.value != "ext.Abrufnummer" and row.value != "PRODN" and \
                        row.value != "Ext.JIT Call No" and (len(ws.cell(row=row.row, column=2).value) == 13 or
                                                            len(ws.cell(row=row.row, column=2).value) == 14):
                    lista_calloff_total.append(row.value)
            lista_calloff_unice = list(dict.fromkeys(lista_calloff_total))
            kskcounter = len(lista_calloff_unice)
            kskprogres = 0
            start = time.time()
            # Extract each KSK from JIT file
            for element in lista_calloff_unice:
                is_light = "NO"
                array_database = []
                kskprogres = kskprogres + 1
                statuslabel2["text"] = "                 " + str(kskprogres) + " / " + str(
                    kskcounter) + " : " + element[1:]
                pbar['value'] += 2
                pbargui.update_idletasks()
                array_temporar = []
                array_temporar_module = []
                for row in ws['A']:
                    if row.value == element:
                        if len(str(ws.cell(row=row.row, column=4).value)) > 4:
                            qty = "0"
                        elif ws.cell(row=row.row, column=4).value == 1000:
                            qty = "1"
                        elif ws.cell(row=row.row, column=4).value == 2000:
                            qty = "2"
                        elif ws.cell(row=row.row, column=4).value == 3000:
                            qty = "3"
                        else:
                            qty = ws.cell(row=row.row, column=4).value
                        array_temporar.append(
                            [ws.cell(row=row.row, column=1).value[1:], ws.cell(row=row.row, column=2).value,
                             ws.cell(row=row.row, column=3).value, qty, ws.cell(row=row.row, column=5).value,
                             ws.cell(row=row.row, column=12).value, ws.cell(row=row.row, column=13).value,
                             ws.cell(row=row.row, column=16).value, ws.cell(row=row.row, column=14).value])
                        array_temporar_module.append(
                            ws.cell(row=row.row, column=2).value.replace('PM.', '81.').replace('VM.', '81.'))
                for i in range(len(array_temporar)):
                    if "PM." in array_temporar[i][1]:
                        array_temporar[i][1] = array_temporar[i][1].replace('PM.', '81.')
                for i in range(len(array_temporar)):
                    if "VM." in array_temporar[i][1]:
                        array_temporar[i][1] = array_temporar[i][1].replace('VM.', '81.')

                for item in array_temporar:
                    if item[1] in array_sortare_module[0]:
                        tip = "8000"
                        c8000 = c8000 + 1
                        break
                    elif item[1] in array_sortare_module[1]:
                        tip = "8011"
                        c8011 = c8011 + 1
                        break
                    elif item[1] in array_sortare_module[2]:
                        tip = "8023"
                        c8023 = c8023 + 1
                        break
                    else:
                        tip = "Necunoscut"
                if tip == "Necunoscut":
                    cnec = cnec + 1
                # Chech KSK if part of KSK Light project
                if set(array_temporar_module).issubset(array_sortare_light[0]):
                    prn_excel_separare_ksk(array_temporar_module, element[1:])
                    is_light = "YES"
                else:
                    prn_excel_ksk_neprelucrate(array_temporar_module, element[1:])
                # Check harness type
                harnesstype = []
                for m in range(len(array_temporar_module)):
                    for n in range(len(array_module_active)):
                        if array_temporar_module[m] == array_module_active[n][0] and \
                                array_module_active[n][3] != "XXXX":
                            harnesstype.append(array_module_active[n][3].replace(' LHD', '').replace(' RHD', ''))
                harnesstype = list(set(harnesstype))
                # Check ss type
                sstype = "None"
                for x in range(len(array_temporar_module)):
                    if array_temporar_module[x] in normal:
                        sstype = "NON ADR"
                        break
                    elif array_temporar_module[x] in ADR:
                        sstype = "ADR"
                        break
                # Write to database
                primarykey = os.path.basename(fisier_calloff) + element[1:]
                array_database.append([primarykey, os.path.basename(fisier_calloff), ';'.join(harnesstype), is_light,
                                       array_temporar[1][8], data_download, element[1:], array_temporar[1][7], sstype,
                                       ';'.join(array_temporar_module)])

                # create a table
                cursor.execute("""CREATE TABLE IF NOT EXISTS KSKDatabase
                                  (primarykey text UNIQUE, numejit text, TipHarness text, Light text, DataLivrare text, 
                                  DataJIT text, KSKNo text, TrailerNO text, SSType text, Module text) """)
                # insert multiple records using the more secure "?" method
                cursor.executemany("INSERT OR IGNORE INTO KSKDatabase VALUES (?,?,?,?,?,?,?,?,?,?)", array_database)
                conn.commit()
                with open(os.path.abspath(os.curdir) + "/MAN/Input/Module Files/" + tip + "/"
                          + element[1:] + ".csv", 'w', newline='') as myfile:
                    wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
                    wr.writerow(['Harness', 'Module', 'Side', 'Quantity', tip, "Date", "Time", "Trailer No"])
                    wr.writerows(array_temporar)
                del array_temporar
                end = time.time()

            file_progres = file_progres + 1
            statuslabel["text"] = str(file_progres) + " / " + str(file_counter) + " : " + file_all
            pbar['value'] += 2
            pbargui.update_idletasks()
    conn.close()
    pbar.destroy()
    pbargui.destroy()
    messagebox.showinfo('Finalizat!', str(file_progres) + " fisiere din " + str(file_counter))


def wires():
    pbargui = Tk()
    pbargui.title("Prelucrare WIRELIST-uri")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    array_wirelisturi = ["8000", "8001", "8011", "8012", "8013", "8014", "8023", "8024", "8025", "8026", "8027",
                         "8030", "8031", "8032", "8052", "8053"]

    dir_wirelist = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                           title="Selectati directorul cu fisiere:")
    start = time.time()
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_wirelist):
        if file_all.endswith(".csv"):
            file_counter = file_counter + 1
    for file_all in os.listdir(dir_wirelist):
        if file_all.endswith(".csv"):
            file_progres = file_progres + 1
            statuslabel["text"] = str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 2
            pbargui.update_idletasks()
            for item in array_wirelisturi:
                if item in file_all:
                    try:
                        array_original = []
                        with open(dir_wirelist + "/" + file_all, newline='') as csvfile:
                            array_incarcat = list(csv.reader(csvfile, delimiter=';'))
                        if "AEM" in array_incarcat[0][1]:
                            for i in range(len(array_incarcat)):
                                del array_incarcat[i][1]
                        for x in range(len(array_incarcat)):
                            array_t = []
                            for i in range(len(array_incarcat[x])):
                                if i < 27 and array_incarcat[x][0] == "3":
                                    if array_incarcat[x][i] == "":
                                        array_t.append("X")
                                    else:
                                        array_t.append(array_incarcat[x][i])
                                elif array_incarcat[x][i] != "":
                                    array_t.append(array_incarcat[x][i])
                            array_original.append(array_t)
                        if array_original[0][0] != 1 and len(array_original[0]) != 2:
                            messagebox.showerror('Eroare fisier' + file_all, 'Nu ai incarcat fisierul corect ...')
                            return
                        if len(array_original) < 50:
                            for i in range(0, len(array_original)):
                                if array_original[i][0] == "0":
                                    if str(array_original[i][1]) != "Ltg-Nr.":
                                        messagebox.showerror('Eroare fisier' + file_all,
                                                             'Nu ai incarcat fisierul corect')
                                        return
                        else:
                            for i in range(0, 50):
                                if array_original[i][0] == "0":
                                    if str(array_original[i][1]) != "Ltg-Nr.":
                                        messagebox.showerror('Eroare fisier' + file_all,
                                                             'Nu ai incarcat fisierul corect')
                                        return
                    except:
                        messagebox.showerror('Eroare fisier' + file_all, 'Nu ai incarcat fisierul corect')
                        return
                    nume_fisier = os.path.splitext(os.path.basename(file_all))[0]
                    array_output = [
                        ["Module", "Ltg No", "Leitung", "Farbe", "Quer.", "Kurzname", "Pin", "Kurzname", "Pin",
                         "Sonderltg.",
                         "Lange"]]
                    array_module = []
                    array_wires = []
                    for i in range(1, len(array_incarcat)):
                        if i == len(array_incarcat) - 1:
                            array_wires.append(array_incarcat[i])
                            array_out_temp = []
                            for verwires in range(len(array_wires[0])):
                                for vermodule in range(len(array_module)):
                                    if array_wires[0][verwires] == array_module[vermodule][1] or \
                                            array_wires[0][verwires] == "Length  " + array_module[vermodule][1]:
                                        array_wires[0][verwires] = array_module[vermodule][2]
                            pot_position = array_wires[0].index('Pot.') + 1
                            ltgno_position = array_wires[0][pot_position:].index('Ltg-Nr.') + pot_position
                            index_ltgno = array_wires[0].index('Ltg-Nr.')
                            index_leitung = array_wires[0].index('Leitung')
                            index_farbe = array_wires[0].index('Farbe')
                            index_quer = array_wires[0].index('Quer.')
                            index_kurz1 = array_wires[0].index('Kurzname')
                            index_pin1 = array_wires[0].index('Pin')
                            index_kurz2 = array_wires[0][index_kurz1 + 1:].index('Kurzname') + index_kurz1 + 1
                            index_pin2 = array_wires[0][index_pin1 + 1:].index('Pin') + index_pin1 + 1
                            index_sonder = array_wires[0].index('Sonderltg.')
                            for wire in range(1, len(array_wires)):
                                for index in range(pot_position, ltgno_position):
                                    if array_wires[wire][index] != "-":
                                        array_out_temp.append([array_wires[0][index], array_wires[wire][index_ltgno],
                                                               array_wires[wire][index_leitung],
                                                               array_wires[wire][index_farbe],
                                                               float(array_wires[wire][index_quer].replace(',', '.')),
                                                               array_wires[wire][index_kurz1],
                                                               array_wires[wire][index_pin1],
                                                               array_wires[wire][index_kurz2],
                                                               array_wires[wire][index_pin2],
                                                               array_wires[wire][index_sonder],
                                                               array_wires[wire][index]])
                            array_module = []
                            array_wires = []
                            array_output.extend(array_out_temp)
                        else:
                            if array_incarcat[i][0] == "2":
                                array_module.append(array_incarcat[i])
                            elif array_incarcat[i][0] == "3" or array_incarcat[i][0] == "0":
                                array_wires.append(array_incarcat[i])
                            elif array_incarcat[i][0] == "1":
                                for verwires in range(len(array_wires[0])):
                                    for vermodule in range(len(array_module)):
                                        if array_wires[0][verwires] == array_module[vermodule][1] or \
                                                array_wires[0][verwires] == "Length  " + array_module[vermodule][1]:
                                            array_wires[0][verwires] = array_module[vermodule][2]
                                array_out_temp = []
                                index_ltgno = array_wires[0].index('Ltg-Nr.')
                                index_leitung = array_wires[0].index('Leitung')
                                index_farbe = array_wires[0].index('Farbe')
                                index_quer = array_wires[0].index('Quer.')
                                index_kurz1 = array_wires[0].index('Kurzname')
                                index_pin1 = array_wires[0].index('Pin')
                                index_kurz2 = array_wires[0][index_kurz1 + 1:].index('Kurzname') + index_kurz1 + 1
                                index_pin2 = array_wires[0][index_pin1 + 1:].index('Pin') + index_pin1 + 1
                                index_sonder = array_wires[0].index('Sonderltg.')
                                pot_position = array_wires[0].index('Pot.') + 1
                                ltgno_position = array_wires[0][pot_position:].index('Ltg-Nr.') + pot_position
                                for wire in range(1, len(array_wires)):
                                    for index in range(pot_position, ltgno_position):
                                        if array_wires[wire][index] != "-":
                                            array_out_temp.append(
                                                [array_wires[0][index], array_wires[wire][index_ltgno],
                                                 array_wires[wire][index_leitung], array_wires[wire][index_farbe],
                                                 float(array_wires[wire][index_quer].replace(',', '.')),
                                                 array_wires[wire][index_kurz1], array_wires[wire][index_pin1],
                                                 array_wires[wire][index_kurz2], array_wires[wire][index_pin2],
                                                 array_wires[wire][index_sonder],
                                                 array_wires[wire][index]])
                                array_module = []
                                array_wires = []
                                array_output.extend(array_out_temp)
                    with open(os.path.abspath(os.curdir) + "/MAN/Input/Wire Lists/" + nume_fisier + ".csv", 'w',
                              newline='') as myfile:
                        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
                        wr.writerows(array_output)
                        log_file("Creat " + nume_fisier)

    end = time.time()
    pbar.destroy()
    pbargui.destroy()
    messagebox.showinfo('Finalizat!', "Prelucrate in " + str(end - start)[:6] + " secunde.")
    return None


def boms():
    pbargui = Tk()
    pbargui.title("Prelucrare BOM-uri")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    global nume_fisier
    array_boms = ["8000", "8001", "8011", "8012", "8013", "8014", "8023", "8024", "8025", "8026", "8027",
                  "8030", "8031", "8032", "8052", "8053"]
    dir_BOM = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir), title="Selectati directorul cu fisiere:")
    start = time.time()
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_BOM):
        if file_all.endswith(".csv"):
            file_counter = file_counter + 1
    for file_all in os.listdir(dir_BOM):
        if file_all.endswith(".csv"):
            file_progres = file_progres + 1
            statuslabel["text"] = str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 2
            pbargui.update_idletasks()
            for item in array_boms:
                if item in file_all:
                    try:
                        with open(dir_BOM + "/" + file_all, newline='') as csvfile:
                            array_incarcat = list(csv.reader(csvfile, delimiter=';'))
                        if "AEM" in array_incarcat[0][1]:
                            for i in range(len(array_incarcat)):
                                del array_incarcat[i][1]
                        if array_incarcat[0][0] != "1" and len(array_incarcat[0]) != 2:
                            messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect')
                            return
                        if len(array_incarcat) < 50:
                            for i in range(0, len(array_incarcat)):
                                if array_incarcat[i][0] == "0":
                                    if str(array_incarcat[i][-1])[:1] != "V":
                                        messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect')
                                        return
                        else:
                            for i in range(0, 50):
                                if array_incarcat[i][0] == "0":
                                    if str(array_incarcat[i][-1])[:1] != "V":
                                        messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect')
                                        return
                    except:
                        messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect')
                        return
                    nume_fisier = os.path.splitext(os.path.basename(file_all))[0]
                    array_output = [
                        ["Module", "Quantity", "Bezeichnung", "VOBES-ID", "Benennung", "Verwendung", "Verwendung",
                         "Kurzname", "xy", "Teilenummer", "Vorzugsteil", "TAB-Nummer", "Referenzteil", "Farbe",
                         "E-Komponente", "E-Komponente Part-Nr.", "Einh."]]
                    array_module = []
                    array_comp = []
                    for i in range(1, len(array_incarcat)):
                        if i == len(array_incarcat) - 1:
                            array_comp.append(array_incarcat[i])
                            array_out_temp = []
                            for vercomp in range(len(array_comp[0])):
                                for vermodule in range(len(array_module)):
                                    if array_module[vermodule][1] in array_comp[0][vercomp]:
                                        array_comp[0][vercomp] = array_module[vermodule][2]
                            last_position = len(array_comp[0])
                            index_beze = array_comp[0].index('Bezeichnung')
                            index_VID = array_comp[0].index('VOBES-ID')
                            try:
                                index_bene = array_comp[0].index('Benennung')
                            except ValueError:
                                index_bene = index_VID
                            index_verew1 = array_comp[0].index('Verwendung')
                            index_verew2 = array_comp[0][index_verew1:].index('Verwendung') + index_verew1
                            index_kurz = array_comp[0].index('Kurzname')
                            index_xy = array_comp[0].index('xy')
                            index_teile = array_comp[0].index('Teilenummer')
                            try:
                                index_vorzug = array_comp[0].index('Vorzugsteil')
                            except ValueError:
                                index_vorzug = index_teile
                            index_tab = array_comp[0].index('TAB-Nummer')
                            index_refe = array_comp[0].index('Referenzteil')
                            index_farbe = array_comp[0].index('Farbe')
                            try:
                                index_ekomp = array_comp[0].index('E-Komponente')
                            except ValueError:
                                index_ekomp = index_VID
                            try:
                                index_ekomppn = array_comp[0].index('E-Komponente Part-Nr.')
                            except ValueError:
                                index_ekomppn = index_VID
                            index_einh = array_comp[0].index('Einh.')
                            for comp in range(1, len(array_comp)):
                                for index in range(index_einh + 1, last_position):
                                    if array_comp[comp][index] != "0" and array_comp[comp][index] != "-":
                                        array_out_temp.append([array_comp[0][index], array_comp[comp][index],
                                                               array_comp[comp][index_beze],
                                                               array_comp[comp][index_VID],
                                                               array_comp[comp][index_bene],
                                                               array_comp[comp][index_verew1],
                                                               array_comp[comp][index_verew2],
                                                               array_comp[comp][index_kurz],
                                                               array_comp[comp][index_xy],
                                                               array_comp[comp][index_teile],
                                                               array_comp[comp][index_vorzug],
                                                               array_comp[comp][index_tab],
                                                               array_comp[comp][index_refe],
                                                               array_comp[comp][index_farbe],
                                                               array_comp[comp][index_ekomp],
                                                               array_comp[comp][index_ekomppn],
                                                               array_comp[comp][index_einh]])
                            array_module = []
                            array_comp = []
                            array_output.extend(array_out_temp)
                        else:
                            if array_incarcat[i][0] == "2":
                                array_module.append(array_incarcat[i])
                            elif array_incarcat[i][0] == "3" or array_incarcat[i][0] == "4" or \
                                    array_incarcat[i][0] == "0":
                                array_comp.append(array_incarcat[i])
                            elif array_incarcat[i][0] == "1":
                                for vercomp in range(len(array_comp[0])):
                                    for vermodule in range(len(array_module)):
                                        if array_module[vermodule][1] in array_comp[0][vercomp]:
                                            array_comp[0][vercomp] = array_module[vermodule][2]
                                array_out_temp = []
                                last_position = len(array_comp[0])
                                index_beze = array_comp[0].index('Bezeichnung')
                                index_VID = array_comp[0].index('VOBES-ID')
                                try:
                                    index_bene = array_comp[0].index('Benennung')
                                except ValueError:
                                    index_bene = index_VID
                                index_verew1 = array_comp[0].index('Verwendung')
                                index_verew2 = array_comp[0][index_verew1:].index('Verwendung') + index_verew1
                                index_kurz = array_comp[0].index('Kurzname')
                                index_xy = array_comp[0].index('xy')
                                index_teile = array_comp[0].index('Teilenummer')
                                try:
                                    index_vorzug = array_comp[0].index('Vorzugsteil')
                                except ValueError:
                                    index_vorzug = index_teile
                                index_tab = array_comp[0].index('TAB-Nummer')
                                index_refe = array_comp[0].index('Referenzteil')
                                index_farbe = array_comp[0].index('Farbe')
                                try:
                                    index_ekomp = array_comp[0].index('E-Komponente')
                                except ValueError:
                                    index_ekomp = index_VID
                                try:
                                    index_ekomppn = array_comp[0].index('E-Komponente Part-Nr.')
                                except ValueError:
                                    index_ekomppn = index_VID
                                index_einh = array_comp[0].index('Einh.')

                                for comp in range(1, len(array_comp)):
                                    for index in range(index_einh + 1, last_position):
                                        if array_comp[comp][index] != "0" and array_comp[comp][index] != "-":
                                            array_out_temp.append([array_comp[0][index], array_comp[comp][index],
                                                                   array_comp[comp][index_beze],
                                                                   array_comp[comp][index_VID],
                                                                   array_comp[comp][index_bene],
                                                                   array_comp[comp][index_verew1],
                                                                   array_comp[comp][index_verew2],
                                                                   array_comp[comp][index_kurz],
                                                                   array_comp[comp][index_xy],
                                                                   array_comp[comp][index_teile],
                                                                   array_comp[comp][index_vorzug],
                                                                   array_comp[comp][index_tab],
                                                                   array_comp[comp][index_refe],
                                                                   array_comp[comp][index_farbe],
                                                                   array_comp[comp][index_ekomp],
                                                                   array_comp[comp][index_ekomppn],
                                                                   array_comp[comp][index_einh]])
                                array_module = []
                                array_comp = []
                                array_output.extend(array_out_temp)
                    with open(os.path.abspath(os.curdir) + "/MAN/Input/BOMs/" + nume_fisier + ".csv", 'w',
                              newline='') as myfile:
                        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
                        wr.writerows(array_output)
    log_file("Creat " + nume_fisier)
    end = time.time()
    pbar.destroy()
    pbargui.destroy()
    messagebox.showinfo('Finalizat!', "Prelucrate in " + str(end - start)[:6] + " secunde.")
    return None


def wires_complet():
    pbargui = Tk()
    pbargui.title("Prelucrare WIRELIST-uri")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    array_wirelisturi = ["8000", "8001", "8011", "8012", "8013", "8014", "8023", "8024", "8025", "8026", "8027",
                         "8030", "8031", "8032", "8052", "8053"]

    dir_wirelist = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                           title="Selectati directorul cu fisiere:")
    start = time.time()
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_wirelist):
        if file_all.endswith(".csv"):
            file_counter = file_counter + 1
    for file_all in os.listdir(dir_wirelist):
        if file_all.endswith(".csv"):
            file_progres = file_progres + 1
            statuslabel["text"] = str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 2
            pbargui.update_idletasks()
            for item in array_wirelisturi:
                if item in file_all:
                    try:
                        array_original = []
                        with open(dir_wirelist + "/" + file_all, newline='') as csvfile:
                            array_incarcat = list(csv.reader(csvfile, delimiter=';'))
                        if "AEM" in array_incarcat[0][1]:
                            for i in range(len(array_incarcat)):
                                del array_incarcat[i][1]
                        for x in range(len(array_incarcat)):
                            array_t = []
                            for i in range(len(array_incarcat[x])):
                                if i < 27 and array_incarcat[x][0] == "3":
                                    if array_incarcat[x][i] == "":
                                        array_t.append("X")
                                    else:
                                        array_t.append(array_incarcat[x][i])
                                elif array_incarcat[x][i] != "":
                                    array_t.append(array_incarcat[x][i])
                            array_original.append(array_t)
                        if array_original[0][0] != 1 and len(array_original[0]) != 2:
                            messagebox.showerror('Eroare fisier' + file_all, 'Nu ai incarcat fisierul corect ...')
                            return
                        if len(array_original) < 50:
                            for i in range(0, len(array_original)):
                                if array_original[i][0] == "0":
                                    if str(array_original[i][1]) != "Ltg-Nr.":
                                        messagebox.showerror('Eroare fisier' + file_all,
                                                             'Nu ai incarcat fisierul corect')
                                        return
                        else:
                            for i in range(0, 50):
                                if array_original[i][0] == "0":
                                    if str(array_original[i][1]) != "Ltg-Nr.":
                                        messagebox.showerror('Eroare fisier' + file_all,
                                                             'Nu ai incarcat fisierul corect')
                                        return
                    except:
                        messagebox.showerror('Eroare fisier' + file_all, 'Nu ai incarcat fisierul corect')
                        return
                    nume_fisier = os.path.splitext(os.path.basename(file_all))[0]
                    array_output = []
                    array_module = []
                    array_wires = []
                    for i in range(1, len(array_incarcat)):
                        if array_incarcat[i][0] == "0":
                            pot_position = array_incarcat[i].index('Pot.') + 1
                            array_output.append(array_incarcat[i][1:pot_position])
                            array_output[0].append("Lange")
                            break
                    for i in range(1, len(array_incarcat)):
                        if i == len(array_incarcat) - 1:
                            array_wires.append(array_incarcat[i])

                            for verwires in range(len(array_wires[0])):
                                for vermodule in range(len(array_module)):
                                    if array_wires[0][verwires] == array_module[vermodule][1] or \
                                            array_wires[0][verwires] == "Length  " + array_module[vermodule][1]:
                                        array_wires[0][verwires] = array_module[vermodule][2]
                            pot_position = array_wires[0].index('Pot.') + 1
                            ltgno_position = array_wires[0][pot_position:].index('Ltg-Nr.') + pot_position
                            array_out_temp = []
                            for wire in range(1, len(array_wires)):
                                for index in range(pot_position, ltgno_position):
                                    if array_wires[wire][index] != "-":
                                        temp_list = array_wires[wire][1:pot_position]
                                        temp_list.append(array_wires[wire][index])
                                        array_out_temp.append(temp_list)
                            array_module = []
                            array_wires = []
                            array_output.extend(array_out_temp)
                        else:
                            if array_incarcat[i][0] == "2":
                                array_module.append(array_incarcat[i])
                            elif array_incarcat[i][0] == "3" or array_incarcat[i][0] == "0":
                                array_wires.append(array_incarcat[i])
                            elif array_incarcat[i][0] == "1":
                                for verwires in range(len(array_wires[0])):
                                    for vermodule in range(len(array_module)):
                                        if array_wires[0][verwires] == array_module[vermodule][1] or \
                                                array_wires[0][verwires] == "Length  " + array_module[vermodule][1]:
                                            array_wires[0][verwires] = array_module[vermodule][2]
                                pot_position = array_wires[0].index('Pot.') + 1
                                ltgno_position = array_wires[0][pot_position:].index('Ltg-Nr.') + pot_position
                                array_out_temp = []
                                for wire in range(1, len(array_wires)):
                                    for index in range(pot_position, ltgno_position):
                                        if array_wires[wire][index] != "-":
                                            temp_list = array_wires[wire][1:pot_position]
                                            temp_list.append(array_wires[wire][index])
                                            array_out_temp.append(temp_list)
                                array_module = []
                                array_wires = []
                                array_output.extend(array_out_temp)
                    prn_excel_wires_complete(array_output, nume_fisier)
    end = time.time()
    pbar.destroy()
    pbargui.destroy()
    messagebox.showinfo('Finalizat!', "Prelucrate in " + str(end - start)[:6] + " secunde.")
    return None


def wires_pnleoni():
    pbargui = Tk()
    pbargui.title("Prelucrare WIRELIST-uri")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    array_wirelisturi = ["8000", "8001", "8011", "8012", "8013", "8014", "8023", "8024", "8025", "8026", "8027",
                         "8030", "8031", "8032", "8052", "8053"]
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Component Overview.txt",
              newline='') as csvfile:
        array_componente = list(csv.reader(csvfile, delimiter=';'))
    dir_wirelist = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                           title="Selectati directorul cu fisiere:")
    start = time.time()
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_wirelist):
        if file_all.endswith(".csv"):
            file_counter = file_counter + 1
    for file_all in os.listdir(dir_wirelist):
        if file_all.endswith(".csv"):
            file_progres = file_progres + 1
            statuslabel["text"] = str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 2
            pbargui.update_idletasks()
            for item in array_wirelisturi:
                if item in file_all:
                    try:
                        array_original = []
                        with open(dir_wirelist + "/" + file_all, newline='') as csvfile:
                            array_incarcat = list(csv.reader(csvfile, delimiter=';'))
                        if "AEM" in array_incarcat[0][1]:
                            for i in range(len(array_incarcat)):
                                del array_incarcat[i][1]
                        for x in range(len(array_incarcat)):
                            array_t = []
                            for i in range(len(array_incarcat[x])):
                                if i < 27 and array_incarcat[x][0] == "3":
                                    if array_incarcat[x][i] == "":
                                        array_t.append("X")
                                    else:
                                        array_t.append(array_incarcat[x][i])
                                elif array_incarcat[x][i] != "":
                                    array_t.append(array_incarcat[x][i])
                            array_original.append(array_t)
                        if array_original[0][0] != 1 and len(array_original[0]) != 2:
                            messagebox.showerror('Eroare fisier' + file_all, 'Nu ai incarcat fisierul corect ...')
                            return
                        if len(array_original) < 50:
                            for i in range(0, len(array_original)):
                                if array_original[i][0] == "0":
                                    if str(array_original[i][1]) != "Ltg-Nr.":
                                        messagebox.showerror('Eroare fisier' + file_all,
                                                             'Nu ai incarcat fisierul corect')
                                        return
                        else:
                            for i in range(0, 50):
                                if array_original[i][0] == "0":
                                    if str(array_original[i][1]) != "Ltg-Nr.":
                                        messagebox.showerror('Eroare fisier' + file_all,
                                                             'Nu ai incarcat fisierul corect')
                                        return
                    except:
                        messagebox.showerror('Eroare fisier' + file_all, 'Nu ai incarcat fisierul corect')
                        return
                    nume_fisier = os.path.splitext(os.path.basename(file_all))[0]
                    array_output = []
                    array_module = []
                    array_wires = []
                    for i in range(1, len(array_incarcat)):
                        if array_incarcat[i][0] == "0":
                            pot_position = array_incarcat[i].index('Pot.') + 1
                            array_output.append(array_incarcat[i][1:pot_position])
                            array_output[0].append("Lange")
                            break
                    for i in range(1, len(array_incarcat)):
                        if i == len(array_incarcat) - 1:
                            array_wires.append(array_incarcat[i])

                            for verwires in range(len(array_wires[0])):
                                for vermodule in range(len(array_module)):
                                    if array_wires[0][verwires] == array_module[vermodule][1] or \
                                            array_wires[0][verwires] == "Length  " + array_module[vermodule][1]:
                                        array_wires[0][verwires] = array_module[vermodule][2]
                            pot_position = array_wires[0].index('Pot.') + 1
                            ltgno_position = array_wires[0][pot_position:].index('Ltg-Nr.') + pot_position
                            array_out_temp = []
                            for wire in range(1, len(array_wires)):
                                for index in range(pot_position, ltgno_position):
                                    if array_wires[wire][index] != "-":
                                        temp_list = array_wires[wire][1:pot_position]
                                        temp_list.append(array_wires[wire][index])
                                        array_out_temp.append(temp_list)
                            array_module = []
                            array_wires = []
                            array_output.extend(array_out_temp)
                        else:
                            if array_incarcat[i][0] == "2":
                                array_module.append(array_incarcat[i])
                            elif array_incarcat[i][0] == "3" or array_incarcat[i][0] == "0":
                                array_wires.append(array_incarcat[i])
                            elif array_incarcat[i][0] == "1":
                                for verwires in range(len(array_wires[0])):
                                    for vermodule in range(len(array_module)):
                                        if array_wires[0][verwires] == array_module[vermodule][1] or \
                                                array_wires[0][verwires] == "Length  " + array_module[vermodule][1]:
                                            array_wires[0][verwires] = array_module[vermodule][2]
                                pot_position = array_wires[0].index('Pot.') + 1
                                ltgno_position = array_wires[0][pot_position:].index('Ltg-Nr.') + pot_position
                                array_out_temp = []
                                for wire in range(1, len(array_wires)):
                                    for index in range(pot_position, ltgno_position):
                                        if array_wires[wire][index] != "-":
                                            temp_list = array_wires[wire][1:pot_position]
                                            temp_list.append(array_wires[wire][index])
                                            array_out_temp.append(temp_list)
                                array_module = []
                                array_wires = []
                                array_output.extend(array_out_temp)
                    array_output_comp1 = array_output
                    for x in range(len(array_output_comp1)):
                        for y in range(len(array_output_comp1[x])):
                            for z in range(len(array_componente)):
                                if array_output_comp1[x][y] == array_componente[z][0]:
                                    array_output_comp1[x][y] = array_componente[z][2]
                    array_output_comp2 = array_output
                    for x in range(len(array_output_comp2)):
                        for y in range(len(array_output_comp2[x])):
                            for z in range(len(array_componente)):
                                if array_output_comp2[x][y] == array_componente[z][0]:
                                    array_output_comp2[x][y] = array_componente[z][2]

                    prn_excel_wires_complete_leoni(array_output, array_output_comp1, array_output_comp2, nume_fisier)
    end = time.time()
    pbar.destroy()
    pbargui.destroy()
    messagebox.showinfo('Finalizat!', "Prelucrate in " + str(end - start)[:6] + " secunde.")
    return None


def boms_pnleoni():
    pbargui = Tk()
    pbargui.title("Prelucrare BOM-uri")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    global nume_fisier
    array_boms = ["8000", "8001", "8011", "8012", "8013", "8014", "8023", "8024", "8025", "8026", "8027",
                  "8030", "8031", "8032", "8052", "8053"]
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Component Overview.txt",
              newline='') as csvfile:
        array_componente = list(csv.reader(csvfile, delimiter=';'))
    dir_BOM = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir), title="Selectati directorul cu fisiere:")
    start = time.time()
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_BOM):
        if file_all.endswith(".csv"):
            file_counter = file_counter + 1
    for file_all in os.listdir(dir_BOM):
        if file_all.endswith(".csv"):
            file_progres = file_progres + 1
            statuslabel["text"] = str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 2
            pbargui.update_idletasks()
            for item in array_boms:
                if item in file_all:
                    try:
                        with open(dir_BOM + "/" + file_all, newline='') as csvfile:
                            array_incarcat = list(csv.reader(csvfile, delimiter=';'))
                        if "AEM" in array_incarcat[0][1]:
                            for i in range(len(array_incarcat)):
                                del array_incarcat[i][1]
                        if array_incarcat[0][0] != "1" and len(array_incarcat[0]) != 2:
                            messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect')
                            return
                        if len(array_incarcat) < 50:
                            for i in range(0, len(array_incarcat)):
                                if array_incarcat[i][0] == "0":
                                    if str(array_incarcat[i][-1])[:1] != "V":
                                        messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect')
                                        return
                        else:
                            for i in range(0, 50):
                                if array_incarcat[i][0] == "0":
                                    if str(array_incarcat[i][-1])[:1] != "V":
                                        messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect')
                                        return
                    except:
                        messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect')
                        return
                    nume_fisier = os.path.splitext(os.path.basename(file_all))[0]
                    array_output = [
                        ["Module", "Quantity", "Bezeichnung", "VOBES-ID", "Benennung", "Verwendung", "Verwendung",
                         "Kurzname", "xy", "Teilenummer", "Vorzugsteil", "TAB-Nummer", "Referenzteil", "Farbe",
                         "E-Komponente", "E-Komponente Part-Nr.", "Einh."]]
                    array_module = []
                    array_comp = []
                    for i in range(1, len(array_incarcat)):
                        if i == len(array_incarcat) - 1:
                            array_comp.append(array_incarcat[i])
                            array_out_temp = []
                            for vercomp in range(len(array_comp[0])):
                                for vermodule in range(len(array_module)):
                                    if array_module[vermodule][1] in array_comp[0][vercomp]:
                                        array_comp[0][vercomp] = array_module[vermodule][2]
                            last_position = len(array_comp[0])
                            index_beze = array_comp[0].index('Bezeichnung')
                            index_VID = array_comp[0].index('VOBES-ID')
                            try:
                                index_bene = array_comp[0].index('Benennung')
                            except ValueError:
                                index_bene = index_VID
                            index_verew1 = array_comp[0].index('Verwendung')
                            index_verew2 = array_comp[0][index_verew1:].index('Verwendung') + index_verew1
                            index_kurz = array_comp[0].index('Kurzname')
                            index_xy = array_comp[0].index('xy')
                            index_teile = array_comp[0].index('Teilenummer')
                            try:
                                index_vorzug = array_comp[0].index('Vorzugsteil')
                            except ValueError:
                                index_vorzug = index_teile
                            index_tab = array_comp[0].index('TAB-Nummer')
                            index_refe = array_comp[0].index('Referenzteil')
                            index_farbe = array_comp[0].index('Farbe')
                            try:
                                index_ekomp = array_comp[0].index('E-Komponente')
                            except ValueError:
                                index_ekomp = index_VID
                            try:
                                index_ekomppn = array_comp[0].index('E-Komponente Part-Nr.')
                            except ValueError:
                                index_ekomppn = index_VID
                            index_einh = array_comp[0].index('Einh.')
                            for comp in range(1, len(array_comp)):
                                for index in range(index_einh + 1, last_position):
                                    if array_comp[comp][index] != "0" and array_comp[comp][index] != "-":
                                        array_out_temp.append([array_comp[0][index], array_comp[comp][index],
                                                               array_comp[comp][index_beze],
                                                               array_comp[comp][index_VID],
                                                               array_comp[comp][index_bene],
                                                               array_comp[comp][index_verew1],
                                                               array_comp[comp][index_verew2],
                                                               array_comp[comp][index_kurz],
                                                               array_comp[comp][index_xy],
                                                               array_comp[comp][index_teile],
                                                               array_comp[comp][index_vorzug],
                                                               array_comp[comp][index_tab],
                                                               array_comp[comp][index_refe],
                                                               array_comp[comp][index_farbe],
                                                               array_comp[comp][index_ekomp],
                                                               array_comp[comp][index_ekomppn],
                                                               array_comp[comp][index_einh]])
                            array_module = []
                            array_comp = []
                            array_output.extend(array_out_temp)
                        else:
                            if array_incarcat[i][0] == "2":
                                array_module.append(array_incarcat[i])
                            elif array_incarcat[i][0] == "3" or array_incarcat[i][0] == "4" or \
                                    array_incarcat[i][0] == "0":
                                array_comp.append(array_incarcat[i])
                            elif array_incarcat[i][0] == "1":
                                for vercomp in range(len(array_comp[0])):
                                    for vermodule in range(len(array_module)):
                                        if array_module[vermodule][1] in array_comp[0][vercomp]:
                                            array_comp[0][vercomp] = array_module[vermodule][2]
                                array_out_temp = []
                                last_position = len(array_comp[0])
                                index_beze = array_comp[0].index('Bezeichnung')
                                index_VID = array_comp[0].index('VOBES-ID')
                                try:
                                    index_bene = array_comp[0].index('Benennung')
                                except ValueError:
                                    index_bene = index_VID
                                index_verew1 = array_comp[0].index('Verwendung')
                                index_verew2 = array_comp[0][index_verew1:].index('Verwendung') + index_verew1
                                index_kurz = array_comp[0].index('Kurzname')
                                index_xy = array_comp[0].index('xy')
                                index_teile = array_comp[0].index('Teilenummer')
                                try:
                                    index_vorzug = array_comp[0].index('Vorzugsteil')
                                except ValueError:
                                    index_vorzug = index_teile
                                index_tab = array_comp[0].index('TAB-Nummer')
                                index_refe = array_comp[0].index('Referenzteil')
                                index_farbe = array_comp[0].index('Farbe')
                                try:
                                    index_ekomp = array_comp[0].index('E-Komponente')
                                except ValueError:
                                    index_ekomp = index_VID
                                try:
                                    index_ekomppn = array_comp[0].index('E-Komponente Part-Nr.')
                                except ValueError:
                                    index_ekomppn = index_VID
                                index_einh = array_comp[0].index('Einh.')

                                for comp in range(1, len(array_comp)):
                                    for index in range(index_einh + 1, last_position):
                                        if array_comp[comp][index] != "0" and array_comp[comp][index] != "-":
                                            array_out_temp.append([array_comp[0][index], array_comp[comp][index],
                                                                   array_comp[comp][index_beze],
                                                                   array_comp[comp][index_VID],
                                                                   array_comp[comp][index_bene],
                                                                   array_comp[comp][index_verew1],
                                                                   array_comp[comp][index_verew2],
                                                                   array_comp[comp][index_kurz],
                                                                   array_comp[comp][index_xy],
                                                                   array_comp[comp][index_teile],
                                                                   array_comp[comp][index_vorzug],
                                                                   array_comp[comp][index_tab],
                                                                   array_comp[comp][index_refe],
                                                                   array_comp[comp][index_farbe],
                                                                   array_comp[comp][index_ekomp],
                                                                   array_comp[comp][index_ekomppn],
                                                                   array_comp[comp][index_einh]])
                                array_module = []
                                array_comp = []
                                array_output.extend(array_out_temp)
                    array_output[0].append("PN Leoni 1")
                    array_output[0].append("PN Leoni 2")
                    for x in range(len(array_output)):
                        for y in range(len(array_componente)):
                            if array_output[x][9] == array_componente[y][0]:
                                array_output[x].append(array_componente[y][1])
                                array_output[x].append(array_componente[y][2])
                    prn_excel_bom_complete(array_output, nume_fisier)
    end = time.time()
    pbar.destroy()
    pbargui.destroy()
    messagebox.showinfo('Finalizat!', "Prelucrate in " + str(end - start)[:6] + " secunde.")
    return None


def boms_cumulat():
    pbargui = Tk()
    pbargui.title("Prelucrare BOM-uri")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    global nume_fisier
    array_boms = ["8000", "8001", "8011", "8012", "8013", "8014", "8023", "8024", "8025", "8026", "8027",
                  "8030", "8031", "8032", "8052", "8053"]
    array_output = [["Module", "Quantity", "Bezeichnung", "VOBES-ID", "Benennung", "Verwendung", "Verwendung",
                     "Kurzname", "xy", "Teilenummer", "Vorzugsteil", "TAB-Nummer", "Referenzteil", "Farbe",
                     "E-Komponente", "E-Komponente Part-Nr.", "Einh.", "Platforma"]]
    dir_BOM = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir), title="Selectati directorul cu fisiere:")
    start = time.time()
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_BOM):
        if file_all.endswith(".csv"):
            file_counter = file_counter + 1
    for file_all in os.listdir(dir_BOM):
        if file_all.endswith(".csv"):
            file_progres = file_progres + 1
            statuslabel["text"] = str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 2
            pbargui.update_idletasks()
            for item in array_boms:
                if item in file_all:
                    try:
                        with open(dir_BOM + "/" + file_all, newline='') as csvfile:
                            array_incarcat = list(csv.reader(csvfile, delimiter=';'))
                        if "AEM" in array_incarcat[0][1]:
                            for i in range(len(array_incarcat)):
                                del array_incarcat[i][1]
                        if array_incarcat[0][0] != "1" and len(array_incarcat[0]) != 2:
                            messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect')
                            return
                        if len(array_incarcat) < 50:
                            for i in range(0, len(array_incarcat)):
                                if array_incarcat[i][0] == "0":
                                    if str(array_incarcat[i][-1])[:1] != "V":
                                        messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect')
                                        return
                        else:
                            for i in range(0, 50):
                                if array_incarcat[i][0] == "0":
                                    if str(array_incarcat[i][-1])[:1] != "V":
                                        messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect')
                                        return
                    except:
                        messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect')
                        return
                    nume_fisier = os.path.splitext(os.path.basename(file_all))[0]

                    array_module = []
                    array_comp = []
                    for i in range(1, len(array_incarcat)):
                        if i == len(array_incarcat) - 1:
                            array_comp.append(array_incarcat[i])
                            array_out_temp = []
                            for vercomp in range(len(array_comp[0])):
                                for vermodule in range(len(array_module)):
                                    if array_module[vermodule][1] in array_comp[0][vercomp]:
                                        array_comp[0][vercomp] = array_module[vermodule][2]
                            last_position = len(array_comp[0])
                            index_beze = array_comp[0].index('Bezeichnung')
                            index_VID = array_comp[0].index('VOBES-ID')
                            try:
                                index_bene = array_comp[0].index('Benennung')
                            except ValueError:
                                index_bene = index_VID
                            index_verew1 = array_comp[0].index('Verwendung')
                            index_verew2 = array_comp[0][index_verew1:].index('Verwendung') + index_verew1
                            index_kurz = array_comp[0].index('Kurzname')
                            index_xy = array_comp[0].index('xy')
                            index_teile = array_comp[0].index('Teilenummer')
                            try:
                                index_vorzug = array_comp[0].index('Vorzugsteil')
                            except ValueError:
                                index_vorzug = index_teile
                            index_tab = array_comp[0].index('TAB-Nummer')
                            index_refe = array_comp[0].index('Referenzteil')
                            index_farbe = array_comp[0].index('Farbe')
                            try:
                                index_ekomp = array_comp[0].index('E-Komponente')
                            except ValueError:
                                index_ekomp = index_VID
                            try:
                                index_ekomppn = array_comp[0].index('E-Komponente Part-Nr.')
                            except ValueError:
                                index_ekomppn = index_VID
                            index_einh = array_comp[0].index('Einh.')
                            for comp in range(1, len(array_comp)):
                                for index in range(index_einh + 1, last_position):
                                    if array_comp[comp][index] != "0" and array_comp[comp][index] != "-":
                                        array_out_temp.append([array_comp[0][index], array_comp[comp][index],
                                                               array_comp[comp][index_beze],
                                                               array_comp[comp][index_VID],
                                                               array_comp[comp][index_bene],
                                                               array_comp[comp][index_verew1],
                                                               array_comp[comp][index_verew2],
                                                               array_comp[comp][index_kurz],
                                                               array_comp[comp][index_xy],
                                                               array_comp[comp][index_teile],
                                                               array_comp[comp][index_vorzug],
                                                               array_comp[comp][index_tab],
                                                               array_comp[comp][index_refe],
                                                               array_comp[comp][index_farbe],
                                                               array_comp[comp][index_ekomp],
                                                               array_comp[comp][index_ekomppn],
                                                               array_comp[comp][index_einh], nume_fisier])
                            array_module = []
                            array_comp = []
                            array_output.extend(array_out_temp)
                        else:
                            if array_incarcat[i][0] == "2":
                                array_module.append(array_incarcat[i])
                            elif array_incarcat[i][0] == "3" or array_incarcat[i][0] == "4" or \
                                    array_incarcat[i][0] == "0":
                                array_comp.append(array_incarcat[i])
                            elif array_incarcat[i][0] == "1":
                                for vercomp in range(len(array_comp[0])):
                                    for vermodule in range(len(array_module)):
                                        if array_module[vermodule][1] in array_comp[0][vercomp]:
                                            array_comp[0][vercomp] = array_module[vermodule][2]
                                array_out_temp = []
                                last_position = len(array_comp[0])
                                index_beze = array_comp[0].index('Bezeichnung')
                                index_VID = array_comp[0].index('VOBES-ID')
                                try:
                                    index_bene = array_comp[0].index('Benennung')
                                except ValueError:
                                    index_bene = index_VID
                                index_verew1 = array_comp[0].index('Verwendung')
                                index_verew2 = array_comp[0][index_verew1:].index('Verwendung') + index_verew1
                                index_kurz = array_comp[0].index('Kurzname')
                                index_xy = array_comp[0].index('xy')
                                index_teile = array_comp[0].index('Teilenummer')
                                try:
                                    index_vorzug = array_comp[0].index('Vorzugsteil')
                                except ValueError:
                                    index_vorzug = index_teile
                                index_tab = array_comp[0].index('TAB-Nummer')
                                index_refe = array_comp[0].index('Referenzteil')
                                index_farbe = array_comp[0].index('Farbe')
                                try:
                                    index_ekomp = array_comp[0].index('E-Komponente')
                                except ValueError:
                                    index_ekomp = index_VID
                                try:
                                    index_ekomppn = array_comp[0].index('E-Komponente Part-Nr.')
                                except ValueError:
                                    index_ekomppn = index_VID
                                index_einh = array_comp[0].index('Einh.')

                                for comp in range(1, len(array_comp)):
                                    for index in range(index_einh + 1, last_position):
                                        if array_comp[comp][index] != "0" and array_comp[comp][index] != "-":
                                            array_out_temp.append([array_comp[0][index], array_comp[comp][index],
                                                                   array_comp[comp][index_beze],
                                                                   array_comp[comp][index_VID],
                                                                   array_comp[comp][index_bene],
                                                                   array_comp[comp][index_verew1],
                                                                   array_comp[comp][index_verew2],
                                                                   array_comp[comp][index_kurz],
                                                                   array_comp[comp][index_xy],
                                                                   array_comp[comp][index_teile],
                                                                   array_comp[comp][index_vorzug],
                                                                   array_comp[comp][index_tab],
                                                                   array_comp[comp][index_refe],
                                                                   array_comp[comp][index_farbe],
                                                                   array_comp[comp][index_ekomp],
                                                                   array_comp[comp][index_ekomppn],
                                                                   array_comp[comp][index_einh], nume_fisier])
                                array_module = []
                                array_comp = []
                                array_output.extend(array_out_temp)
    prn_excel_boomcumulat(array_output)
    end = time.time()
    pbar.destroy()
    pbargui.destroy()
    messagebox.showinfo('Finalizat!', "Prelucrate in " + str(end - start)[:6] + " secunde.")
    return None


def wires_cumulat():
    pbargui = Tk()
    pbargui.title("Prelucrare WIRELIST-uri")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    array_wirelisturi = ["8000", "8001", "8011", "8012", "8013", "8014", "8023", "8024", "8025", "8026", "8027",
                         "8030", "8031", "8032", "8052", "8053"]
    array_output = [["Ltg-Nr.", "Verbindung", "von", "Kurzname", "Pin", "xy", "Kontakt", "Dichtung",
                     "nach", "Kurzname", "Pin", "xy", "Kontakt", "Dichtung", "Leitung", "Sonderltg.",
                     "Farbe", "Quer.", "Pot.", "Lange", "Platforma"]]
    dir_wirelist = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                           title="Selectati directorul cu fisiere:")
    start = time.time()
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_wirelist):
        if file_all.endswith(".csv"):
            file_counter = file_counter + 1
    for file_all in os.listdir(dir_wirelist):
        if file_all.endswith(".csv"):
            file_progres = file_progres + 1
            statuslabel["text"] = str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 2
            pbargui.update_idletasks()
            for item in array_wirelisturi:
                if item in file_all:
                    try:
                        array_original = []
                        with open(dir_wirelist + "/" + file_all, newline='') as csvfile:
                            array_incarcat = list(csv.reader(csvfile, delimiter=';'))
                        if "AEM" in array_incarcat[0][1]:
                            for i in range(len(array_incarcat)):
                                del array_incarcat[i][1]
                        for x in range(len(array_incarcat)):
                            array_t = []
                            for i in range(len(array_incarcat[x])):
                                if i < 27 and array_incarcat[x][0] == "3":
                                    if array_incarcat[x][i] == "":
                                        array_t.append("X")
                                    else:
                                        array_t.append(array_incarcat[x][i])
                                elif array_incarcat[x][i] != "":
                                    array_t.append(array_incarcat[x][i])
                            array_original.append(array_t)
                        if array_original[0][0] != 1 and len(array_original[0]) != 2:
                            messagebox.showerror('Eroare fisier' + file_all, 'Nu ai incarcat fisierul corect ...')
                            return
                        if len(array_original) < 50:
                            for i in range(0, len(array_original)):
                                if array_original[i][0] == "0":
                                    if str(array_original[i][1]) != "Ltg-Nr.":
                                        messagebox.showerror('Eroare fisier' + file_all,
                                                             'Nu ai incarcat fisierul corect')
                                        return
                        else:
                            for i in range(0, 50):
                                if array_original[i][0] == "0":
                                    if str(array_original[i][1]) != "Ltg-Nr.":
                                        messagebox.showerror('Eroare fisier' + file_all,
                                                             'Nu ai incarcat fisierul corect')
                                        return
                    except:
                        messagebox.showerror('Eroare fisier' + file_all, 'Nu ai incarcat fisierul corect')
                        return
                    nume_fisier = os.path.splitext(os.path.basename(file_all))[0]
                    platforma = nume_fisier[0:4]
                    array_module = []
                    array_wires = []

                    for i in range(1, len(array_incarcat)):
                        if i == len(array_incarcat) - 1:
                            array_out_temp = []
                            array_wires.append(array_incarcat[i])
                            for verwires in range(len(array_wires[0])):
                                for vermodule in range(len(array_module)):
                                    if array_wires[0][verwires] == array_module[vermodule][1] or \
                                            array_wires[0][verwires] == "Length  " + array_module[vermodule][1]:
                                        array_wires[0][verwires] = array_module[vermodule][2]
                            pot_position = array_wires[0].index('Pot.') + 1
                            ltgno_position = array_wires[0][pot_position:].index('Ltg-Nr.') + pot_position
                            index_ltgno = array_wires[0].index('Ltg-Nr.')
                            index_verbindung = array_wires[0].index('Verbindung')
                            index_von = array_wires[0].index('von')
                            index_nach = array_wires[0].index('nach')
                            index_quer = array_wires[0].index('Quer.')
                            index_kurz1 = array_wires[0].index('Kurzname')
                            index_pin1 = array_wires[0].index('Pin')
                            index_kurz2 = array_wires[0][index_kurz1 + 1:].index('Kurzname') + index_kurz1 + 1
                            index_pin2 = array_wires[0][index_pin1 + 1:].index('Pin') + index_pin1 + 1
                            index_xy1 = array_wires[0].index('xy')
                            index_kontakt1 = array_wires[0].index('Kontakt')
                            index_xy2 = array_wires[0][index_xy1 + 1:].index('xy') + index_xy1 + 1
                            index_kontakt2 = array_wires[0][index_kontakt1 + 1:].index('Kontakt') + index_kontakt1 + 1
                            index_dichtung = array_wires[0].index('Dichtung')
                            index_dichtung2 = array_wires[0][index_dichtung + 1:].index('Dichtung') + index_dichtung + 1
                            index_leitung = array_wires[0].index('Leitung')
                            index_farbe = array_wires[0].index('Farbe')
                            index_sonder = array_wires[0].index('Sonderltg.')
                            for wire in range(1, len(array_wires)):
                                for index in range(pot_position, ltgno_position):
                                    if array_wires[wire][index] != "-":
                                        array_out_temp.append([array_wires[wire][index_ltgno],
                                                               array_wires[wire][index_verbindung],
                                                               array_wires[wire][index_von],
                                                               array_wires[wire][index_kurz1],
                                                               array_wires[wire][index_pin1],
                                                               array_wires[wire][index_xy1],
                                                               array_wires[wire][index_kontakt1],
                                                               array_wires[wire][index_dichtung],
                                                               array_wires[wire][index_nach],
                                                               array_wires[wire][index_kurz2],
                                                               array_wires[wire][index_pin2],
                                                               array_wires[wire][index_xy2],
                                                               array_wires[wire][index_kontakt2],
                                                               array_wires[wire][index_dichtung2],
                                                               array_wires[wire][index_leitung],
                                                               array_wires[wire][index_sonder],
                                                               array_wires[wire][index_farbe],
                                                               array_wires[wire][index_quer],
                                                               array_wires[wire][pot_position],
                                                               array_wires[wire][index],
                                                               platforma])
                            array_module = []
                            array_wires = []
                            array_output.extend(array_out_temp)
                        else:
                            array_out_temp = []
                            if array_incarcat[i][0] == "2":
                                array_module.append(array_incarcat[i])
                            elif array_incarcat[i][0] == "3" or array_incarcat[i][0] == "0":
                                array_wires.append(array_incarcat[i])
                            elif array_incarcat[i][0] == "1":
                                for verwires in range(len(array_wires[0])):
                                    for vermodule in range(len(array_module)):
                                        if array_wires[0][verwires] == array_module[vermodule][1] or \
                                                array_wires[0][verwires] == "Length  " + array_module[vermodule][1]:
                                            array_wires[0][verwires] = array_module[vermodule][2]
                                pot_position = array_wires[0].index('Pot.') + 1
                                ltgno_position = array_wires[0][pot_position:].index('Ltg-Nr.') + pot_position
                                index_ltgno = array_wires[0].index('Ltg-Nr.')
                                index_verbindung = array_wires[0].index('Verbindung')
                                index_von = array_wires[0].index('von')
                                index_nach = array_wires[0].index('nach')
                                index_quer = array_wires[0].index('Quer.')
                                index_kurz1 = array_wires[0].index('Kurzname')
                                index_pin1 = array_wires[0].index('Pin')
                                index_kurz2 = array_wires[0][index_kurz1 + 1:].index('Kurzname') + index_kurz1 + 1
                                index_pin2 = array_wires[0][index_pin1 + 1:].index('Pin') + index_pin1 + 1
                                index_xy1 = array_wires[0].index('xy')
                                index_kontakt1 = array_wires[0].index('Kontakt')
                                index_xy2 = array_wires[0][index_xy1 + 1:].index('xy') + index_xy1 + 1
                                index_kontakt2 = array_wires[0][index_kontakt1 + 1:].index(
                                    'Kontakt') + index_kontakt1 + 1
                                index_dichtung = array_wires[0].index('Dichtung')
                                index_dichtung2 = array_wires[0][index_dichtung + 1:].index(
                                    'Dichtung') + index_dichtung + 1
                                index_leitung = array_wires[0].index('Leitung')
                                index_farbe = array_wires[0].index('Farbe')
                                index_sonder = array_wires[0].index('Sonderltg.')
                                for wire in range(1, len(array_wires)):
                                    for index in range(pot_position, ltgno_position):
                                        if array_wires[wire][index] != "-":
                                            array_out_temp.append([array_wires[wire][index_ltgno],
                                                                   array_wires[wire][index_verbindung],
                                                                   array_wires[wire][index_von],
                                                                   array_wires[wire][index_kurz1],
                                                                   array_wires[wire][index_pin1],
                                                                   array_wires[wire][index_xy1],
                                                                   array_wires[wire][index_kontakt1],
                                                                   array_wires[wire][index_dichtung],
                                                                   array_wires[wire][index_nach],
                                                                   array_wires[wire][index_kurz2],
                                                                   array_wires[wire][index_pin2],
                                                                   array_wires[wire][index_xy2],
                                                                   array_wires[wire][index_kontakt2],
                                                                   array_wires[wire][index_dichtung2],
                                                                   array_wires[wire][index_leitung],
                                                                   array_wires[wire][index_sonder],
                                                                   array_wires[wire][index_farbe],
                                                                   array_wires[wire][index_quer],
                                                                   array_wires[wire][pot_position],
                                                                   array_wires[wire][index],
                                                                   platforma])
                                array_module = []
                                array_wires = []
                                array_output.extend(array_out_temp)

    prn_excel_wirelistsallinone(array_output)
    end = time.time()
    pbar.destroy()
    pbargui.destroy()
    messagebox.showinfo('Finalizat!', "Prelucrate in " + str(end - start)[:6] + " secunde.")
    return None
