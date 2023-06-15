import csv
import os
import time
from tkinter import Tk, ttk, HORIZONTAL, Label, filedialog, messagebox

import pandas as pd
from openpyxl import load_workbook
from functii_print import prn_excel_diagrame, prn_excel_infoindiagrame, prn_excel_matrixmodule, prn_excel_bmmodule, \
    prn_excel_diagrameinksk


def comparatiediagrame():
    pbargui = Tk()
    pbargui.title("Comparatie diagrame")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    dir_old = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                      title="Selectati directorul cu diagramele vechi:")
    dir_new = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                      title="Selectati directorul cu diagramele noi:")
    start = time.time()
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_new):
        if file_all.endswith(".xlsx"):
            file_counter = file_counter + 1
    if file_counter == 0:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showwarning('Eroare!', "Directorul selectat este gol.")

    # verificare diagrame in ambele directoare
    array_log = []
    array_new = []
    for file_all in os.listdir(dir_new):
        if file_all.endswith(".xlsx"):
            file_progres = file_progres + 1
            statuslabel["text"] = str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 2
            pbargui.update_idletasks()
            wb1 = load_workbook(dir_new + "/" + file_all)
            try:
                wb2 = load_workbook(dir_old + "/" + file_all)
                sheet1 = wb1.worksheets[0]
                sheet2 = wb2.worksheets[0]

                # iterate through the rows and columns of both worksheets
                for row in range(1, sheet1.max_row + 1):
                    for col in range(1, sheet1.max_column + 1):
                        cell1 = sheet1.cell(row, col)
                        cell2 = sheet2.cell(row, col)
                        if cell1.value != cell2.value:
                            array_log.append([file_all, cell1.value, cell2.value, row, col])
                    pbar['value'] += 2
                    pbargui.update_idletasks()
            except:
                array_new.append([file_all])
    array_log.insert(0, ["Fisier", "Valoare noua", "Valoare veche", "Rand", "Coloana"])
    prn_excel_diagrame(array_log, array_new)

    pbar.destroy()
    pbargui.destroy()
    end = time.time()
    messagebox.showinfo('Finalizat!', "Prelucrate in " + str(end - start)[:6] + " secunde.")


# this code defines a function called "comparatiediagrame" that creates a GUI window using the tkinter library.
# Within the GUI, the function prompts the user to select two directories containing Excel files, and then compares
# the cells of each Excel file in the two directories to see if they have the same values. If a difference is found,
# the function logs the file name, the old and new cell values, and the row and column where the difference occurred.
# The function then calls another function called "prn_excel_diagrame" and passes in the logged data. Finally, the GUI
# window is destroyed and a message box appears to inform the user that the operation is completed, along with the time
# it took to complete the operation.

def extragere_informatii_diagrame():
    pbargui = Tk()
    pbargui.title("Extragere informatii din diagrame")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    dir_diagrame = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                           title="Selectati directorul cu diagramele:")
    start = time.time()
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_diagrame):
        if file_all.endswith(".xlsx"):
            file_counter = file_counter + 1
    if file_counter == 0:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showwarning('Eroare!', "Directorul selectat este gol.")
    array_informatii = []
    for file_all in os.listdir(dir_diagrame):
        if file_all.endswith(".xlsx"):
            file_progres = file_progres + 1
            statuslabel["text"] = str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 2
            pbargui.update_idletasks()
            wb1 = load_workbook(dir_diagrame + "/" + file_all)
            sheet1 = wb1.worksheets[0]
            for row in range(1, sheet1.max_row + 1):
                for col in range(1, sheet1.max_column + 1):
                    cell1 = sheet1.cell(row, col)
                    if cell1.value is not None:
                        array_informatii.append([file_all, cell1.value, row, col])
    array_informatii.insert(0, ["Nume Diagrama", "Text celula", "Rand", "Coloana"])
    prn_excel_infoindiagrame(array_informatii)
    pbar.destroy()
    pbargui.destroy()
    end = time.time()
    messagebox.showinfo('Finalizat!', "Prelucrate in " + str(end - start)[:6] + " secunde.")


# This is a Python script that extracts information from Excel charts. It allows the user to select a directory
# containing Excel chart files, and then extracts the text from each cell of each chart, along with the row and
# column numbers.
# The script uses the tkinter library to create a simple GUI with a progress bar and a status label. It then uses the
# os library to iterate over all files in the selected directory, and the openpyxl library to load each Excel file and
# access its worksheets and cells.
# The extracted information is stored in a two-dimensional array called array_informatii, which is then written to an
# Excel file using the prn_excel_infoindiagrame function.
# The script is well-documented and easy to follow. Overall, it provides a useful tool for quickly extracting
# information from Excel charts.

def crearematrixmodule():
    global headerrow
    pbargui = Tk()
    pbargui.title("Matrix module complet")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    array_output = [
        ["Drawing", "Place", "PB", "Group", "Module ID", "Name Normal", "Name ADR", "Conector", "Wire Print",
         "Wire Number", "XCode1", "Cavity1", "XCode2", "Cavity2", "All Supersleeve", "Module ID SuperSleeve",
         "Module ID"]]
    array_output_basic = [["Basic_Module", "Module", "Index", "Group", ]]

    dir_matrix = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                         title="Selectati directorul cu fisiere Matrix Module:")
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_matrix):
        if file_all.endswith(".xlsm"):
            file_counter = file_counter + 1
    if file_counter == 0:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showinfo("Fisier invalid", "Nu am gasit fisiere de prelucrat!")
        return None
    for file_all in os.listdir(dir_matrix):
        try:
            statuslabel["text"] = "Loading " + file_all + " . . . "
            pbar['value'] += 2
            pbargui.update_idletasks()
            if file_all.endswith(".xlsm"):
                place_index = 0
                pb_index = 0
                group_index = 0
                modid_index = 0
                normal_index = 0
                adr_index = 0
                con_index = 0
                wirep_index = 0
                wiren_index = 0
                xcode1_index = 0
                cavity1_index = 0
                xcode2_index = 0
                cavity2_index = 0
                allss_index = 0
                modidadr_index = 0
                modidnonadr_index = 0
                module1 = []
                module2 = []
                module3 = []
                module4 = []
                module5 = []
                module6 = []
                module7 = []
                module8 = []
                drwno = file_all[8:12]
                array_output_temp = []
                wb = load_workbook(dir_matrix + "/" + file_all, data_only=True)
                ws = wb.active
                for row in ws.iter_rows():
                    array_temp = []
                    for cell in row:
                        array_temp.append(cell.value)
                    array_output_temp.append(array_temp)
                for i in range(len(array_output_temp)):
                    pbar['value'] += 2
                    pbargui.update_idletasks()
                    if array_output_temp[i][0] == "№":
                        headerrow = i + 1
                        for x, y in enumerate(array_output_temp[i]):
                            if y == 'Місце' or y == 'Place':
                                place_index = x
                            elif y == '№ПБ/Ст.' or y == 'PB':
                                pb_index = x
                            elif y == 'GROUP' or y == 'Group':
                                group_index = x
                            elif y == 'Назва модуля':
                                modid_index = x
                            elif y == 'Name_Norm' or y == 'NAME_norm' or y == 'Name':
                                normal_index = x
                            elif y == 'Name_ADR' or y == 'NAME_AD':
                                adr_index = x
                            elif y == "Роз'єм":
                                con_index = x
                            elif y == "Провід":
                                wirep_index = x
                            elif y == "KN":
                                wiren_index = x
                            elif y == "XCode1":
                                xcode1_index = x
                            elif y == "Cavity1":
                                cavity1_index = x
                            elif y == "XCode2":
                                xcode2_index = x
                            elif y == "Cavity2":
                                cavity2_index = x
                            elif y == "ALL_Supersleeve":
                                allss_index = x
                            elif y == "Module_ADR" or y == "Adr" or y == "ADR":
                                modidadr_index = x
                            elif y == "Module_NonADR" or y == "NotADR" or y == "normal":
                                modidnonadr_index = x
                            elif y == "Module 1" or y == "Варіант_в'язки_1" or y == "Module1":
                                module1.append(x)
                            elif y == "Module 2" or y == "Варіант_в'язки_2" or y == "Module2":
                                module2.append(x)
                            elif y == "Module 3" or y == "Варіант_в'язки_3" or y == "Module3":
                                module3.append(x)
                            elif y == "Module 4" or y == "Варіант_в'язки_4" or y == "Module4":
                                module4.append(x)
                            elif y == "Module 5" or y == "Варіант_в'язки_5" or y == "Module5":
                                module5.append(x)
                            elif y == "Module 6" or y == "Варіант_в'язки_6" or y == "Module6":
                                module6.append(x)
                            elif y == "Module 7" or y == "Варіант_в'язки_7" or y == "Module7":
                                module7.append(x)
                            elif y == "Module 8" or y == "Варіант_в'язки_8" or y == "Module8":
                                module8.append(x)
                        break
                statuslabel["text"] = "Extracting informations . . ."
                pbar['value'] += 2
                pbargui.update_idletasks()
                for i in range(headerrow, len(array_output_temp)):
                    pbar['value'] += 2
                    pbargui.update_idletasks()
                    if array_output_temp[i][modidadr_index] is not None:
                        for element in array_output_temp[i][modidadr_index].split("/"):
                            # Module1 split
                            try:
                                if array_output_temp[i][module1[0]] is not None:
                                    for module in array_output_temp[i][module1[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module2 split
                            try:
                                if array_output_temp[i][module2[0]] is not None:
                                    for module in array_output_temp[i][module2[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module3 split
                            try:
                                if array_output_temp[i][module3[0]] is not None:
                                    for module in array_output_temp[i][module3[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module4 split
                            try:
                                if array_output_temp[i][module4[0]] is not None:
                                    for module in array_output_temp[i][module4[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module5 split
                            try:
                                if array_output_temp[i][module5[0]] is not None:
                                    for module in array_output_temp[i][module5[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module6 split
                            try:
                                if array_output_temp[i][module6[0]] is not None:
                                    for module in array_output_temp[i][module6[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module7 split
                            try:
                                if array_output_temp[i][module7[0]] is not None:
                                    for module in array_output_temp[i][module7[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module8 split
                            try:
                                if array_output_temp[i][module8[0]] is not None:
                                    for module in array_output_temp[i][module8[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                for i in range(headerrow, len(array_output_temp)):
                    pbar['value'] += 2
                    pbargui.update_idletasks()
                    if array_output_temp[i][modidnonadr_index] is not None:
                        for element in array_output_temp[i][modidnonadr_index].split("/"):
                            # Module1 split
                            try:
                                if array_output_temp[i][module1[0]] is not None:
                                    for module in array_output_temp[i][module1[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module2 split
                            try:
                                if array_output_temp[i][module2[0]] is not None:
                                    for module in array_output_temp[i][module2[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module3 split
                            try:
                                if array_output_temp[i][module3[0]] is not None:
                                    for module in array_output_temp[i][module3[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module4 split
                            try:
                                if array_output_temp[i][module4[0]] is not None:
                                    for module in array_output_temp[i][module4[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module5 split
                            try:
                                if array_output_temp[i][module5[0]] is not None:
                                    for module in array_output_temp[i][module5[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module6 split
                            try:
                                if array_output_temp[i][module6[0]] is not None:
                                    for module in array_output_temp[i][module6[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module7 split
                            try:
                                if array_output_temp[i][module7[0]] is not None:
                                    for module in array_output_temp[i][module7[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module8 split
                            try:
                                if array_output_temp[i][module8[0]] is not None:
                                    for module in array_output_temp[i][module8[0]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                for i in range(headerrow, len(array_output_temp)):
                    pbar['value'] += 2
                    pbargui.update_idletasks()
                    if array_output_temp[i][modidnonadr_index] is None and \
                            array_output_temp[i][modidadr_index] is None:
                        # Module1 split
                        try:
                            if array_output_temp[i][module1[0]] is not None:
                                for module in array_output_temp[i][module1[0]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue
                        # Module2 split
                        try:
                            if array_output_temp[i][module2[0]] is not None:
                                for module in array_output_temp[i][module2[0]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue
                        # Module3 split
                        try:
                            if array_output_temp[i][module3[0]] is not None:
                                for module in array_output_temp[i][module3[0]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue
                        # Module4 split
                        try:
                            if array_output_temp[i][module4[0]] is not None:
                                for module in array_output_temp[i][module4[0]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue
                        # Module5 split
                        try:
                            if array_output_temp[i][module5[0]] is not None:
                                for module in array_output_temp[i][module5[0]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue
                        # Module6 split
                        try:
                            if array_output_temp[i][module6[0]] is not None:
                                for module in array_output_temp[i][module6[0]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue
                        # Module7 split
                        try:
                            if array_output_temp[i][module7[0]] is not None:
                                for module in array_output_temp[i][module7[0]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue
                        # Module8 split
                        try:
                            if array_output_temp[i][module8[0]] is not None:
                                for module in array_output_temp[i][module8[0]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue

                # MY23
                drwno = file_all[8:12] + "MY23"
                for i in range(headerrow, len(array_output_temp)):
                    pbar['value'] += 2
                    pbargui.update_idletasks()
                    if array_output_temp[i][modidadr_index] is not None:
                        for element in array_output_temp[i][modidadr_index].split("/"):
                            # Module1 split
                            try:
                                if array_output_temp[i][module1[1]] is not None:
                                    for module in array_output_temp[i][module1[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module2 split
                            try:
                                if array_output_temp[i][module2[1]] is not None:
                                    for module in array_output_temp[i][module2[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module3 split
                            try:
                                if array_output_temp[i][module3[1]] is not None:
                                    for module in array_output_temp[i][module3[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module4 split
                            try:
                                if array_output_temp[i][module4[1]] is not None:
                                    for module in array_output_temp[i][module4[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module5 split
                            try:
                                if array_output_temp[i][module5[1]] is not None:
                                    for module in array_output_temp[i][module5[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module6 split
                            try:
                                if array_output_temp[i][module6[1]] is not None:
                                    for module in array_output_temp[i][module6[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module7 split
                            try:
                                if array_output_temp[i][module7[1]] is not None:
                                    for module in array_output_temp[i][module7[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module8 split
                            try:
                                if array_output_temp[i][module8[1]] is not None:
                                    for module in array_output_temp[i][module8[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                for i in range(headerrow, len(array_output_temp)):
                    pbar['value'] += 2
                    pbargui.update_idletasks()
                    if array_output_temp[i][modidnonadr_index] is not None:
                        for element in array_output_temp[i][modidnonadr_index].split("/"):
                            # Module1 split
                            try:
                                if array_output_temp[i][module1[1]] is not None:
                                    for module in array_output_temp[i][module1[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module2 split
                            try:
                                if array_output_temp[i][module2[1]] is not None:
                                    for module in array_output_temp[i][module2[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module3 split
                            try:
                                if array_output_temp[i][module3[1]] is not None:
                                    for module in array_output_temp[i][module3[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module4 split
                            try:
                                if array_output_temp[i][module4[1]] is not None:
                                    for module in array_output_temp[i][module4[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module5 split
                            try:
                                if array_output_temp[i][module5[1]] is not None:
                                    for module in array_output_temp[i][module5[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module6 split
                            try:
                                if array_output_temp[i][module6[1]] is not None:
                                    for module in array_output_temp[i][module6[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module7 split
                            try:
                                if array_output_temp[i][module7[1]] is not None:
                                    for module in array_output_temp[i][module7[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                            # Module8 split
                            try:
                                if array_output_temp[i][module8[1]] is not None:
                                    for module in array_output_temp[i][module8[1]].split("/"):
                                        array_output.append([drwno, array_output_temp[i][place_index],
                                                             array_output_temp[i][pb_index],
                                                             array_output_temp[i][group_index],
                                                             array_output_temp[i][modid_index],
                                                             array_output_temp[i][normal_index],
                                                             array_output_temp[i][adr_index],
                                                             array_output_temp[i][con_index],
                                                             array_output_temp[i][wirep_index],
                                                             array_output_temp[i][wiren_index],
                                                             array_output_temp[i][xcode1_index],
                                                             array_output_temp[i][cavity1_index],
                                                             array_output_temp[i][xcode2_index],
                                                             array_output_temp[i][cavity2_index],
                                                             array_output_temp[i][allss_index], element, module])
                                else:
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, ""])
                            except IndexError:
                                continue
                for i in range(headerrow, len(array_output_temp)):
                    pbar['value'] += 2
                    pbargui.update_idletasks()
                    if array_output_temp[i][modidnonadr_index] is None and \
                            array_output_temp[i][modidadr_index] is None:
                        # Module1 split
                        try:
                            if array_output_temp[i][module1[1]] is not None:
                                for module in array_output_temp[i][module1[1]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue
                        # Module2 split
                        try:
                            if array_output_temp[i][module2[1]] is not None:
                                for module in array_output_temp[i][module2[1]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue
                        # Module3 split
                        try:
                            if array_output_temp[i][module3[1]] is not None:
                                for module in array_output_temp[i][module3[1]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue
                        # Module4 split
                        try:
                            if array_output_temp[i][module4[1]] is not None:
                                for module in array_output_temp[i][module4[1]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue
                        # Module5 split
                        try:
                            if array_output_temp[i][module5[1]] is not None:
                                for module in array_output_temp[i][module5[1]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue
                        # Module6 split
                        try:
                            if array_output_temp[i][module6[1]] is not None:
                                for module in array_output_temp[i][module6[1]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue
                        # Module7 split
                        try:
                            if array_output_temp[i][module7[1]] is not None:
                                for module in array_output_temp[i][module7[1]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue
                        # Module8 split
                        try:
                            if array_output_temp[i][module8[1]] is not None:
                                for module in array_output_temp[i][module8[1]].split("/"):
                                    array_output.append([drwno, array_output_temp[i][place_index],
                                                         array_output_temp[i][pb_index],
                                                         array_output_temp[i][group_index],
                                                         array_output_temp[i][modid_index],
                                                         array_output_temp[i][normal_index],
                                                         array_output_temp[i][adr_index],
                                                         array_output_temp[i][con_index],
                                                         array_output_temp[i][wirep_index],
                                                         array_output_temp[i][wiren_index],
                                                         array_output_temp[i][xcode1_index],
                                                         array_output_temp[i][cavity1_index],
                                                         array_output_temp[i][xcode2_index],
                                                         array_output_temp[i][cavity2_index],
                                                         array_output_temp[i][allss_index], element, module])
                            else:
                                array_output.append([drwno, array_output_temp[i][place_index],
                                                     array_output_temp[i][pb_index],
                                                     array_output_temp[i][group_index],
                                                     array_output_temp[i][modid_index],
                                                     array_output_temp[i][normal_index],
                                                     array_output_temp[i][adr_index],
                                                     array_output_temp[i][con_index],
                                                     array_output_temp[i][wirep_index],
                                                     array_output_temp[i][wiren_index],
                                                     array_output_temp[i][xcode1_index],
                                                     array_output_temp[i][cavity1_index],
                                                     array_output_temp[i][xcode2_index],
                                                     array_output_temp[i][cavity2_index],
                                                     array_output_temp[i][allss_index], element, ""])
                        except IndexError:
                            continue


        except PermissionError:
            messagebox.showerror('Eroare scriere', "Fisierul " + file_all + "este read-only!")
            quit()
    statuslabel["text"] = "Printing EXCEL file . . ."
    pbar['value'] += 2
    pbargui.update_idletasks()
    array_output2 = [array_output[i] for i in range(len(array_output)) if array_output[i][-1] != ""]
    prn_excel_matrixmodule(array_output2)
    pbargui.destroy()
    end = time.time()
    messagebox.showinfo('Finalizat!', 'Prelucrate  fisiere.')


def crearebasicmodule():
    global headerrow
    pbargui = Tk()
    pbargui.title("Matrix module complet")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    array_output_basic = [["Basic_Module", "Module", "Index", "Group", "Place (delete)"]]
    dir_matrix = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                         title="Selectati directorul cu fisiere Matrix Module:")
    file_counter = 0
    file_progres = 0

    for file_all in os.listdir(dir_matrix):
        if file_all.endswith(".xlsm"):
            file_counter = file_counter + 1
    if file_counter == 0:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showinfo("Fisier invalid", "Nu am gasit fisiere de prelucrat!")
        return None
    bmgroup = 0
    for file_all in os.listdir(dir_matrix):
        try:
            statuslabel["text"] = "Loading " + file_all + " . . . "
            pbar['value'] += 2
            pbargui.update_idletasks()
            if file_all.endswith(".xlsm"):
                place_index = 0
                pb_index = 0
                group_index = 0
                modid_index = 0
                normal_index = 0
                adr_index = 0
                con_index = 0
                wirep_index = 0
                wiren_index = 0
                xcode1_index = 0
                cavity1_index = 0
                xcode2_index = 0
                cavity2_index = 0
                allss_index = 0
                modidadr_index = 0
                modidnonadr_index = 0
                module1 = []
                module2 = []
                module3 = []
                module4 = []
                module5 = []
                module6 = []
                module7 = []
                module8 = []
                drwno = file_all[8:12]
                array_output_temp = []
                wb = load_workbook(dir_matrix + "/" + file_all, data_only=True)
                ws = wb.active
                for row in ws.iter_rows():
                    array_temp = []
                    for cell in row:
                        array_temp.append(cell.value)
                    array_output_temp.append(array_temp)
                for i in range(len(array_output_temp)):
                    pbar['value'] += 2
                    pbargui.update_idletasks()
                    if array_output_temp[i][0] == "№":
                        headerrow = i + 1
                        for x, y in enumerate(array_output_temp[i]):
                            if y == 'Місце' or y == 'Place':
                                place_index = x
                            elif y == '№ПБ/Ст.' or y == 'PB':
                                pb_index = x
                            elif y == 'GROUP' or y == 'Group':
                                group_index = x
                            elif y == 'Назва модуля':
                                modid_index = x
                            elif y == 'Name_Norm' or y == 'NAME_norm' or y == 'Name':
                                normal_index = x
                            elif y == 'Name_ADR' or y == 'NAME_AD':
                                adr_index = x
                            elif y == "Роз'єм":
                                con_index = x
                            elif y == "Провід":
                                wirep_index = x
                            elif y == "KN":
                                wiren_index = x
                            elif y == "XCode1":
                                xcode1_index = x
                            elif y == "Cavity1":
                                cavity1_index = x
                            elif y == "XCode2":
                                xcode2_index = x
                            elif y == "Cavity2":
                                cavity2_index = x
                            elif y == "ALL_Supersleeve":
                                allss_index = x
                            elif y == "Module_ADR" or y == "Adr" or y == "ADR":
                                modidadr_index = x
                            elif y == "Module_NonADR" or y == "NotADR" or y == "normal":
                                modidnonadr_index = x
                            elif y == "Module 1" or y == "Варіант_в'язки_1" or y == "Module1":
                                module1.append(x)
                            elif y == "Module 2" or y == "Варіант_в'язки_2" or y == "Module2":
                                module2.append(x)
                            elif y == "Module 3" or y == "Варіант_в'язки_3" or y == "Module3":
                                module3.append(x)
                            elif y == "Module 4" or y == "Варіант_в'язки_4" or y == "Module4":
                                module4.append(x)
                            elif y == "Module 5" or y == "Варіант_в'язки_5" or y == "Module5":
                                module5.append(x)
                            elif y == "Module 6" or y == "Варіант_в'язки_6" or y == "Module6":
                                module6.append(x)
                            elif y == "Module 7" or y == "Варіант_в'язки_7" or y == "Module7":
                                module7.append(x)
                            elif y == "Module 8" or y == "Варіант_в'язки_8" or y == "Module8":
                                module8.append(x)
                        break
                for i in range(headerrow, len(array_output_temp)):
                    pbar['value'] += 2
                    pbargui.update_idletasks()
                    if array_output_temp[i][modid_index] is not None:
                        bmgroup = bmgroup + 1
                        if array_output_temp[i][modidadr_index] is None and \
                                array_output_temp[i][modidnonadr_index] is None:
                            try:
                                if array_output_temp[i][module1[0]] is not None:
                                    for module in array_output_temp[i][module1[1]].split("/"):
                                        array_output_basic.append([array_output_temp[i][modid_index], module, 1,
                                                                   bmgroup, array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module1[1]] is not None:
                                    for module in array_output_temp[i][module1[1]].split("/"):
                                        array_output_basic.append([array_output_temp[i][modid_index], module, 1,
                                                                   bmgroup, array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                        # ADR adr_index normal_index
                        elif array_output_temp[i][modidadr_index] is not None and \
                                array_output_temp[i][modidnonadr_index] is not None:
                            for ssmodule in array_output_temp[i][modidadr_index].split("/"):
                                array_output_basic.append([array_output_temp[i][adr_index], ssmodule, 1, bmgroup,
                                                           array_output_temp[i][place_index]])
                            for ssmodule in array_output_temp[i][modidnonadr_index].split("/"):
                                array_output_basic.append([array_output_temp[i][normal_index], ssmodule, 1, bmgroup,
                                                           array_output_temp[i][place_index]])
                            try:
                                if array_output_temp[i][module1[0]] is not None:
                                    for module in array_output_temp[i][module1[0]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 2, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append([array_output_temp[i][normal_index], module, 2,
                                                                   bmgroup, array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module1[1]] is not None:
                                    for module in array_output_temp[i][module1[1]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 2, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module2[0]] is not None:
                                    for module in array_output_temp[i][module2[0]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 3, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module2[1]] is not None:
                                    for module in array_output_temp[i][module2[1]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 3, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module3[0]] is not None:
                                    for module in array_output_temp[i][module3[0]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 4, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module3[1]] is not None:
                                    for module in array_output_temp[i][module3[1]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 4, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module4[0]] is not None:
                                    for module in array_output_temp[i][module4[0]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 5, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module4[1]] is not None:
                                    for module in array_output_temp[i][module4[1]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 5, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module5[0]] is not None:
                                    for module in array_output_temp[i][module5[0]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 6, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module5[1]] is not None:
                                    for module in array_output_temp[i][module5[1]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 6, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module6[0]] is not None:
                                    for module in array_output_temp[i][module6[0]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 7, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module6[1]] is not None:
                                    for module in array_output_temp[i][module6[1]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 7, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module7[0]] is not None:
                                    for module in array_output_temp[i][module7[0]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 8, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module7[1]] is not None:
                                    for module in array_output_temp[i][module7[1]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 8, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module8[0]] is not None:
                                    for module in array_output_temp[i][module8[0]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 9, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module8[1]] is not None:
                                    for module in array_output_temp[i][module8[1]].split("/"):
                                        array_output_basic.append([array_output_temp[i][adr_index], module, 9, bmgroup,
                                                                   array_output_temp[i][place_index]])
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                        # ADR adr_index normal_index
                        elif array_output_temp[i][modidadr_index] is None and \
                                array_output_temp[i][modidnonadr_index] is not None:
                            for ssmodule in array_output_temp[i][modidnonadr_index].split("/"):
                                array_output_basic.append(
                                    [array_output_temp[i][normal_index], ssmodule, 1, bmgroup,
                                     array_output_temp[i][place_index]])
                            try:
                                if array_output_temp[i][module1[0]] is not None:
                                    for module in array_output_temp[i][module1[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module1[1]] is not None:
                                    for module in array_output_temp[i][module1[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module2[0]] is not None:
                                    for module in array_output_temp[i][module2[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module2[1]] is not None:
                                    for module in array_output_temp[i][module2[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module3[0]] is not None:
                                    for module in array_output_temp[i][module3[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module3[1]] is not None:
                                    for module in array_output_temp[i][module3[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module4[0]] is not None:
                                    for module in array_output_temp[i][module4[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module4[1]] is not None:
                                    for module in array_output_temp[i][module4[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module5[0]] is not None:
                                    for module in array_output_temp[i][module5[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module5[1]] is not None:
                                    for module in array_output_temp[i][module5[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module6[0]] is not None:
                                    for module in array_output_temp[i][module6[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module6[1]] is not None:
                                    for module in array_output_temp[i][module6[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module7[0]] is not None:
                                    for module in array_output_temp[i][module7[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module7[1]] is not None:
                                    for module in array_output_temp[i][module7[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module8[0]] is not None:
                                    for module in array_output_temp[i][module8[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module8[1]] is not None:
                                    for module in array_output_temp[i][module8[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][normal_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                        # ADR adr_index normal_index
                        elif array_output_temp[i][modidadr_index] is not None and \
                                array_output_temp[i][modidnonadr_index] is None:
                            for ssmodule in array_output_temp[i][modidadr_index].split("/"):
                                array_output_basic.append(
                                    [array_output_temp[i][adr_index], ssmodule, 1, bmgroup,
                                     array_output_temp[i][place_index]])
                            try:
                                if array_output_temp[i][module1[0]] is not None:
                                    for module in array_output_temp[i][module1[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module1[1]] is not None:
                                    for module in array_output_temp[i][module1[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module2[0]] is not None:
                                    for module in array_output_temp[i][module2[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module2[1]] is not None:
                                    for module in array_output_temp[i][module2[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module3[0]] is not None:
                                    for module in array_output_temp[i][module3[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module3[1]] is not None:
                                    for module in array_output_temp[i][module3[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module4[0]] is not None:
                                    for module in array_output_temp[i][module4[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module4[1]] is not None:
                                    for module in array_output_temp[i][module4[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module5[0]] is not None:
                                    for module in array_output_temp[i][module5[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module5[1]] is not None:
                                    for module in array_output_temp[i][module5[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module6[0]] is not None:
                                    for module in array_output_temp[i][module6[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module6[1]] is not None:
                                    for module in array_output_temp[i][module6[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module7[0]] is not None:
                                    for module in array_output_temp[i][module7[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module7[1]] is not None:
                                    for module in array_output_temp[i][module7[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module8[0]] is not None:
                                    for module in array_output_temp[i][module8[0]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                            try:
                                if array_output_temp[i][module8[1]] is not None:
                                    for module in array_output_temp[i][module8[1]].split("/"):
                                        array_output_basic.append(
                                            [array_output_temp[i][adr_index], module, 2, bmgroup,
                                             array_output_temp[i][place_index]])
                            except IndexError:
                                continue
                statuslabel["text"] = "Extracting informations . . ."
                pbar['value'] += 2
                pbargui.update_idletasks()
        except PermissionError:
            messagebox.showerror('Eroare scriere', "Fisierul " + file_all + "este read-only!")
            quit()
    statuslabel["text"] = "Printing EXCEL file . . ."
    pbar['value'] += 2
    pbargui.update_idletasks()
    prn_excel_bmmodule(array_output_basic)
    pbargui.destroy()
    end = time.time()
    messagebox.showinfo('Finalizat!', 'Prelucrate  fisiere.')


def digrame_in_ksk():
    pbargui = Tk()
    pbargui.title("Lista diagrame in KSK")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    array_print = []
    basic_module = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                              title="Selectati fisierul Basic Module:")
    statuslabel["text"] = "Loading Basic Module file         "
    pbar['value'] += 2
    pbargui.update_idletasks()
    with open(basic_module, newline='') as csvfile:
        array_bm = list(csv.reader(csvfile, delimiter=';'))
    if array_bm[0][0] != "Basic_Module":
        messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect. Eroare cap de tabel!')
        pbar.destroy()
        pbargui.destroy()
        return
    ksk = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                     title="Selectati fisierul KSK:")
    nume_fisier = os.path.splitext(os.path.basename(ksk))[0]
    statuslabel["text"] = "Loading KSK file         "
    pbar['value'] += 2
    pbargui.update_idletasks()
    with open(ksk, newline='') as csvfile:
        array_ksk = list(csv.reader(csvfile, delimiter=';'))
    if array_ksk[0][0] != "Harness" and array_ksk[0][0] != "Module":
        messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect. Eroare cap de tabel!')
        pbar.destroy()
        pbargui.destroy()
        return
    statuslabel["text"] = "Loading corespondence list         "
    pbar['value'] += 2
    pbargui.update_idletasks()
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Corespondenta.txt", newline='') as csvfile:
        arr_corespondenta = list(csv.reader(csvfile, delimiter=';'))
    statuslabel["text"] = "Searching for diagrams         "
    pbar['value'] += 2
    pbargui.update_idletasks()
    bm_groups = []
    for i in range(1, len(array_bm)):
        if array_bm[i][3] not in bm_groups:
            bm_groups.append(array_bm[i][3])
            pbar['value'] += 2
            pbargui.update_idletasks()
    module_ksk = []
    statuslabel["text"] = "Searching for corespondence         "
    pbar['value'] += 2
    pbargui.update_idletasks()
    for i in range(1, len(array_ksk)):
        for x in range(1, len(arr_corespondenta)):
            if array_ksk[i][1] == arr_corespondenta[x][0]:
                module_ksk.append(arr_corespondenta[x][1])
                pbar['value'] += 2
                pbargui.update_idletasks()
    array_bm_ksk = []
    statuslabel["text"] = "Data analisys in progress         "
    pbar['value'] += 2
    pbargui.update_idletasks()
    array_bm[0].append("Indice")
    for i in range(1, len(array_bm)):
        if array_bm[i][1] in module_ksk:
            array_bm[i].append(1)
        else:
            array_bm[i].append(0)
    groups_unique = []
    for i in range(1, len(array_bm)):
        if array_bm[i][3] not in groups_unique:
            groups_unique.append(array_bm[i][3])
            pbar['value'] += 2
            pbargui.update_idletasks()
    diagrame_ksk =[]
    for q in range(len(groups_unique)):
        diagrame_basicgroup = []
        for w in range(len(array_bm)):
            if array_bm[w][3] == groups_unique[q] and array_bm[w][0] not in diagrame_basicgroup:
                diagrame_basicgroup.append(array_bm[w][0])
        for e in range(len(diagrame_basicgroup)):
            array_diagrama = []
            for r in range(len(array_bm)):
                if diagrame_basicgroup[e] == array_bm[r][0]:
                    array_diagrama.append(array_bm[r])
            index_unique = []
            counter_index = 0
            for z in range(len(array_diagrama)):
                counter_index = counter_index + array_diagrama[z][5]
                if array_diagrama[z][2] not in index_unique:
                    index_unique.append(array_diagrama[z][2])
                    pbar['value'] += 2
                    pbargui.update_idletasks()
            if len(index_unique) == counter_index:
                diagrame_ksk.append(array_diagrama[0][0])
    prn_excel_diagrameinksk(diagrame_ksk, nume_fisier)
    pbar.destroy()
    pbargui.destroy()
    end = time.time()
    messagebox.showinfo('Finalizat!', 'Done')
