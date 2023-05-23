import csv
import os
import sqlite3
from tkinter import *
from tkinter import messagebox, filedialog
import pandas as pd
from openpyxl.reader.excel import load_workbook
from functii_print import prn_excel_export_database


def databasecontent():
    # Create your connection.
    try:
        cnx = sqlite3.connect("//SVRO8FILE01/Groups/General/EFI/DBMAN/database.db")
    except sqlite3.OperationalError:
        cnx = sqlite3.connect(os.path.abspath(os.curdir) + "/MAN/Input/Others/database.db")
        messagebox.showinfo("Local database", "Network database unavailable. Using local database.")
    try:
        df = pd.read_sql_query("SELECT * FROM KSKDatabase", cnx)
    except:
        messagebox.showerror("Eroare baza de date", "Nici o baza de date nu este disponibila")
        return None

    listaallksk = df['KSKNo'].unique()

    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Sortare Module.txt", newline='') as csvfile:
        array_sortare_module = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Module Active.txt", newline='') as csvfile:
        array_module_active = list(csv.reader(csvfile, delimiter=';'))

    def scankey(event):
        val = event.widget.get()
        if val == '':
            data = list1
        else:
            data = []
            for item in list1:
                if val.lower() in item.lower():
                    data.append(item)
        update(data)

    def update(data):
        datalivrare_lb.delete(0, 'end')
        for item in data:
            datalivrare_lb.insert('end', item)

    list1 = df.DataLivrare.unique()

    def scankey2(event):
        val = event.widget.get()
        if val == '':
            data = list2
        else:
            data = []
            for item in list2:
                if val.lower() in item.lower():
                    data.append(item)
        update2(data)

    def scankey3(event):
        val = event.widget.get()
        if val == '':
            data = list3
        else:
            data = []
            for item in list3:
                if val.lower() in item.lower():
                    data.append(item)
        update3(data)

    def update2(data):
        datajit_lb.delete(0, 'end')
        for item in data:
            datajit_lb.insert('end', item)

    list2 = df.DataJIT.unique()

    def update3(data):
        ksk_lb.delete(0, 'end')
        # put new data
        for item in data:
            ksk_lb.insert('end', item)

    list3 = listaallksk

    def select_all():
        ksk_lb.select_set(0, END)

    def deselect_all():
        ksk_lb.selection_clear(0, END)

    def searchksk():
        valuesdatalivrare_lb = [datalivrare_lb.get(idx) for idx in datalivrare_lb.curselection()]
        valuesdatajit_lb = [datajit_lb.get(idx) for idx in datajit_lb.curselection()]
        if len(valuesdatajit_lb) > 0 and len(valuesdatalivrare_lb) > 0:
            xxx = df.query('DataLivrare in @valuesdatalivrare_lb')
            yyy = xxx.query('DataJIT in @valuesdatajit_lb')
            listaallksknew = yyy['KSKNo'].unique()
        elif len(valuesdatajit_lb) > 0 and len(valuesdatalivrare_lb) == 0:
            yyy = df.query('DataJIT in @valuesdatajit_lb')
            listaallksknew = yyy['KSKNo'].unique()
        elif len(valuesdatajit_lb) == 0 and len(valuesdatalivrare_lb) > 0:
            yyy = df.query('DataLivrare in @valuesdatalivrare_lb')
            listaallksknew = yyy['KSKNo'].unique()
        update3(listaallksknew)

    def exportksk():
        ksklist = [ksk_lb.get(idx) for idx in ksk_lb.curselection()]
        for ksk in ksklist:
            exp = df.loc[df['KSKNo'] == ksk]
            moduleksk = list(exp.iloc[0, 9].split(";"))
            trailerno = exp.iloc[0, 7]
            data = exp.iloc[0, 4]

            for item in moduleksk:
                if item in array_sortare_module[0]:
                    tip = "8000"
                    break
                elif item in array_sortare_module[1]:
                    tip = "8011"
                    break
                elif item in array_sortare_module[2]:
                    tip = "8023"
                    break
                else:
                    tip = "Necunoscut"
            arraywrite = [['Harness', 'Module', 'Side', 'Quantity', tip, "Date", "Time", "Trailer No"]]
            for module in moduleksk:
                for i in range(len(array_module_active)):
                    if array_module_active[i][0] == module and "LHD" in array_module_active[i][3]:
                        arraywrite.append([ksk, module, "BODYL", 1, "PC", data, data, trailerno])
                    elif array_module_active[i][0] == module and "RHD" in array_module_active[i][3]:
                        arraywrite.append([ksk, module, "BODYR", 1, "PC", data, data, trailerno])
                    elif array_module_active[i][0] == module:
                        arraywrite.append([ksk, module, "XXXX", 1, "PC", data, data, trailerno])
            with open(os.path.abspath(os.curdir) + "/MAN/Output/Database/KSK Export/" + ksk + ".csv", 'w', newline='',
                      encoding='utf-8') as myfile:
                wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
                wr.writerows(arraywrite)
        messagebox.showinfo('Finalizat!')

    def exportall():
        for ksk in listaallksk:
            exp = df.loc[df['KSKNo'] == ksk]
            moduleksk = list(exp.iloc[0, 9].split(";"))
            trailerno = exp.iloc[0, 7]
            data = exp.iloc[0, 4]

            for item in moduleksk:
                if item in array_sortare_module[0]:
                    tip = "8000"
                    break
                elif item in array_sortare_module[1]:
                    tip = "8011"
                    break
                elif item in array_sortare_module[2]:
                    tip = "8023"
                    break
                else:
                    tip = "Necunoscut"
            arraywrite = [['Harness', 'Module', 'Side', 'Quantity', tip, "Date", "Time", "Trailer No"]]
            for module in moduleksk:
                for i in range(len(array_module_active)):
                    if array_module_active[i][0] == module and "LHD" in array_module_active[i][3]:
                        arraywrite.append([ksk, module, "BODYL", 1, "PC", data, data, trailerno])
                    elif array_module_active[i][0] == module and "RHD" in array_module_active[i][3]:
                        arraywrite.append([ksk, module, "BODYR", 1, "PC", data, data, trailerno])
            with open(os.path.abspath(os.curdir) + "/MAN/Output/Database/KSK Export/" + ksk + ".csv", 'w', newline='',
                      encoding='utf-8') as myfile:
                wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
                wr.writerows(arraywrite)
        messagebox.showinfo('Finalizat!')

    def exportlist():
        exportlist = []
        file_load = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir) + "/MAN/Input/Module Files",
                                               title="Incarcati fisierul lista de KSK-uri:")
        wb = load_workbook(file_load)
        ws1 = wb.active
        for row in ws1['A']:
            if row.value is not None:
                exportlist.append(ws1.cell(row=row.row, column=1).value)
        for ksk in exportlist:
            try:
                exp = df.loc[df['KSKNo'] == ksk]
                moduleksk = list(exp.iloc[0, 9].split(";"))
                trailerno = exp.iloc[0, 7]
                data = exp.iloc[0, 4]

                for item in moduleksk:
                    if item in array_sortare_module[0]:
                        tip = "8000"
                        break
                    elif item in array_sortare_module[1]:
                        tip = "8011"
                        break
                    elif item in array_sortare_module[2]:
                        tip = "8023"
                        break
                    else:
                        tip = "Necunoscut"
                arraywrite = [['Harness', 'Module', 'Side', 'Quantity', tip, "Date", "Time", "Trailer No"]]
                for module in moduleksk:
                    for i in range(len(array_module_active)):
                        if array_module_active[i][0] == module and "LHD" in array_module_active[i][3]:
                            arraywrite.append([ksk, module, "BODYL", 1, "PC", data, data, trailerno])
                        elif array_module_active[i][0] == module and "RHD" in array_module_active[i][3]:
                            arraywrite.append([ksk, module, "BODYR", 1, "PC", data, data, trailerno])
                with open(os.path.abspath(os.curdir) + "/MAN/Output/Database/KSK Export/" + ksk + ".csv", 'w', newline='',
                          encoding='utf-8') as myfile:
                    wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
                    wr.writerows(arraywrite)
            except:
                continue
        messagebox.showinfo('Finalizat!')

    ws = Tk()
    ws.title("2022 MAN KSK Light reports")
    ws.geometry("+570+50")
    l1 = Label(ws, text="Data JIT")
    l2 = Label(ws, text="Data livrare")
    l3 = Label(ws, text="KSK number")

    l1.grid(row=0, column=0)
    l2.grid(row=0, column=1)
    l3.grid(row=0, column=2)
    entry = Entry(ws)
    entry.grid(row=1, column=1)
    entry.bind('<KeyRelease>', scankey)
    entry2 = Entry(ws)
    entry2.grid(row=1, column=0)
    entry2.bind('<KeyRelease>', scankey2)
    entry3 = Entry(ws)
    entry3.grid(row=1, column=2)
    entry3.bind('<KeyRelease>', scankey3)

    bsch1 = Button(ws, text="Search", command=searchksk)
    bsch1.grid(row=4, column=2)
    bsall = Button(ws, text="Select all", command=select_all)
    bsall.grid(row=5, column=2)
    bdsall = Button(ws, text="Deselect all", command=deselect_all)
    bdsall.grid(row=6, column=2)
    bexp = Button(ws, text="Export", command=exportksk)
    bexp.grid(row=4, column=3)
    bexpall = Button(ws, text="Export All", command=exportall)
    bexpall.grid(row=5, column=3)
    bexplist = Button(ws, text="Export List", command=exportlist)
    bexplist.grid(row=6, column=3)

    datalivrare_lb = Listbox(ws, exportselection=0, selectmode="multiple")
    datajit_lb = Listbox(ws, exportselection=0, selectmode="multiple")
    ksk_lb = Listbox(ws, exportselection=0, selectmode="multiple")

    datalivrare_lb.grid(row=2, column=1)
    datajit_lb.grid(row=2, column=0)
    ksk_lb.grid(row=2, column=2)
    update(list1)
    update2(list2)
    update3(list3)
    ws.mainloop()


def exportdatabase():
    # Create your connection.
    try:
        conn = sqlite3.connect("//SVRO8FILE01/Groups/General/EFI/DBMAN/database.db")
    except sqlite3.OperationalError:
        conn = sqlite3.connect(os.path.abspath(os.curdir) + "/MAN/Input/Others/database.db")
        messagebox.showinfo("Local database", "Network database unavailable. Using local database.")
    array_write = []
    c = conn.cursor()
    c.execute("select * from KSKDatabase")
    mysel = c.execute("select * from KSKDatabase ")
    for i, row in enumerate(mysel):
        for j, value in enumerate(row):
            array_write.append([i, j, row[j]])
    prn_excel_export_database(array_write)
    messagebox.showinfo('Finalizat!')


def database_delete_record():
    array_delete_rows = []
    fisier_delete = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                               title="Incarcati fisierul cu cablajele pe stock:")
    if len(fisier_delete) == 0:
        messagebox.showinfo("Nu ati selectat nimic")
        return None
    try:
        wb = load_workbook(fisier_delete)
    except:
        messagebox.showinfo("Fisier invalid", fisier_delete + " extensie incompatibila!")
        return None
    ws = wb.worksheets[0]
    for row in ws['A']:
        if row.value is not None:
            array_delete_rows.append(row.value)

    # Create your connection.
    try:
        conn = sqlite3.connect("//SVRO8FILE01/Groups/General/EFI/DBMAN/database.db")
    except sqlite3.OperationalError:
        conn = sqlite3.connect(os.path.abspath(os.curdir) + "/MAN/Input/Others/database.db")
        messagebox.showinfo("Local database", "Network database unavailable. Using local database.")
    counter = 0
    cursor = conn.cursor()

    for i in range(len(array_delete_rows)):
        counter += 1
        query_string = "DELETE from KSKDatabase where primarykey in (%s)" % ','.join(['?'] * len(array_delete_rows))
        cursor.execute(query_string, array_delete_rows)
        conn.commit()
    conn.close()
    messagebox.showinfo("Finalizat", "Sterse " + str(counter) + " inregistrari")


def database_delete_all_records():
    msg_box = messagebox.askquestion('Exit Application', 'Are you sure you want to delete all database records ?',
                                     icon='warning')
    if msg_box == 'yes':
        # Create your connection.
        try:
            conn = sqlite3.connect("//SVRO8FILE01/Groups/General/EFI/DBMAN/database.db")
        except sqlite3.OperationalError:
            conn = sqlite3.connect(os.path.abspath(os.curdir) + "/MAN/Input/Others/database.db")
            messagebox.showinfo("Local database", "Network database unavailable. Using local database.")
        cursor = conn.cursor()
        # delete all rows from table
        conn.execute('DELETE FROM KSKDatabase;', )

        conn.close()
        messagebox.showinfo("Finalizat", "Sterse toate inregistrarile")
    else:
        messagebox.showinfo('Return', 'You will now return to the application screen')