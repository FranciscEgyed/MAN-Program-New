import csv
import datetime
import difflib
import os
import time
from collections import Counter
from tkinter import messagebox, Tk, ttk, Label, HORIZONTAL, filedialog
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from diverse import log_file
from functii_print import prn_excel_variatii


def extragere_lungimi_ksk():
    pbargui = Tk()
    pbargui.title("Extragere lungimi KSK")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    save_time = datetime.datetime.now().strftime("%d.%m.%Y_%H.%M.%S")
    array_lungimi = [["KSK", "Dwg No.", "Module", "Ltg-Nr.", "Leitung", "Farbe", "Quer.", "Kurzname", "Pin", "Lange"]]
    counter = 0
    dir_salvare = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                          title="Selectati directorul pentru salvare")
    dir_prelucrare = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir) + "/MAN/Output/Excel Files/",
                                             title="Selectati directorul cu fisiere:")

    start = time.time()
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_prelucrare):
        if file_all.endswith(".xlsx") and not file_all.startswith("BOM"):
            file_counter = file_counter + 1
    for file_all in os.listdir(dir_prelucrare):
        try:
            if file_all.endswith(".xlsx") and not file_all.startswith("BOM"):
                counter = counter + 1
                wb = load_workbook(dir_prelucrare + "/" + file_all)
                ws0 = wb.worksheets[0]
                ws1 = wb.worksheets[1]
                ws2 = wb.worksheets[2]
                file_progres = file_progres + 1
                statuslabel["text"] = str(file_progres) + "/" + str(file_counter) + " : " + file_all
                pbar['value'] += 2
                pbargui.update_idletasks()
                for row in ws1['A']:
                    if row.value != "Harness":
                        array_lungimi.append([ws0.cell(row=2, column=1).value, ws1.cell(row=row.row, column=1).value,
                                              ws1.cell(row=row.row, column=2).value,
                                              ws1.cell(row=row.row, column=3).value,
                                              ws1.cell(row=row.row, column=4).value,
                                              ws1.cell(row=row.row, column=5).value,
                                              ws1.cell(row=row.row, column=6).value,
                                              ws1.cell(row=row.row, column=7).value,
                                              ws1.cell(row=row.row, column=8).value,
                                              ws1.cell(row=row.row, column=9).value])
                for row in ws2['A']:
                    if row.value != "Harness":
                        array_lungimi.append([ws0.cell(row=2, column=1).value, ws2.cell(row=row.row, column=1).value,
                                              ws2.cell(row=row.row, column=2).value,
                                              ws2.cell(row=row.row, column=3).value,
                                              ws2.cell(row=row.row, column=4).value,
                                              ws2.cell(row=row.row, column=5).value,
                                              ws2.cell(row=row.row, column=6).value,
                                              ws2.cell(row=row.row, column=7).value,
                                              ws2.cell(row=row.row, column=8).value,
                                              ws2.cell(row=row.row, column=9).value])
        except PermissionError:
            messagebox.showerror('Eroare scriere', "Fisierul " + file_all + "este read-only!")
            quit()

    wbs = Workbook()
    wss1 = wbs.active
    for i in range(len(array_lungimi)):
        for x in range(len(array_lungimi[i])):
            try:
                if "E-" in array_lungimi[i][x]:
                    try:
                        wss1.cell(column=x + 1, row=i + 1, value=str(array_lungimi[i][x]))
                    except:
                        wss1.cell(column=x + 1, row=i + 1, value=str(array_lungimi[i][x]))
                else:
                    try:
                        wss1.cell(column=x + 1, row=i + 1, value=float(array_lungimi[i][x]))
                    except:
                        wss1.cell(column=x + 1, row=i + 1, value=array_lungimi[i][x])
            except TypeError:
                wss1.cell(column=x + 1, row=i + 1, value=array_lungimi[i][x])
    if dir_salvare == "":
        try:
            wbs.save(
                os.path.abspath(os.curdir) + "/MAN/Output/Report Files/Export Lungimi " + save_time + ".xlsx")
            log_file("Creat Export Lungimi.xlsx")
        except PermissionError:
            log_file("Eroare scriere. Nu am salvat Export Lungimi " + save_time + ".xlsx")
            messagebox.showerror('Eroare scriere', "Fisierul este read-only!")
            quit()
    else:
        try:
            wbs.save(dir_salvare + "/Export Lungimi " + save_time + ".xlsx")
            log_file("Creat Export Lungimi.xlsx")
        except PermissionError:
            log_file("Eroare scriere. Nu am salvat Export Lungimi.xlsx")
            messagebox.showerror('Eroare scriere', "Fisierul este read-only!")
            quit()
    pbar.destroy()
    pbargui.destroy()
    end = time.time()
    messagebox.showinfo('Finalizat!', 'Prelucrate ' + str(counter) + " fisiere in " + str(end - start)[:6] + " secunde.")


def extragere_bom_ksk():
    pbargui = Tk()
    pbargui.title("Extragere BOM KSK")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    save_time = datetime.datetime.now().strftime("%d.%m.%Y_%H.%M.%S")
    array_bom = [["KSK", "Module", "Quantity", "Bezei", "VOBES-ID", "Benennung", "Verwendung", "Verwendung",
                  "Kurzname", "xy", "Teilenummer", "Vorzugsteil", "TAB-Nummer", "Referenzteil", "Farbe",
                  "E-Komponente", "E-Komponente Part-Nr.", "Einh."]]
    cc = []
    counter = 0
    dir_salvare = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                          title="Selectati directorul pentru salvare")
    dir_prelucrare = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir) + "/MAN/Output/Excel Files/",
                                          title="Selectati directorul cu fisiere:")
    start = time.time()
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_prelucrare):
        if file_all.endswith(".xlsx") and file_all.startswith("BOM"):
            file_counter = file_counter + 1
    for file_all in os.listdir(dir_prelucrare):
        try:
            if file_all.endswith(".xlsx") and file_all.startswith("BOM"):
                counter = counter + 1
                wb = load_workbook(dir_prelucrare + "/" + file_all)
                ws0 = wb.worksheets[0]
                ws1 = wb.worksheets[1]
                file_progres = file_progres + 1
                statuslabel["text"] = str(file_progres) + "/" + str(file_counter) + " : " + file_all
                pbar['value'] += 2
                pbargui.update_idletasks()
                for row in ws1['A']:
                    if row.value != "Module" and ws1.cell(row=row.row, column=2).value != 0:
                        array_bom.append([ws0.cell(row=2, column=1).value, ws1.cell(row=row.row, column=1).value,
                                          ws1.cell(row=row.row, column=2).value, ws1.cell(row=row.row, column=3).value,
                                          ws1.cell(row=row.row, column=4).value, ws1.cell(row=row.row, column=5).value,
                                          ws1.cell(row=row.row, column=6).value, ws1.cell(row=row.row, column=7).value,
                                          ws1.cell(row=row.row, column=8).value, ws1.cell(row=row.row, column=9).value,
                                          ws1.cell(row=row.row, column=10).value, ws1.cell(row=row.row, column=11).value,
                                          ws1.cell(row=row.row, column=12).value, ws1.cell(row=row.row, column=13).value,
                                          ws1.cell(row=row.row, column=14).value, ws1.cell(row=row.row, column=15).value,
                                          ws1.cell(row=row.row, column=16).value, ws1.cell(row=row.row, column=17).value])
        except PermissionError:
            messagebox.showerror('Eroare scriere', "Fisierul " + file_all + "este read-only!")
            quit()
    for i in range(1, len(array_bom)):
        ksk = ""
        partno = ""
        conectorno = ""
        ksk = array_bom[i][0]
        for x in range(3, len(array_bom[i])):
            if len(str(array_bom[i][x])) == 13 and str(array_bom[i][x]).find(".") == 2 and \
                    str(array_bom[i][x]).find("-") == 8:
                partno = array_bom[i][x]
                break
            else:
                partno = "-"
            for q in range(3, len(array_bom[i])):
                if str(array_bom[i][q]).startswith("X"):
                    conectorno = array_bom[i][q]
                    break
                else:
                    conectorno = "-"
        array_bom[i].append(ksk + partno + conectorno)
        cc.append(ksk + partno + conectorno)
    concatenare = list(Counter(cc).items())
    for i in range(1, len(array_bom)):
        for y in concatenare:
            if array_bom[i][18] == y[0]:
                array_bom[i].append(y[1])
    wbs = Workbook()
    wss1 = wbs.active
    for i in range(len(array_bom)):
        for x in range(len(array_bom[i])):
            try:
                if "E-" in array_bom[i][x]:
                    try:
                        wss1.cell(column=x + 1, row=i + 1, value=str(array_bom[i][x]))
                    except:
                        wss1.cell(column=x + 1, row=i + 1, value=str(array_bom[i][x]))
                else:
                    try:
                        wss1.cell(column=x + 1, row=i + 1, value=float(array_bom[i][x]))
                    except:
                        wss1.cell(column=x + 1, row=i + 1, value=array_bom[i][x])
            except TypeError:
                wss1.cell(column=x + 1, row=i + 1, value=array_bom[i][x])
    wss1.cell(column=19, row=1, value="Concatenare Ksk+Part+Conector")
    wss1.cell(column=20, row=1, value="Count Concatenare")
    pbar['value'] += 2
    pbargui.update_idletasks()
    if dir_salvare == "":
        try:
            wbs.save(
                os.path.abspath(os.curdir) + "/MAN/Output/Report Files/Export BOMs " + save_time + ".xlsx")
            log_file("Creat Export BOMs.xlsx")
        except PermissionError:
            log_file("Eroare scriere. Nu am salvat Export BOMs " + save_time + ".xlsx")
            messagebox.showerror('Eroare scriere', "Fisierul este read-only!")
            quit()
    else:
        try:
            wbs.save(dir_salvare + "/Export BOMs " + save_time + ".xlsx")
            log_file("Creat Export Lungimi.xlsx")
        except PermissionError:
            log_file("Eroare scriere. Nu am salvat Export BOMs.xlsx")
            messagebox.showerror('Eroare scriere', "Fisierul este read-only!")
            quit()
    pbar.destroy()
    pbargui.destroy()
    end = time.time()
    messagebox.showinfo('Finalizat!', 'Prelucrate ' + str(counter) + " fisiere in " + str(end - start)[:6] + " secunde.")


def extragere_variatii():
    pbargui = Tk()
    pbargui.title("Extragere variatii lungimi")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    start = time.time()
    counter = 0
    array_print = [["KSK", "Type", "Abbreviation", "Part_Number_2", "Side", "q-VLA", "r-VLA/RQT", "s-Radstand",	"t-NLA",
                    "u-Aoeberhang", "Heck module", "Identic"]]

    dir_selectat = os.path.abspath(os.curdir) + "/MAN/Output/Excel Files/8000/"
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".xlsx"):
            file_counter = file_counter + 1
    pbar['value'] = 0
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".xlsx"):
            counter = counter + 1
            file_progres = file_progres + 1
            statuslabel["text"] = "8000 = " + str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 1
            pbargui.update_idletasks()
            wb = load_workbook(dir_selectat + "/" + file_all)
            ws = wb.worksheets[3]
            for row in ws['A']:
                if row.value != "Type":
                    array_print.append([file_all[:-5], ws.cell(row=row.row, column=1).value,
                                        ws.cell(row=row.row, column=2).value, ws.cell(row=row.row, column=3).value,
                                        ws.cell(row=row.row, column=4).value, ws.cell(row=row.row, column=5).value,
                                        ws.cell(row=row.row, column=6).value, ws.cell(row=row.row, column=7).value,
                                        ws.cell(row=row.row, column=8).value, ws.cell(row=row.row, column=9).value,
                                        ws.cell(row=row.row, column=10).value, ws.cell(row=row.row, column=11).value])
            continue
        else:
            continue

    dir_selectat = os.path.abspath(os.curdir) + "/MAN/Output/Excel Files/8011/"
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".xlsx"):
            file_counter = file_counter + 1
    pbar['value'] = 0
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".xlsx"):
            counter = counter + 1
            file_progres = file_progres + 1
            statuslabel["text"] = "8011 = " + str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 1
            pbargui.update_idletasks()
            wb = load_workbook(dir_selectat + "/" + file_all)
            ws = wb.worksheets[3]
            for row in ws['A']:
                if row.value != "Type":
                    array_print.append([file_all[:-5], ws.cell(row=row.row, column=1).value,
                                        ws.cell(row=row.row, column=2).value, ws.cell(row=row.row, column=3).value,
                                        ws.cell(row=row.row, column=4).value, ws.cell(row=row.row, column=5).value,
                                        ws.cell(row=row.row, column=6).value, ws.cell(row=row.row, column=7).value,
                                        ws.cell(row=row.row, column=8).value, ws.cell(row=row.row, column=9).value,
                                        ws.cell(row=row.row, column=10).value, ws.cell(row=row.row, column=11).value])
            continue
        else:
            continue

    dir_selectat = os.path.abspath(os.curdir) + "/MAN/Output/Excel Files/8023/"
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".xlsx"):
            file_counter = file_counter + 1
    pbar['value'] = 0
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".xlsx"):
            counter = counter + 1
            file_progres = file_progres + 1
            statuslabel["text"] = "8023 = " + str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 1
            pbargui.update_idletasks()
            wb = load_workbook(dir_selectat + "/" + file_all)
            ws = wb.worksheets[3]
            for row in ws['A']:
                if row.value != "Type":
                    array_print.append([file_all[:-5], ws.cell(row=row.row, column=1).value,
                                        ws.cell(row=row.row, column=2).value, ws.cell(row=row.row, column=3).value,
                                        ws.cell(row=row.row, column=4).value, ws.cell(row=row.row, column=5).value,
                                        ws.cell(row=row.row, column=6).value, ws.cell(row=row.row, column=7).value,
                                        ws.cell(row=row.row, column=8).value, ws.cell(row=row.row, column=9).value,
                                        ws.cell(row=row.row, column=10).value, ws.cell(row=row.row, column=11).value])
            continue
        else:
            continue
    for i in range(len(array_print)):
        for x in range(len(array_print[i])):
            if array_print[i][x] is None:
                array_print[i][x] = ""
        else:
            continue
    pbar.destroy()
    pbargui.destroy()
    prn_excel_variatii(array_print)
    end = time.time()

    messagebox.showinfo('Finalizat!', 'Prelucrate ' + str(counter) + " fisiere in " + str(end - start)[:6] + " secunde.")


def comparatie_fisiere():

    def compare_files_in_folders(folder1, folder2):
        pbargui = Tk()
        pbargui.title("Comparatie fisiere ")
        pbargui.geometry("500x50+50+550")
        pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
        statuslabel = Label(pbargui, text="Waiting . . .")
        pbar.grid(row=1, column=1, padx=5, pady=5)
        statuslabel.grid(row=1, column=2, padx=5, pady=5)
        # Get the list of files in each folder
        files_folder1 = os.listdir(folder1)
        files_folder2 = os.listdir(folder2)

        # Get the common files (files with the same name) in both folders
        common_files = set(files_folder1) & set(files_folder2)

        # Initialize a dictionary to store comparison results
        comparison_results = {}

        # Compare the contents of common files
        for filename in common_files:
            statuslabel["text"] = filename
            pbar['value'] += 1
            pbargui.update_idletasks()
            file1_path = os.path.join(folder1, filename)
            file2_path = os.path.join(folder2, filename)

            with open(file1_path, 'r') as file1:
                content1 = file1.read()

            with open(file2_path, 'r') as file2:
                content2 = file2.read()

            diff_lines = list(difflib.ndiff(content1, content2))
            different_lines = [line[2:] for line in diff_lines if line.startswith('- ')]

            if not different_lines:
                comparison_results[filename] = "Contents are identical."
            else:
                comparison_results[filename] = "Contents are different."
                comparison_results[f"{filename}_diff"] = "".join(different_lines)
            pbar['value'] += 1
            pbargui.update_idletasks()
        # Check for files that exist in one folder but not the other
        for filename in set(files_folder1) ^ set(files_folder2):
            if filename in files_folder1:
                comparison_results[filename] = "File exists only in folder 1."
            else:
                comparison_results[filename] = "File exists only in folder 2."
        pbar.destroy()
        pbargui.destroy()
        return comparison_results

    # Example usage:
    folder1_path = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                                       title="Selectati directorul numarul unu")
    folder2_path = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                                       title="Selectati directorul numarul doi")
    start = time.time()
    results = compare_files_in_folders(folder1_path, folder2_path)
    output = []
    for filename, result in results.items():
        output.append([f"{filename}: {result}"])
    print(output)
    with open(os.path.abspath(os.curdir) + "/MAN/Diferente fisiere.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(output)

    end = time.time()

    messagebox.showinfo('Finalizat!',
                        'Prelucrate in ' + str(end - start)[:6] + " secunde.")