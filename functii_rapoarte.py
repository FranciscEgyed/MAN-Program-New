import os
import time
from tkinter import filedialog, messagebox, Tk, HORIZONTAL, Label, ttk
from openpyxl.reader.excel import load_workbook
import globale
from functii_print import prn_excel_raport
from functii_rapoarte_small import *


def creare_raport():
    globale.director_salvare_raport = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                                              title="Selectati directorul pentru salvare")
    fisier_de_prelucrat = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir) + "/MAN/Output/Excel Files/",
                                                     title="Selectati fisierul")
    raport_final = [["Release in production", "Call off", "Raw call off name", "KSK", "Klappschalle", "Super sleeve",
                     "Extra autarke modul", "Old/new modul not in the UA list", "Heck Module", "BKK", "X1555.1A1",
                     "Splice/wire", "Same wire no in harness", "X2799.1A1 or X2799.1A1_1", "XA.B129.1 or XA.B610.1",
                     "Modules not implemented", "Check", "comment", "Side", "Abbreviation", "q / W", "r / X", "s / Y",
                     "t", "u", "", "Side", "Abbreviation", "q / W", "r / X", "s / Y", "t", "u", "DoKa", "X6616/X6490",
                     "KSW module", "Military", "Arad Prüfung", "MY 2023", "X6490 Module", "Module check", "CKD",
                     "Delivery Date", "Trailer No", "Stvb RHM FHS/RHM"]]
    try:
        wb = load_workbook(fisier_de_prelucrat)
        ws1 = wb.worksheets[0]
        ws2 = wb.worksheets[1]
        ws3 = wb.worksheets[2]
        ws4 = wb.worksheets[3]
        ws5 = wb.worksheets[4]
        ws6 = wb.worksheets[5]
        ws7 = wb.worksheets[6]
        ws8 = wb.worksheets[7]
        raport_final.append(
            ["", "", "", file_namer(ws1), klappschaller(ws6, ws1), supersleever(ws1), extraautarker(ws1),
             oldnewcheckr(ws1), heckmoduler(ws4), bkkr(ws6), x1555r(ws2, ws3), splicewirer(ws5), samewirer(ws7),
             x2799r(ws2, ws3), xab6101r(ws2, ws3), module_implementater(ws1), "", verificarelungimir(ws6, ws4),
             copylenghtvaluesrightr(ws4)[0], copylenghtvaluesrightr(ws4)[1], copylenghtvaluesrightr(ws4)[2],
             copylenghtvaluesrightr(ws4)[3], copylenghtvaluesrightr(ws4)[4], copylenghtvaluesrightr(ws4)[5],
             copylenghtvaluesrightr(ws4)[6], "", copylenghtvaluesleftr(ws4)[0], copylenghtvaluesleftr(ws4)[1],
             copylenghtvaluesleftr(ws4)[2], copylenghtvaluesleftr(ws4)[3], copylenghtvaluesleftr(ws4)[4],
             copylenghtvaluesleftr(ws4)[5], copylenghtvaluesleftr(ws4)[6], dokar(ws1), x6616r(ws6), kswr(ws1),
             militaryr(ws1), prufungr(ws6, ws1, ws2, ws3), my2023r(ws6, ws1), x6616stvbr(ws1, ws8, ws6),
             module_check(ws1), ckd(ws1), delivery(ws1), ws1.cell(row=2, column=11).value, stvb(ws6)])
    except PermissionError:
        messagebox.showerror('Eroare scriere', "Fisierul " + fisier_de_prelucrat + "este read-only!")
        quit()

    verificare_raport(raport_final)
    prn_excel_raport(raport_final)
    messagebox.showinfo('Finalizat!', fisier_de_prelucrat)


def creare_raport_all():
    pbargui = Tk()
    pbargui.title("Progres . . . ")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    timelabel = Label(pbargui, text="Time . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    timelabel.grid(row=2, column=2)
    counter = 0
    globale.director_salvare_raport = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                                              title="Selectati directorul pentru salvare")
    raport_final = [["Release in production", "Call off", "Raw call off name", "KSK", "Klappschalle", "Super sleeve",
                     "Extra autarke modul", "Old/new modul not in the UA list", "Heck Module", "BKK", "X1555.1A1",
                     "Splice/wire", "Same wire no in harness", "X2799.1A1 or X2799.1A1_1", "XA.B129.1 or XA.B610.1",
                     "Modules not implemented", "Check", "comment", "Side", "Abbreviation", "q / W", "r / X", "s / Y",
                     "t", "u", "", "Side", "Abbreviation", "q / W", "r / X", "s / Y", "t", "u", "DoKa", "X6616/X6490",
                     "KSW module", "Military", "Arad Prüfung", "MY 2023", "X6490 Module", "Module check", "CKD",
                     "Delivery Date", "Trailer No", "Stvb RHM FHS/RHM"]]
    start = time.time()
    dir_prelucrare = os.path.abspath(os.curdir) + "/MAN/Output/Excel Files/8000"
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_prelucrare):
        if file_all.endswith(".xlsx") and not file_all.startswith("BOM"):
            file_counter = file_counter + 1
    start0 = time.time()
    for file_all in os.listdir(dir_prelucrare):
        try:
            if file_all.endswith(".xlsx") and not file_all.startswith("BOM"):
                counter = counter + 1
                wb = load_workbook(dir_prelucrare + "/" + file_all)
                ws1 = wb.worksheets[0]
                ws2 = wb.worksheets[1]
                ws3 = wb.worksheets[2]
                ws4 = wb.worksheets[3]
                ws5 = wb.worksheets[4]
                ws6 = wb.worksheets[5]
                ws7 = wb.worksheets[6]
                ws8 = wb.worksheets[7]
                raport_final.append(
                    ["", "", "", file_namer(ws1), klappschaller(ws6, ws1), supersleever(ws1), extraautarker(ws1),
                     oldnewcheckr(ws1), heckmoduler(ws4), bkkr(ws6), x1555r(ws2, ws3), splicewirer(ws5), samewirer(ws7),
                     x2799r(ws2, ws3), xab6101r(ws2, ws3), module_implementater(ws1), "", verificarelungimir(ws6, ws4),
                     copylenghtvaluesrightr(ws4)[0], copylenghtvaluesrightr(ws4)[1], copylenghtvaluesrightr(ws4)[2],
                     copylenghtvaluesrightr(ws4)[3], copylenghtvaluesrightr(ws4)[4], copylenghtvaluesrightr(ws4)[5],
                     copylenghtvaluesrightr(ws4)[6], "", copylenghtvaluesleftr(ws4)[0], copylenghtvaluesleftr(ws4)[1],
                     copylenghtvaluesleftr(ws4)[2], copylenghtvaluesleftr(ws4)[3], copylenghtvaluesleftr(ws4)[4],
                     copylenghtvaluesleftr(ws4)[5], copylenghtvaluesleftr(ws4)[6], dokar(ws1), x6616r(ws6), kswr(ws1),
                     militaryr(ws1), prufungr(ws6, ws1, ws2, ws3), my2023r(ws6, ws1), x6616stvbr(ws1, ws8, ws6),
                     module_check(ws1), ckd(ws1), delivery(ws1), ws1.cell(row=2, column=11).value, stvb(ws6)])
                end0 = time.time()
                file_progres = file_progres + 1
                statuslabel["text"] = "                 8000: " + str(file_progres) + "/" + str(file_counter) \
                                      + " : " + file_all + "    "
                timelabel["text"] = "Estimated time to complete : " + \
                                    str(((file_counter * 0.3) - (end0 - start0))/60)[:5] + " minutes."
                pbar['value'] += 2
                pbargui.update_idletasks()
        except PermissionError:
            messagebox.showerror('Eroare scriere', "Fisierul " + file_all + "este read-only!")
            quit()

    dir_prelucrare = os.path.abspath(os.curdir) + "/MAN/Output/Excel Files/8011"
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_prelucrare):
        if file_all.endswith(".xlsx") and not file_all.startswith("BOM"):
            file_counter = file_counter + 1
    start1 = time.time()
    for file_all in os.listdir(dir_prelucrare):
        try:
            if file_all.endswith(".xlsx") and not file_all.startswith("BOM"):
                counter = counter + 1
                wb = load_workbook(dir_prelucrare + "/" + file_all)
                ws1 = wb.worksheets[0]
                ws2 = wb.worksheets[1]
                ws3 = wb.worksheets[2]
                ws4 = wb.worksheets[3]
                ws5 = wb.worksheets[4]
                ws6 = wb.worksheets[5]
                ws7 = wb.worksheets[6]
                ws8 = wb.worksheets[7]
                raport_final.append(
                    ["", "", "", file_namer(ws1), klappschaller(ws6, ws1), supersleever(ws1), extraautarker(ws1),
                     oldnewcheckr(ws1), heckmoduler(ws4), bkkr(ws6), x1555r(ws2, ws3), splicewirer(ws5), samewirer(ws7),
                     x2799r(ws2, ws3), xab6101r(ws2, ws3), module_implementater(ws1), "", verificarelungimir(ws6, ws4),
                     copylenghtvaluesrightr(ws4)[0], copylenghtvaluesrightr(ws4)[1], copylenghtvaluesrightr(ws4)[2],
                     copylenghtvaluesrightr(ws4)[3], copylenghtvaluesrightr(ws4)[4], copylenghtvaluesrightr(ws4)[5],
                     copylenghtvaluesrightr(ws4)[6], "", copylenghtvaluesleftr(ws4)[0], copylenghtvaluesleftr(ws4)[1],
                     copylenghtvaluesleftr(ws4)[2], copylenghtvaluesleftr(ws4)[3], copylenghtvaluesleftr(ws4)[4],
                     copylenghtvaluesleftr(ws4)[5], copylenghtvaluesleftr(ws4)[6], dokar(ws1), x6616r(ws6), kswr(ws1),
                     militaryr(ws1), prufungr(ws6, ws1, ws2, ws3), my2023r(ws6, ws1), x6616stvbr(ws1, ws8, ws6),
                     module_check(ws1), ckd(ws1), delivery(ws1), ws1.cell(row=2, column=11).value, stvb(ws6)])
                end1 = time.time()
                file_progres = file_progres + 1
                statuslabel["text"] = "                 8011: " + str(file_progres) + "/" + \
                                      str(file_counter) + " : " + file_all + "    "
                timelabel["text"] = "Estimated time to complete : " + \
                                    str(((file_counter * 0.3) - (end1 - start1))/60)[:5] + " minutes."
                pbar['value'] += 2
                pbargui.update_idletasks()
        except PermissionError:
            messagebox.showerror('Eroare scriere', "Fisierul " + file_all + "este read-only!")
            quit()

    dir_prelucrare = os.path.abspath(os.curdir) + "/MAN/Output/Excel Files/8023"
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_prelucrare):
        if file_all.endswith(".xlsx") and not file_all.startswith("BOM"):
            file_counter = file_counter + 1
    start2 = time.time()
    for file_all in os.listdir(dir_prelucrare):
        try:
            if file_all.endswith(".xlsx") and not file_all.startswith("BOM"):
                counter = counter + 1
                wb = load_workbook(dir_prelucrare + "/" + file_all)
                ws1 = wb.worksheets[0]
                ws2 = wb.worksheets[1]
                ws3 = wb.worksheets[2]
                ws4 = wb.worksheets[3]
                ws5 = wb.worksheets[4]
                ws6 = wb.worksheets[5]
                ws7 = wb.worksheets[6]
                ws8 = wb.worksheets[7]
                raport_final.append(
                    ["", "", "", file_namer(ws1), klappschaller(ws6, ws1), supersleever(ws1), extraautarker(ws1),
                     oldnewcheckr(ws1), heckmoduler(ws4), bkkr(ws6), x1555r(ws2, ws3), splicewirer(ws5), samewirer(ws7),
                     x2799r(ws2, ws3), xab6101r(ws2, ws3), module_implementater(ws1), "", verificarelungimir(ws6, ws4),
                     copylenghtvaluesrightr(ws4)[0], copylenghtvaluesrightr(ws4)[1], copylenghtvaluesrightr(ws4)[2],
                     copylenghtvaluesrightr(ws4)[3], copylenghtvaluesrightr(ws4)[4], copylenghtvaluesrightr(ws4)[5],
                     copylenghtvaluesrightr(ws4)[6], "", copylenghtvaluesleftr(ws4)[0], copylenghtvaluesleftr(ws4)[1],
                     copylenghtvaluesleftr(ws4)[2], copylenghtvaluesleftr(ws4)[3], copylenghtvaluesleftr(ws4)[4],
                     copylenghtvaluesleftr(ws4)[5], copylenghtvaluesleftr(ws4)[6], dokar(ws1), x6616r(ws6), kswr(ws1),
                     militaryr(ws1), prufungr(ws6, ws1, ws2, ws3), my2023r(ws6, ws1), x6616stvbr(ws1, ws8, ws6),
                     module_check(ws1), ckd(ws1), delivery(ws1), ws1.cell(row=2, column=11).value, stvb(ws6)])
                end2 = time.time()
                file_progres = file_progres + 1
                statuslabel["text"] = "                 8023: " + str(file_progres) + "/" + \
                                      str(file_counter) + " : " + file_all + "    "
                timelabel["text"] = "Estimated time to complete : " + \
                                    str(((file_counter * 0.3) - (end2 - start2))/60)[:5] + " minutes."
                pbar['value'] += 2
                pbargui.update_idletasks()
        except PermissionError:
            messagebox.showerror('Eroare scriere', "Fisierul " + file_all + "este read-only!")
            quit()
    statuslabel["text"] = "Verificare raport . . ."
    timelabel["text"] = "==================="
    pbar['value'] += 2
    pbargui.update_idletasks()
    verificare_raport(raport_final)
    statuslabel["text"] = "Printare raport . . ."
    timelabel["text"] = "==================="
    pbar['value'] += 2
    pbargui.update_idletasks()
    prn_excel_raport(raport_final)
    pbar.destroy()
    pbargui.destroy()
    end = time.time()
    messagebox.showinfo('Finalizat!', 'Prelucrate ' +
                        str(counter) + " fisiere in " + str(end - start)[:6] + " secunde.")


def creare_raport_director():
    pbargui = Tk()
    pbargui.title("Progres . . . ")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    counter = 0
    globale.director_salvare_raport = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                                              title="Selectati directorul pentru salvare")
    dir_prelucrare = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir) + '/MAN/Output/Excel Files',
                                             title="Selectati directorul cu fisiere:")
    raport_final = [["Release in production", "Call off", "Raw call off name", "KSK", "Klappschalle", "Super sleeve",
                     "Extra autarke modul", "Old/new modul not in the UA list", "Heck Module", "BKK", "X1555.1A1",
                     "Splice/wire", "Same wire no in harness", "X2799.1A1 or X2799.1A1_1", "XA.B129.1 or XA.B610.1",
                     "Modules not implemented", "Check", "comment", "Side", "Abbreviation", "q / W", "r / X", "s / Y",
                     "t", "u", "", "Side", "Abbreviation", "q / W", "r / X", "s / Y", "t", "u", "DoKa", "X6616/X6490",
                     "KSW module", "Military", "Arad Prüfung", "MY 2023", "X6490 Module", "Module check", "CKD",
                     "Delivery Date", "Trailer No", "Stvb RHM FHS/RHM"]]
    start = time.time()
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_prelucrare):
        if file_all.endswith(".xlsx") and not file_all.startswith("BOM"):
            file_counter = file_counter + 1
    if file_counter == 0:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showinfo("Fisier invalid", "Nu am gasit fisiere de prelucrat!")
        return None
    for file_all in os.listdir(dir_prelucrare):
        try:
            if file_all.endswith(".xlsx") and not file_all.startswith("BOM"):
                counter = counter + 1
                wb = load_workbook(dir_prelucrare + "/" + file_all)
                ws1 = wb.worksheets[0]
                ws2 = wb.worksheets[1]
                ws3 = wb.worksheets[2]
                ws4 = wb.worksheets[3]
                ws5 = wb.worksheets[4]
                ws6 = wb.worksheets[5]
                ws7 = wb.worksheets[6]
                ws8 = wb.worksheets[7]
                raport_final.append(
                    ["", "", "", file_namer(ws1), klappschaller(ws6, ws1), supersleever(ws1), extraautarker(ws1),
                     oldnewcheckr(ws1), heckmoduler(ws4), bkkr(ws6), x1555r(ws2, ws3), splicewirer(ws5), samewirer(ws7),
                     x2799r(ws2, ws3), xab6101r(ws2, ws3), module_implementater(ws1), "", verificarelungimir(ws6, ws4),
                     copylenghtvaluesrightr(ws4)[0], copylenghtvaluesrightr(ws4)[1], copylenghtvaluesrightr(ws4)[2],
                     copylenghtvaluesrightr(ws4)[3], copylenghtvaluesrightr(ws4)[4], copylenghtvaluesrightr(ws4)[5],
                     copylenghtvaluesrightr(ws4)[6], "", copylenghtvaluesleftr(ws4)[0], copylenghtvaluesleftr(ws4)[1],
                     copylenghtvaluesleftr(ws4)[2], copylenghtvaluesleftr(ws4)[3], copylenghtvaluesleftr(ws4)[4],
                     copylenghtvaluesleftr(ws4)[5], copylenghtvaluesleftr(ws4)[6], dokar(ws1), x6616r(ws6), kswr(ws1),
                     militaryr(ws1), prufungr(ws6, ws1, ws2, ws3), my2023r(ws6, ws1), x6616stvbr(ws1, ws8, ws6),
                     module_check(ws1), ckd(ws1), delivery(ws1), ws1.cell(row=2, column=11).value, stvb(ws6)])
                file_progres = file_progres + 1
                statuslabel["text"] = "                 " + str(file_progres) + "/" + \
                                      str(file_counter) + " : " + file_all
                pbar['value'] += 2
                pbargui.update_idletasks()
        except PermissionError:
            messagebox.showerror('Eroare scriere', "Fisierul " + file_all + "este read-only!")
            quit()
    verificare_raport(raport_final)
    prn_excel_raport(raport_final)
    pbargui.destroy()
    end = time.time()
    messagebox.showinfo('Finalizat!', 'Prelucrate '
                        + str(counter) + " fisiere in " + str(end - start)[:6] + " secunde.")


def verificare_raport(arr_raport):
    exception_array = ["433014_6", "591003_1", "713004_4", "310000_509"]
    for i in range(1, len(arr_raport)):
        output = []
        txt = arr_raport[i][12].replace("Wire No. - ", "")
        txt_arr = txt.rsplit(",")
        for x in range(len(txt_arr)):
            if txt_arr[x].strip() not in exception_array:
                output.append(txt_arr[x])
        if "Error" in arr_raport[i][4]:
            arr_raport[i][0] = "Klappschalle error"
        elif "wrong" in arr_raport[i][4]:
            arr_raport[i][0] = "Klappschalle error"
        elif "Missing" in arr_raport[i][4]:
            arr_raport[i][0] = "Klappschalle error"
        elif "found" in arr_raport[i][4]:
            arr_raport[i][0] = "Klappschalle error"
        elif arr_raport[i][5] == "No Module":
            arr_raport[i][0] = "No Supersleeve Module"
        elif "Error" in arr_raport[i][5]:
            arr_raport[i][0] = arr_raport[i][5]
        elif arr_raport[i][7] != "None":
            arr_raport[i][0] = "Old/new error"
        elif arr_raport[i][8] == "No Heck module":
            arr_raport[i][0] = "Heckmodule error"
        elif arr_raport[i][10] != "OK":
            arr_raport[i][0] = "X1555 error"
        elif arr_raport[i][11] != "OK":
            arr_raport[i][0] = "Splice wire error"
        elif arr_raport[i][12] != "OK" and len(output) > 0:
            arr_raport[i][0] = "Same wire error"
        elif arr_raport[i][13] != "OK":
            arr_raport[i][0] = "X2799 error"
        elif arr_raport[i][15] != "OK":
            arr_raport[i][0] = "Modules not implemented error"
        elif "Error" in arr_raport[i][17]:
            arr_raport[i][0] = "Comment error"
        elif "Mixed" in arr_raport[i][17]:
            arr_raport[i][0] = "Comment error"
        elif "only" in arr_raport[i][17]:
            arr_raport[i][0] = "Comment error"
        elif "error" in arr_raport[i][17]:
            arr_raport[i][0] = "Comment error"
        elif "NOT" in arr_raport[i][34]:
            arr_raport[i][0] = "X6616/X6490 error"
        elif arr_raport[i][37] != "OK":
            arr_raport[i][0] = "Prufung error"
        elif arr_raport[i][39] != "OK":
            arr_raport[i][0] = "X6490 module missing"
        elif arr_raport[i][40] != "OK":
            arr_raport[i][0] = "Module check error"
        elif "Missing" in arr_raport[i][44]:
            arr_raport[i][0] = "Missing Stvb RHM FHS/RHM"
    return None
