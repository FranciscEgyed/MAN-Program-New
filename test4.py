import os

from openpyxl.reader.excel import load_workbook


def sort_by_first_two_characters(name):
    # Extract the first two characters of the filename
    first_two_characters = name[:2]

    # Convert the first two characters to an integer (if possible)
    try:
        return int(first_two_characters)
    except ValueError:
        # If conversion to an integer fails, return a large value
        return float('inf')


def copy_excel_to_list(file_path):
    # Load the Excel workbook
    workbook = load_workbook(file_path)
    # Select the specific sheet
    sheet = workbook.worksheets[0]
    # Create an empty list to store all rows as lists
    all_rows_list = []
    # Loop through all rows in the sheet
    for row in sheet.iter_rows(values_only=True):
        # Append each row as a list to the main list
        all_rows_list.append(list(row))
    # Close the workbook
    workbook.close()
    return all_rows_list

output = copy_excel_to_list("F:/Python Projects/MAN 2022\MAN\Output\Diagrame\EXCELS/Lista output.xlsx")

contents = os.listdir("F:/Python Projects/MAN 2022\MAN\Output\Diagrame\EXCELS/Conectori/Compatibili/")
# Sort the contents based on the first two characters as numbers
contents.sort(key=sort_by_first_two_characters)
lista_fire = output
for i in range(len(contents)):

    print(contents[i][contents[i].find(' ') + 1:-5])
    module_diagrama = copy_excel_to_list("F:/Python Projects/MAN 2022\MAN\Output\Diagrame\EXCELS/Conectori/Compatibili/" +contents[i])
    for combinatie in module_diagrama:
        output_diagrama = []
        for modul in combinatie:
            output_diagrama.append(modul)
        print(output_diagrama)
        print()
        for modul_diagrama in output_diagrama:
            diagrama = []
            if modul_diagrama is not None:
                for x in range(len(lista_fire)):
                    if lista_fire[x][0] == contents[i][contents[i].find(' ') + 1:-5] and lista_fire[x][11] == modul_diagrama:
                        diagrama.append([modul_diagrama, lista_fire[x][7], lista_fire[x][12]])
                for dia in diagrama:
                    print(dia)
        print()