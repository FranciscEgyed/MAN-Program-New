import csv
import datetime
import itertools
import math
import os
import time
from tkinter import messagebox, Tk, ttk, HORIZONTAL, Label, Entry, Listbox, END, Button
import pandas as pd
from openpyxl.reader.excel import load_workbook
from functii_print import prn_excel_cutting, prn_excel_supers_ksk_all, prn_excel_compare_ksk_light, \
    prn_excel_moduleinksk
import sqlite3


def cutting_ksklight():
    pbargui = Tk()
    pbargui.title("Lista taiere KSK Light")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    start = time.time()
    control_matrix = []
    try:
        with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Control_Matrix_CSL.txt", newline='') as csvfile:
            control_matrix_csl = list(csv.reader(csvfile, delimiter=';'))
        with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Control_Matrix_CSR.txt", newline='') as csvfile:
            control_matrix_csr = list(csv.reader(csvfile, delimiter=';'))
    except FileNotFoundError:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showerror('Eroare fisier', 'Lipsa fisier Control_Matrix din INPUT')
        return None
    # Liste excludere si inlocuire
    excluderecst = ["81.25482-6147", "81.25482-6148", "81.25480-5681"]
    for i in range(len(control_matrix_csl)):
        if control_matrix_csl[i][1] != "":
            control_matrix.append(control_matrix_csl[i])
    pbar['value'] += 2
    pbargui.update_idletasks()
    for i in range(len(control_matrix_csr)):
        if control_matrix_csr[i][1] != "":
            control_matrix.append(control_matrix_csr[i])
    pbar['value'] += 2
    pbargui.update_idletasks()

    array_wires_all = []
    try:
        with open(os.path.abspath(os.curdir) + "/MAN/Input/Wire Lists/8011.Wirelist.csv",
                  newline='') as csvfile:
            array_wires_8011 = list(csv.reader(csvfile, delimiter=';'))
        with open(os.path.abspath(os.curdir) + "/MAN/Input/Wire Lists/8013.Wirelist.csv",
                  newline='') as csvfile:
            array_wires_8013 = list(csv.reader(csvfile, delimiter=';'))
    except:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showerror('Eroare fisier', 'Lipsa fisiere Wire List')
        return None
    array_wires_all.extend(array_wires_8011)
    array_wires_all.extend(array_wires_8013)
    lista_cutting = [["CC", "KSK No", "Module ID", "Wire No.", "LTG PMD", "Color", "Cross Sec", "Conector 1", "Pin 1",
                      "Conector 2", "Pin 2", "Sonderltg", "Length", "KANBAN-AG", "REAL NAME", "Ledset"]]
    dir_selectat = os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/Beius/Neprelucrate/"
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".xlsx"):
            file_counter = file_counter + 1
    if file_counter == 0:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showerror('Eroare fisier', 'Lipsa fisiere in director')
        return None
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".xlsx"):
            array_fire_ksk = []
            pbar['value'] += 2
            file_progres = file_progres + 1
            statuslabel["text"] = str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbargui.update_idletasks()
            path = os.path.join(dir_selectat, file_all)
            wb = load_workbook(path)
            ws = wb.worksheets[0]
            lista_module = [ws.cell(row=row.row, column=1).value for row in ws['A'] if row.value is not None]
            for q in range(len(array_wires_all)):
                for x in range(len(lista_module)):
                    if array_wires_all[q][0] == lista_module[x] and array_wires_all[q][0] not in excluderecst:
                        array_fire_ksk.append([file_all.split(".")[0], array_wires_all[q][0],
                                               array_wires_all[q][1].lower(),
                                               array_wires_all[q][2], array_wires_all[q][3], array_wires_all[q][4],
                                               array_wires_all[q][5], array_wires_all[q][6], array_wires_all[q][7],
                                               array_wires_all[q][8], array_wires_all[q][9], array_wires_all[q][10]])
            for i in range(len(array_fire_ksk)):
                lista_cutting.append([array_fire_ksk[i][0] + array_fire_ksk[i][2] + array_fire_ksk[i][6] +
                                      array_fire_ksk[i][7] + array_fire_ksk[i][8] + array_fire_ksk[i][9],
                                      array_fire_ksk[i][0], array_fire_ksk[i][1], array_fire_ksk[i][2],
                                      array_fire_ksk[i][3], array_fire_ksk[i][4], array_fire_ksk[i][5],
                                      array_fire_ksk[i][6], array_fire_ksk[i][7], array_fire_ksk[i][8],
                                      array_fire_ksk[i][9], array_fire_ksk[i][10], array_fire_ksk[i][11]])
    pbar['value'] += 2
    statuslabel["text"] = "Searching Control matrix        "
    pbargui.update_idletasks()
    for i in range(1, len(lista_cutting)):
        for x in range(len(control_matrix)):
            if lista_cutting[i][2] + lista_cutting[i][3] == control_matrix[x][0]:
                lista_cutting[i].extend([control_matrix[x][1], control_matrix[x][3], control_matrix[x][5]])
                break
        pbar['value'] += 2
        pbargui.update_idletasks()
    for i in range(len(lista_cutting)):
        try:
            if lista_cutting[i][13] == "-":
                for x in range(len(control_matrix)):
                    if lista_cutting[i][14] == control_matrix[x][3] and control_matrix[x][1] != "-":
                        lista_cutting[i][13] = control_matrix[x][1]
                        break
            pbar['value'] += 2
            pbargui.update_idletasks()
        except IndexError:
            continue
    for i in range(1, len(lista_cutting)):
        try:
            lista_cutting[i][13] = lista_cutting[i][13].split("-")[0].replace("U", "W")
        except IndexError:
            continue
    lista_unice = []
    lista_cutting_unice = []
    lista_cutting_twisturi = []
    lista_unice_twisturi = []
    for i in range(len(lista_cutting)):
        try:
            if lista_cutting[i][0] not in lista_unice and lista_cutting[i][15] != "":
                lista_unice.append(lista_cutting[i][0])
        except IndexError:
            continue
    for i in range(len(lista_cutting)):
        try:
            if [lista_cutting[i][1] + lista_cutting[i][14]] not in lista_unice_twisturi and lista_cutting[i][15] == "":
                lista_cutting_twisturi.append(lista_cutting[i][0])
                lista_unice_twisturi.append([lista_cutting[i][1] + lista_cutting[i][14]])
        except IndexError:
            continue
    for x in range(len(lista_unice)):
        for i in range(len(lista_cutting)):
            if lista_cutting[i][0] == lista_unice[x]:
                lista_cutting_unice.append(lista_cutting[i])
                break
    for x in range(len(lista_cutting_twisturi)):
        for i in range(len(lista_cutting)):
            if lista_cutting[i][0] == lista_cutting_twisturi[x]:
                lista_cutting_unice.append(lista_cutting[i])
                break
    lista_wire_no = []
    for i in range(len(lista_cutting_unice)):
        try:
            if [lista_cutting_unice[i][14], lista_cutting_unice[i][13]] not in lista_wire_no:
                lista_wire_no.append([lista_cutting_unice[i][14], lista_cutting_unice[i][13]])
        except IndexError:
            continue
    lista_wire_no[0].append("Cantitate")
    for i in range(1, len(lista_wire_no)):
        counter = 0
        for x in range(len(lista_cutting_unice)):
            try:
                if lista_wire_no[i][0] == lista_cutting_unice[x][14]:
                    counter += 1
            except IndexError:
                continue
        lista_wire_no[i].append(counter)
    pbar['value'] += 2
    statuslabel["text"] = "Printing file                     "
    pbargui.update_idletasks()
    prn_excel_cutting(lista_cutting, lista_cutting_unice, lista_wire_no)
    end = time.time()
    pbar.destroy()
    pbargui.destroy()
    messagebox.showinfo('Finalizat!', str(end - start)[:6] + " secunde.")


def ss_ksklight():
    pbargui = Tk()
    pbargui.title("Lista SuperSleeve KSK Light")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    timelabel = Label(pbargui, text="Time . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    timelabel.grid(row=2, column=2)
    counter = 0
    start = time.time()
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/SSDrawings.txt", newline='') as csvfile:
        drawings = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Component Overview.txt", newline='') as csvfile:
        compover = list(csv.reader(csvfile, delimiter=';'))
    dir_selectat = os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/Beius/Neprelucrate/"
    normal = ["04.37161-9100", "81.25484-5259", "81.25484-5263", "81.25484-5260", "81.25484-5264", "81.25484-5273",
              "81.25484-5272", "81.25484-5267", "81.25484-5268"]
    ADR = ["04.37161-9000", "81.25484-5261", "81.25484-5265", "81.25484-5262", "81.25484-5266", "81.25484-5275",
           "81.25484-5274", "81.25484-5271", "81.25484-5270"]
    array_corespondenta = [["04.37161-9144", "P00194490", "P00154252"], ["04.37161-9133", "P00194492", "P00154251"],
                           ["04.37161-9123", "P00194489", "P00154250"], ["04.37161-9115", "P00194491", "P00154249"],
                           ["04.37161-9149", "P00194421", "P00154253"], ["04.37161-9158", "P00194420", "P00157590"]]

    array_print = [["KSK No", "Basic_Module", "Module", "[Index]", "Module", "Part Number", "ID_TAP", "CA_SegmentD",
                    "Lengs_DWмм", "In KSK"]]
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".xlsx"):
            file_counter = file_counter + 1

    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".xlsx"):
            start1 = time.time()
            with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/SSDatabase.txt", newline='') as csvfile:
                ssdatabase = list(csv.reader(csvfile, delimiter=';'))
            ssdatabase_de_prelucrat = ssdatabase
            file_progres = file_progres + 1
            statuslabel["text"] = str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 2
            pbargui.update_idletasks()
            path = os.path.join(dir_selectat, file_all)
            wb = load_workbook(path)
            ws = wb.worksheets[0]
            lista_module_ksk_light = [ws.cell(row=row.row, column=1).value for row in ws['A'] if row.value is not None]

            for i in range(len(ssdatabase_de_prelucrat)):
                if ssdatabase_de_prelucrat[i][1] in lista_module_ksk_light:
                    ssdatabase_de_prelucrat[i].append(1)
                else:
                    ssdatabase_de_prelucrat[i].append(0)

            ss_prelucrat = [ssdatabase_de_prelucrat[i] for i in range(len(ssdatabase_de_prelucrat))
                            if ssdatabase_de_prelucrat[i][8] != 0]
            ss_prelucrat.insert(0, ["Basic_Module", "Module", "Index", "Module", "Part Number", "NORMAL_COD", "Nume",
                                    "Lungime mm"])

            df1 = pd.DataFrame(ss_prelucrat)
            df1.columns = df1.iloc[0]
            df1 = df1[1:]
            df2 = pd.pivot_table(df1, index=["Basic_Module"], columns=["Index"], fill_value=0)
            indexes = df2.index.values.tolist()
            valori = df2.values.tolist()
            # index_value_pairs = [indexes[i] for i in range(len(valori)) if valori[i][0] != 0 and valori[i][1] != 0]
            index_value_pairs = []
            for i in range(len(valori)):
                if valori[i][0] != 0 and valori[i][1] != 0:
                    index_value_pairs.append(indexes[i])
            add_to_print = []
            for i in range(len(index_value_pairs)):
                for x in range(len(ssdatabase)):
                    templist = []
                    if index_value_pairs[i] == ssdatabase[x][0] and ssdatabase[x] not in add_to_print:
                        templist = ssdatabase[x]
                        templist.insert(0, file_all[:-5])
                        add_to_print.append(templist)
            counter = counter + 1
            array_print.extend(add_to_print)
            end1 = time.time()
            timelabel["text"] = "Estimated time to complete : " \
                                + str((file_counter * (end1-start1)) / 60)[:5] + " minutes."
            pbargui.update_idletasks()
    statuslabel["text"] = "Creare lista SuperSleeve         "
    pbar['value'] += 2
    pbargui.update_idletasks()
    array_unic_print = [[array_print[i][1], array_print[i][0] + array_print[i][1], array_print[i][2],
                         array_print[i][3], array_print[i][9], array_print[i][5]] for i in range(1, len(array_print))]
    array_unic_print2 = [array_unic_print[i] for i in range(len(array_unic_print))
                         if array_unic_print[i][3] == str(array_unic_print[i][4])]
    for i in range(len(array_unic_print2)):
        if array_unic_print2[i][2] in normal:
            array_unic_print2[i][3] = "Normal"
        elif array_unic_print2[i][2] in ADR:
            array_unic_print2[i][3] = "ADR"
    for i in range(len(array_unic_print2)):
        for x in range(len(array_print)):
            if array_unic_print2[i][0] == array_print[x][1]:
                array_unic_print2[i].append(array_print[x][7])
                break
    statuslabel["text"] = "Creare lista completa SuperSleeve         "
    pbar['value'] += 2
    pbargui.update_idletasks()
    array_unic_print2.insert(0, ["Basic Module", "CC: KSK+BM", "Modul ID", "Tip", "Cantitate", "Part No", "Segment"])
    for i in range(len(array_unic_print2)):
        for x in range(len(array_corespondenta)):
            if array_unic_print2[i][3] == "Normal" and array_unic_print2[i][5] == array_corespondenta[x][0]:
                array_unic_print2[i][5] = array_corespondenta[x][1]
            elif array_unic_print2[i][3] == "ADR" and array_unic_print2[i][5] == array_corespondenta[x][0]:
                array_unic_print2[i][5] = array_corespondenta[x][2]
    df = pd.DataFrame(array_unic_print2)
    df = df.reset_index(drop=True)
    df.columns = df.iloc[0]
    df = df[1:]
    statuslabel["text"] = "Creare lista taiere SuperSleeve          "
    pbar['value'] += 2
    pbargui.update_idletasks()
    pivot = df.pivot_table(index="Segment", columns="Part No", values="Cantitate", fill_value=0, aggfunc='count')
    indexes = pivot.index.values.tolist()
    valori = pivot.values.tolist()
    coloane = pivot.columns.tolist()
    array_taiere_print = [["Segment"]]
    array_taiere_print[0].extend(coloane)
    for i in range(len(indexes)):
        temp = []
        temp.append(indexes[i])
        temp.extend(valori[i])
        array_taiere_print.append(temp)
    array_taiere_print[0].append("Lungime mm")
    for i in range(1, len(array_taiere_print)):
        for x in range(len(array_print)):
            if array_taiere_print[i][0] == array_print[x][7]:
                array_taiere_print[i].append(array_print[x][8])
                break
    statuslabel["text"] = "Cautare part number LEONI             "
    pbar['value'] += 2
    pbargui.update_idletasks()
    # for i in range(1, len(array_taiere_print[0]) - 1):
    #    pbar['value'] += 2
    #    pbargui.update_idletasks()
    #    for x in range(len(compover)):
    #        if array_taiere_print[0][i] == compover[x][0]:
    #            array_taiere_print[0][i] = compover[x][2]
    # answer = askyesno(title='Optiuni printare', message='Doriti printare completa?')
    statuslabel["text"] = "Printare lista             "
    pbar['value'] += 2
    array_taiere_print[0].append("Desen")
    for i in range(1, len(array_taiere_print)):
        for x in range(len(drawings)):
            if array_taiere_print[i][0] == drawings[x][0]:
                array_taiere_print[i].append(drawings[x][1])
                break
    #pbargui.update_idletasks()
    #if answer:
    prn_excel_supers_ksk_all(array_print, array_unic_print2, array_taiere_print)
    #else:
    #    prn_excel_supers_ksk_all_simplu(array_taiere_print)

    pbar.destroy()
    pbargui.destroy()
    end = time.time()
    messagebox.showinfo('Finalizat!', 'Prelucrate ' + str(counter) + " fisiere in "
                        + str(end - start)[:6][:6] + " secunde.")


def compare_ksk_light():
    pbargui = Tk()
    pbargui.title("Comparatie KSK Light")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    timelabel = Label(pbargui, text="Time . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    timelabel.grid(row=2, column=2)
    lista_excludere = []
    lista_identice = []
    # lista_nonidentice = []
    start = time.time()
    files_dir = os.listdir(os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/Beius/Neprelucrate/")
    my_dir = os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/Beius/Neprelucrate/"
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(my_dir):
        if file_all.endswith(".xlsx"):
            file_counter = file_counter + 1
    file_counter = int(math.factorial(file_counter) / ((math.factorial(2)) * math.factorial((file_counter - 2))))
    timelabel["text"] = "             Estimated time to complete : " + \
                        str(((file_counter * 0.02988) / 60) - ((file_progres * 0.02988) / 60)) + " minutes.            "
    pbargui.update_idletasks()
    if file_counter == 0:
        messagebox.showerror('Eroare fisier', 'Lipsa fisiere in director')
        return None
    for file1, file2 in itertools.combinations(files_dir, 2):
        pbar['value'] += 2
        file_progres = file_progres + 1
        statuslabel["text"] = "                Combinatii verificate " + str(file_progres) + "/" + str(file_counter) +\
                              "                "
        end = time.time()
        timelabel["text"] = "         Estimated time to complete : " + \
                            str(((file_counter * 0.02988) / 60) - ((file_progres * 0.02988) / 60)) + " minutes.        "
        pbargui.update_idletasks()
        path1 = os.path.join(my_dir, file1)
        path2 = os.path.join(my_dir, file2)
        df1 = pd.read_excel(path1, header=None)
        df2 = pd.read_excel(path2, header=None)
        if len(df1.index) == len(df2.index):
            if df1.equals(df2):
                lista_identice.append([file1, file2])
                lista_excludere.append(file1)
                lista_excludere.append(file2)
    lista_nonidentice = [[files_dir[i], ""] for i in range(len(files_dir)) if files_dir[i] not in lista_excludere]
    lista = lista_identice
    lista_xxx = lista_identice
    for i in range(1, len(lista)):
        for x in range(len(lista_xxx)):
            if lista[i] not in lista_xxx:
                lista_xxx.append(lista[i])
            elif lista[i][0] in lista_xxx[x]:
                if lista[i][1] not in lista_xxx[x]:
                    lista_xxx[x].append(lista[i][1])
            elif lista[i][1] in lista_xxx[x]:
                if lista[i][0] not in lista_xxx[x]:
                    lista_xxx[x].append(lista[i][0])
    lista_print = []
    for i in range(len(lista_xxx)):
        if sorted(lista_xxx[i]) not in lista_print:
            lista_print.append(sorted(lista_xxx[i]))

    prn_excel_compare_ksk_light(lista_print, lista_nonidentice)
    pbar.destroy()
    pbargui.destroy()
    messagebox.showinfo('Finalizat!')


def raport_light():
    # Create your connection.
    try:
        cnx = sqlite3.connect("//SVRO8FILE01/Groups/General/EFI/DBMAN/database.db")
    except sqlite3.OperationalError:
        cnx = sqlite3.connect(os.path.abspath(os.curdir) + "/MAN/Input/Others/database.db")
        messagebox.showinfo("Local database", "Network database unavailable. Using local database.")
    df = pd.read_sql_query("SELECT * FROM KSKDatabase", cnx)

    def scankey(event):
        val = event.widget.get()
        if val == '':
            data = list1
        else:
            data = []
            for item in list1:
                if val.lower() in item.lower():
                    data.append(item)
        update(data)

    def update(data):
        datalivrare_lb.delete(0, 'end')
        for item in data:
            datalivrare_lb.insert('end', item)

    def select_all():
        datalivrare_lb.select_set(0, END)

    def deselect_all():
        datalivrare_lb.selection_clear(0, END)

    list1 = df.DataLivrare.unique()

    def scankey2(event):
        val = event.widget.get()
        if val == '':
            data = list2
        else:
            data = []
            for item in list2:
                if val.lower() in item.lower():
                    data.append(item)
        update2(data)

    def update2(data):
        datajit_lb.delete(0, 'end')
        for item in data:
            datajit_lb.insert('end', item)

    def select_all2():
        datajit_lb.select_set(0, END)

    def deselect_all2():
        datajit_lb.selection_clear(0, END)

    list2 = df.DataJIT.unique()

    def update3(data):
        indexe.delete(0, 'end')
        # put new data
        for item in data:
            indexe.insert('end', item)

    list3 = df.columns.values.tolist()[2:]

    def update4(data):
        coloane.delete(0, 'end')
        for item in data:
            coloane.insert('end', item)

    list4 = df.columns.values.tolist()[2:]

    def run():
        save_time = datetime.datetime.now().strftime("%d.%m.%Y_%H.%M.%S")
        valuesdatalivrare_lb = [datalivrare_lb.get(idx) for idx in datalivrare_lb.curselection()]
        valuesdatajit_lb = [datajit_lb.get(idx) for idx in datajit_lb.curselection()]
        xxx = df.query('DataLivrare in @valuesdatalivrare_lb')
        yyy = xxx.query('DataJIT in @valuesdatajit_lb')
        indexlist = [indexe.get(idx) for idx in indexe.curselection()]
        columnlist = [coloane.get(idx) for idx in coloane.curselection()]
        if len(indexlist) + len(columnlist) == 0 or len(valuesdatalivrare_lb) + len(valuesdatajit_lb) == 0:
            messagebox.showerror("Valori gresite", "Nu ati selectat nimic")
        else:
            try:
                pivot = yyy.pivot_table(index=indexlist, columns=columnlist, values="primarykey", fill_value=0,
                                        aggfunc='count')
                # print(pivot.to_string())
                pivot.to_excel(os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/Raport -" + indexlist[0] +
                               "-" + save_time + ".xlsx")
                ws.destroy()
                messagebox.showinfo("Finalizat", "Raportul " + indexlist[0] + "/" + columnlist[0] + " a fost salvat!")
            except ValueError:
                ws.destroy()
                messagebox.showerror("Valori gresite", "Indexul si coloanele nu pot contine aceasi informatii.")


    def moduleinksk():
        printer = []
        printer2 = []
        for x in range(0, len(df.index)):
            printer.append([df.iloc[x, 6], df.iloc[x, 9].split(";")])
        for i in range(len(printer)):
            for x in range(len(printer[i][1])):
                printer2.append([printer[i][0], printer[i][1][x]])
        try:
            prn_excel_moduleinksk(printer2)
            ws.destroy()
            messagebox.showinfo("Finalizat", "Raportul Lista module in KSK a fost salvat!")
        except ValueError:
            ws.destroy()
            messagebox.showerror("Error", "Error")

    def comparatie():
        save_time = datetime.datetime.now().strftime("%d.%m.%Y_%H.%M.%S")
        valuesdatalivrare_lb = [datalivrare_lb.get(idx) for idx in datalivrare_lb.curselection()]
        valuesdatajit_lb = [datajit_lb.get(idx) for idx in datajit_lb.curselection()]
        xxx = df.query('DataLivrare in @valuesdatalivrare_lb')
        yyy = xxx.query('DataJIT in @valuesdatajit_lb')
        if len(valuesdatalivrare_lb) + len(valuesdatajit_lb) == 0:
            messagebox.showerror("Valori gresite", "Nu ati selectat nimic")
        else:

            pivot = yyy.pivot_table(index="Module", columns="KSKNo", values="primarykey", fill_value=0,
                                    aggfunc='count')

            pivot.loc[:, 'Total'] = pivot.iloc[:, 1:].sum(axis=1)
            printarray = []
            for x in range(len(pivot.index)):
                if pivot.iloc[x, -1] > 1:
                    templist = []
                    for y in range(len(pivot.columns)-1):
                        if pivot.iloc[x, y] != 0:
                            templist.append(pivot.columns[y])
                    printarray.append(templist)
            prn_excel_compare_ksk_light(printarray, save_time, valuesdatajit_lb)

            ws.destroy()
            messagebox.showinfo("Finalizat", "Comparatia " + save_time + " a fost salvat!")


    ws = Tk()
    ws.title("2022 MAN KSK Light reports")
    ws.geometry("+570+50")
    l1 = Label(ws, text="Data JIT")
    l2 = Label(ws, text="Data livrare")
    l3 = Label(ws, text="Index pentru pivot")
    l4 = Label(ws, text="Coloane pentru pivot")
    l1.grid(row=0, column=1)
    l2.grid(row=0, column=0)
    l3.grid(row=0, column=2)
    l4.grid(row=0, column=3)
    entry = Entry(ws)
    entry.grid(row=1, column=0)
    entry.bind('<KeyRelease>', scankey)
    entry2 = Entry(ws)
    entry2.grid(row=1, column=1)
    entry2.bind('<KeyRelease>', scankey2)

    bs1 = Button(ws, text="Select All", command=select_all)
    bs2 = Button(ws, text="Select All", command=select_all2)
    bs1.grid(row=3, column=0)
    bs2.grid(row=3, column=1)
    bds1 = Button(ws, text="DeSelect All", command=deselect_all)
    bds2 = Button(ws, text="DeSelect All", command=deselect_all2)
    bds1.grid(row=4, column=0)
    bds2.grid(row=4, column=1)

    brun = Button(ws, text="Generate report", command=run, bg="green", font="Arial 10 bold")
    brun.grid(row=5, column=4)
    bmoduleinksk = Button(ws, text="Raport module in KSK", command=moduleinksk, bg="yellow", font="Arial 10 bold")
    bmoduleinksk.grid(row=6, column=4)
    bmoduleinksk = Button(ws, text="Raport comparatie KSK", command=comparatie, bg="blue", font="Arial 10 bold")
    bmoduleinksk.grid(row=7, column=4)

    datalivrare_lb = Listbox(ws, exportselection=0, selectmode="multiple")
    datajit_lb = Listbox(ws, exportselection=0, selectmode="multiple")
    indexe = Listbox(ws, exportselection=0)
    coloane = Listbox(ws, exportselection=0, selectmode="multiple")
    datalivrare_lb.grid(row=2, column=0)
    datajit_lb.grid(row=2, column=1)
    indexe.grid(row=2, column=2)
    coloane.grid(row=2, column=3)
    update(list1)
    update2(list2)
    update3(list3)
    update4(list4)
    ws.mainloop()
