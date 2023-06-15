import os
from tkinter import filedialog, messagebox
from tkinter.colorchooser import askcolor

from openpyxl.reader.excel import load_workbook
import qrcode


def eticheteqr():
    colors = askcolor(title="Code Color Chooser")
    bcolors = askcolor(title="Background Color Chooser")
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    file_load = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                           title="Incarcati fisierul EXCEl cu informatie de pe eticheta:")
    try:
        wb = load_workbook(file_load)
        ws1 = wb.active
        for row in ws1['A']:
            if row.value is not None:
                # img = qrcode.make(ws1.cell(row=row.row, column=1).value, back_color=bcolors[1], fill_color=colors[1])
                print(ws1.cell(row=row.row, column=1).value)
                qr.add_data(ws1.cell(row=row.row, column=1).value)
                img = qr.make_image(back_color=bcolors[1], fill_color=colors[1])
                img = img.resize((200, 200))
                img.save(os.path.abspath(os.curdir) + "/MAN/Output/QR Images/" +
                         str(ws1.cell(row=row.row, column=1).value.replace("/", " ")) + ".jpg")
                qr.clear()
        messagebox.showinfo("Finalizat", "Finalizat!")
    except:
        messagebox.showerror("No file . . . ", "Nu ati selectat nici un fisier!!")
