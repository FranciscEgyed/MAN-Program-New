import datetime
import os
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
import pandas as pd
import globale
from diverse import istsoll, log_file


def prn_excel_wires(sheet1, sheet2, sheet3, sheet4, sheet5, sheet6, sheet7, sheet8):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    dir_salvare_default = ""
    if sheet2[1][0] == "8011" or sheet2[1][0] == "8012" or sheet2[1][0] == "8013" or sheet2[1][0] == "8014":
        dir_salvare_default = "8011/"
    if sheet2[1][0] == "8023" or sheet2[1][0] == "8024" or sheet2[1][0] == "8025" or sheet2[1][0] == "8026":
        dir_salvare_default = "8023/"
    if sheet2[1][0] == "8000" or sheet2[1][0] == "8001":
        dir_salvare_default = "8000/"
    if sheet2[1][0] == "8011.MY23" or sheet2[1][0] == "8012.MY23" or sheet2[1][0] == "8013.MY23" or \
            sheet2[1][0] == "8014.MY23":
        dir_salvare_default = "/8011/"
    if sheet2[1][0] == "8023.MY23" or sheet2[1][0] == "8024.MY23" or sheet2[1][0] == "8025.MY23" or \
            sheet2[1][0] == "8026.MY23":
        dir_salvare_default = "/8023/"
    if sheet2[1][0] == "8000.MY23" or sheet2[1][0] == "8001.MY23":
        dir_salvare_default = "/8000/"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = sheet1[1][0]
    ws2 = wb.create_sheet(sheet2[1][0])
    ws3 = wb.create_sheet(sheet3[1][0])
    ws4 = wb.create_sheet("Variatie Lungimi")
    ws5 = wb.create_sheet("Erori")
    ws6 = wb.create_sheet("Bracket")
    ws7 = wb.create_sheet("Same Wire")
    ws8 = wb.create_sheet("X6616X6490")
    ws9 = wb.create_sheet("Variatie Lungimi IST SOLL")
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            if "E-" in sheet1[i][x]:
                try:
                    ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
                except:
                    ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
            else:
                try:
                    ws1.cell(column=x + 1, row=i + 1, value=float(sheet1[i][x]))
                except:
                    ws1.cell(column=x + 1, row=i + 1, value=sheet1[i][x])
    for i in range(len(sheet2)):
        for x in range(len(sheet2[i])):
            if x == 2:
                ws2.cell(column=x + 1, row=i + 1, value=str(sheet2[i][x]))
            else:
                try:
                    ws2.cell(column=x + 1, row=i + 1, value=float(sheet2[i][x]))
                except:
                    ws2.cell(column=x + 1, row=i + 1, value=sheet2[i][x])
    for i in range(len(sheet3)):
        for x in range(len(sheet3[i])):
            if x == 2:
                ws3.cell(column=x + 1, row=i + 1, value=str(sheet3[i][x]))
            else:
                try:
                    ws3.cell(column=x + 1, row=i + 1, value=float(sheet3[i][x]))
                except:
                    ws3.cell(column=x + 1, row=i + 1, value=sheet3[i][x])
    for i in range(len(sheet4)):
        for x in range(len(sheet4[i])):
            try:
                ws4.cell(column=x + 1, row=i + 1, value=float(sheet4[i][x]))
            except:
                ws4.cell(column=x + 1, row=i + 1, value=sheet4[i][x])
    for i in range(len(sheet5)):
        for x in range(len(sheet5[i])):
            if x == 2:
                ws5.cell(column=x + 1, row=i + 1, value=str(sheet5[i][x]))
            else:
                try:
                    ws5.cell(column=x + 1, row=i + 1, value=float(sheet5[i][x]))
                except:
                    ws5.cell(column=x + 1, row=i + 1, value=sheet5[i][x])
    for i in range(len(sheet6)):
        for x in range(len(sheet6[i])):
            try:
                ws6.cell(column=x + 1, row=i + 1, value=float(sheet6[i][x]))
            except:
                ws6.cell(column=x + 1, row=i + 1, value=sheet6[i][x])
    """Cosmetica"""
    colorsgray = PatternFill(start_color='aabbcc', end_color='aabbcc', fill_type='solid')
    colorsdoi = PatternFill(start_color='7EF1EA', end_color='7EF1EA', fill_type='solid')
    for i in range(1, 7):
        for x in range(1, len(ws6['A']) + 1):
            ws6.cell(column=i, row=x).border = thin_border
            ws6.cell(column=i, row=x).alignment = Alignment(horizontal='center')
    for row in ws6['C']:
        if row.value == "RHD":
            for i in range(1, 7):
                ws6.cell(column=i, row=row.row).fill = colorsgray
        elif row.value == "LHD":
            for i in range(1, 7):
                ws6.cell(column=i, row=row.row).fill = colorsdoi
    for cells in ws1['1']:
        ws1.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws2['1']:
        ws2.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws3['1']:
        ws3.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws4['1']:
        ws4.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws5['1']:
        ws5.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws6['1']:
        ws6.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws7['1']:
        ws7.cell(column=cells.column, row=1).font = Font(bold=True)
    """Cosmetica"""
    for i in range(len(sheet7)):
        for x in range(len(sheet7[i])):
            if x == 2:
                ws7.cell(column=x + 1, row=i + 1, value=str(sheet7[i][x]))
            else:
                try:
                    ws7.cell(column=x + 1, row=i + 1, value=float(sheet7[i][x]))
                except:
                    ws7.cell(column=x + 1, row=i + 1, value=sheet7[i][x])
    for x in range(len(sheet8[0])):
        try:
            ws8.cell(column=1, row=x + 1, value=float(sheet8[0][x][0]))
            ws8.cell(column=2, row=x + 1, value=float(sheet8[0][x][1]))
        except:
            ws8.cell(column=1, row=x + 1, value=sheet8[0][x][0])
            ws8.cell(column=2, row=x + 1, value=sheet8[0][x][1])
    for x in range(len(sheet8[1])):
        try:
            ws8.cell(column=4, row=x + 1, value=float(sheet8[1][x][0]))
            ws8.cell(column=5, row=x + 1, value=float(sheet8[1][x][1]))
        except:
            ws8.cell(column=4, row=x + 1, value=sheet8[1][x][0])
            ws8.cell(column=5, row=x + 1, value=sheet8[1][x][1])
    for x in range(len(sheet8[2])):
        try:
            ws8.cell(column=7, row=x + 1, value=float(sheet8[2][x][0]))
            ws8.cell(column=8, row=x + 1, value=float(sheet8[2][x][1]))
        except:
            ws8.cell(column=7, row=x + 1, value=sheet8[2][x][0])
            ws8.cell(column=8, row=x + 1, value=sheet8[2][x][1])
    for x in range(len(sheet8[3])):
        try:
            ws8.cell(column=10, row=x + 1, value=float(sheet8[3][x][0]))
            ws8.cell(column=11, row=x + 1, value=float(sheet8[3][x][1]))
        except:
            ws8.cell(column=10, row=x + 1, value=sheet8[3][x][0])
            ws8.cell(column=11, row=x + 1, value=sheet8[3][x][1])
    for i in range(len(sheet4)):
        ws9.cell(column=1, row=i + 1, value="IST")
        for x in range(len(sheet4[i])):
            try:
                ws9.cell(column=x + 2, row=i + 1, value=float(sheet4[i][x]))
            except:
                ws9.cell(column=x + 2, row=i + 1, value=sheet4[i][x])
    ws9.cell(column=1, row=1, value="IST / SOLL")
    ws9.insert_cols(5)
    ws9.cell(column=5, row=1, value="Description")
    istsoll(ws6, ws9)

    if globale.director_salvare == "":
        try:
            wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Excel Files/" + dir_salvare_default +
                    sheet1[1][0] + ".xlsx")
            log_file("Creat " + sheet1[1][0] + ".xlsx")
        except PermissionError:
            log_file("Eroare scriere. Nu am salvat " + sheet1[1][0] + ".xlsx")
            messagebox.showerror('Eroare scriere', "Fisierul " + sheet1[1][0] + "este read-only!")
            return None
        except FileNotFoundError:
            messagebox.showerror('Eroare scriere', "Directorul " + dir_salvare_default + " nu exista!")
            return None
    else:
        try:
            wb.save(globale.director_salvare + "/" + sheet1[1][0] + ".xlsx")
            log_file("Creat " + sheet1[1][0] + ".xlsx")
        except PermissionError:
            log_file("Eroare scriere. Nu am salvat " + sheet1[1][0] + ".xlsx")
            messagebox.showerror('Eroare scriere', "Fisierul " + sheet1[1][0] + "este read-only!")
            return None

    return None


def prn_excel_bom(sheet1, sheet2, lista_fisiere):
    dir_salvare_default = ""
    if "8011" in lista_fisiere or "8012" in lista_fisiere or "8013" in lista_fisiere or "8014" in lista_fisiere:
        dir_salvare_default = "8011/"
    if "8023" in lista_fisiere or "8024" in lista_fisiere or "8025" in lista_fisiere or "8026" in lista_fisiere:
        dir_salvare_default = "8023/"
    if "8001" in lista_fisiere or "8000" in lista_fisiere:
        dir_salvare_default = "8000/"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = sheet1[1][0]
    ws2 = wb.create_sheet("BOM")
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            if "E-" in sheet1[i][x]:
                try:
                    ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
                except:
                    ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))
            else:
                try:
                    ws1.cell(column=x + 1, row=i + 1, value=float(sheet1[i][x]))
                except:
                    ws1.cell(column=x + 1, row=i + 1, value=sheet1[i][x])
    for i in range(len(sheet2)):
        for x in range(len(sheet2[i])):
            try:
                ws2.cell(column=x + 1, row=i + 1, value=float(sheet2[i][x]))
            except:
                ws2.cell(column=x + 1, row=i + 1, value=sheet2[i][x])
    if globale.director_salvare == "":
        try:
            wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Excel Files/" + dir_salvare_default +
                    "/BOM " + sheet1[1][0] + ".xlsx")
            log_file("Creat BOM " + sheet1[1][0] + ".xlsx")
        except PermissionError:
            log_file("Eroare salvare. Nu am salvat BOM " + sheet1[1][0] + ".xlsx")
            messagebox.showerror('Eroare scriere', "Fisierul " + sheet1[1][0] + "este read-only!")
            return None
    else:
        try:
            wb.save(globale.director_salvare + "/BOM " + sheet1[1][0] + ".xlsx")
            log_file("Creat BOM " + sheet1[1][0] + ".xlsx")
        except PermissionError:
            log_file("Eroare salvare. Nu am salvat BOM " + sheet1[1][0] + ".xlsx")
            messagebox.showerror('Eroare scriere', "Fisierul " + sheet1[1][0] + "este read-only!")
            return None


def prn_excel_raport(sheet):
    save_time = datetime.datetime.now().strftime("%d.%m.%Y_%H.%M.%S")
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Raport"
    for i in range(len(sheet)):
        for x in range(len(sheet[i])):
            if "E-" in str(sheet[i][x]):
                try:
                    ws1.cell(column=x + 1, row=i + 1, value=str(sheet[i][x]))
                except:
                    ws1.cell(column=x + 1, row=i + 1, value=str(sheet[i][x]))
            else:
                try:
                    ws1.cell(column=x + 1, row=i + 1, value=float(sheet[i][x]))
                except:
                    ws1.cell(column=x + 1, row=i + 1, value=str(sheet[i][x]))
    for cells in ws1['1']:
        ws1.cell(column=cells.column, row=1).font = Font(bold=True)
    if globale.director_salvare_raport == "":
        try:
            wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Report Files/Error file " + save_time + ".xlsx")
            log_file("Creat Error file " + save_time + ".xlsx")
        except PermissionError:
            log_file("Eroare scriere. Nu am salvat Error file " + save_time + ".xlsx")
            messagebox.showerror('Eroare scriere', "Fisierul Error file " + save_time + "este read-only!")
            return None
    else:
        try:
            wb.save(globale.director_salvare_raport + "/Error file " + save_time + ".xlsx")
        except PermissionError:
            messagebox.showerror('Eroare scriere', "Fisierul Error file " + save_time + "este read-only!")
            return None


def prn_excel_separare_ksk(sheet, nume_fisier):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = nume_fisier
    for i in range(len(sheet)):
        ws1.cell(column=1, row=i + 1, value=str(sheet[i]))
    if globale.director_salvare_raport == "":
        try:
            wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/Beius/Neprelucrate/Light/" +
                    nume_fisier + ".xlsx")
            log_file("Creat Error file " + nume_fisier + ".xlsx")
        except PermissionError:
            log_file("Eroare scriere. Nu am salvat Error file " + nume_fisier + ".xlsx")
            messagebox.showerror('Eroare scriere', "Fisierul Error file " + nume_fisier + "este read-only!")
            return None
    else:
        try:
            wb.save(globale.director_salvare_raport + "/Error file " + nume_fisier + ".xlsx")
        except PermissionError:
            messagebox.showerror('Eroare scriere', "Fisierul Error file " + nume_fisier + "este read-only!")
            return None


def prn_excel_bom_complete(sheet1, nume_fisier):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = nume_fisier
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))

    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Complete BOM and WIRELIST/BOM/" + nume_fisier + ".xlsx")
        log_file("Creat BOM " + nume_fisier + ".xlsx")
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat BOM " + nume_fisier + ".xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul " + nume_fisier + "este read-only!")
        return None


def prn_excel_wires_complete(sheet1, nume_fisier):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "MAN Pn"
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=sheet1[i][x])
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Complete BOM and WIRELIST/Wirelist/" + nume_fisier + ".xlsx")
        log_file("Creat wire " + nume_fisier + ".xlsx")
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat wirelist " + nume_fisier + ".xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul " + nume_fisier + "este read-only!")
        return None


def prn_excel_wirelistsallinone(sheet1):
    wb = Workbook()
    ws1 = wb.active

    ws1.title = "Wirelists All"
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Wirelists All.xlsx")
        log_file("Creat wire Wirelists All.xlsx")
        messagebox.showinfo('Finalizat!')
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat wirelist Wirelists All.xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul Wirelists All.xlsx este read-only!")
        return None


def prn_excel_boomcumulat(sheet1):
    wb = Workbook()
    ws1 = wb.active

    ws1.title = "Wirelists All"
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/BOM All.xlsx")
        log_file("Creat wire BOM All.xlsx")
        messagebox.showinfo('Finalizat!')
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat wirelist BOM All.xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul BOM All.xlsx este read-only!")
        return None


def prn_excel_cutting(sheet1, sheet2, sheet3):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Lista fire"
    ws2 = wb.create_sheet("Lista fire unice KSK")
    ws3 = wb.create_sheet("Lista Fire CST")
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=sheet1[i][x])
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))
    for i in range(len(sheet2)):
        for x in range(len(sheet2[i])):
            try:
                ws2.cell(column=x + 1, row=i + 1, value=sheet2[i][x])
            except:
                ws2.cell(column=x + 1, row=i + 1, value=str(float(sheet2[i][x])))
    for i in range(len(sheet3)):
        for x in range(len(sheet3[i])):
            try:
                ws3.cell(column=x + 1, row=i + 1, value=sheet3[i][x])
            except:
                ws3.cell(column=x + 1, row=i + 1, value=str(float(sheet3[i][x])))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/Lista Cutting.xlsx")
        log_file("Creat wire Lista Cutting.xlsx")

    except PermissionError:
        log_file("Eroare salvare. Nu am salvat wirelist Lista Cutting.xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul Lista Cutting.xlsx este read-only!")
        return None


def prn_excel_module_ldorado(sheet1, title):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = title
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))

    try:
        # wb.save("D:\Fertzy\Python Projects\MAN 2022" + title + ".xlsx")
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/LDorado/" + title + ".xlsx")
        log_file("Creat")
        messagebox.showinfo('Finalizat!')
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat ")
        messagebox.showerror('Eroare scriere', "Fisierul este read-only!")
        return None


def prn_excel_raport_ksk_light(sheet1, pivotraport, array_total_ksk):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Lista KSK light"
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            ws1.cell(column=x + 1, row=i + 1, value=sheet1[i][x])
    for i in range(len(array_total_ksk)):
        for x in range(len(array_total_ksk[i])):
            ws1.cell(column=x + 6, row=i + 1, value=array_total_ksk[i][x])
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/Raport KSK Light.xlsx")
        log_file("Creat Raport KSK Light")
        with pd.ExcelWriter(os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/Raport KSK Light.xlsx",
                            mode='a') as writer:
            pivotraport.to_excel(writer, sheet_name='Raport')

    except PermissionError:
        log_file("Eroare salvare. Nu am salvat ")
        messagebox.showerror('Eroare scriere', "Fisierul este read-only!")
        return None


def prn_excel_supers_ksk(sheet1, nume):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = nume
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))

    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/SuperSleeve/KSK/" + nume + ".xlsx")
        log_file("Creat")
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat ")
        messagebox.showerror('Eroare scriere', "Fisierul este read-only!")
        return None


def prn_excel_supers_ksk_all(sheet1, sheet2, sheet3):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Toate"
    ws2 = wb.create_sheet("Lista Basic module")
    ws3 = wb.create_sheet("Lista taiere")
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=float(sheet1[i][x]))
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
    for i in range(len(sheet2)):
        for x in range(len(sheet2[i])):
            try:
                ws2.cell(column=x + 1, row=i + 1, value=float(sheet2[i][x]))
            except:
                ws2.cell(column=x + 1, row=i + 1, value=str(sheet2[i][x]))
    for i in range(len(sheet3)):
        for x in range(len(sheet3[i])):
            try:
                ws3.cell(column=x + 1, row=i + 1, value=float(sheet3[i][x]))
            except:
                ws3.cell(column=x + 1, row=i + 1, value=str(sheet3[i][x]))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/All Super Sleeve.xlsx")
        log_file("Creat")
    except PermissionError:

        log_file("Eroare salvare. Nu am salvat ")
        messagebox.showerror('Eroare scriere', "Fisierul este read-only!")
        return None


def prn_excel_supers_ksk_all_simplu(sheet1):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Lista taiere"
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=float(sheet1[i][x]))
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))

    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/Cutting Super Sleeve.xlsx")
        log_file("Creat")
    except PermissionError:

        log_file("Eroare salvare. Nu am salvat ")
        messagebox.showerror('Eroare scriere', "Fisierul este read-only!")
        return None


def prn_excel_splksklight(sheet1):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Toate"
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))

    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/All SPL KSK Light.xlsx")
        log_file("Creat")

    except PermissionError:
        log_file("Eroare salvare. Nu am salvat ")
        messagebox.showerror('Eroare scriere', "Fisierul este read-only!")
        return None


def prn_excel_wires_complete_leoni(sheet1, sheet2, sheet3, nume_fisier):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "MAN Pn"
    ws2 = wb.create_sheet("LEONI Pn1")
    ws3 = wb.create_sheet("LEONI Pn2")
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=sheet1[i][x])
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))
    for i in range(len(sheet2)):
        for x in range(len(sheet2[i])):
            try:
                ws2.cell(column=x + 1, row=i + 1, value=sheet2[i][x])
            except:
                ws2.cell(column=x + 1, row=i + 1, value=str(float(sheet2[i][x])))
    for i in range(len(sheet3)):
        for x in range(len(sheet3[i])):
            try:
                ws3.cell(column=x + 1, row=i + 1, value=sheet3[i][x])
            except:
                ws3.cell(column=x + 1, row=i + 1, value=str(float(sheet3[i][x])))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Complete BOM and WIRELIST/Wirelist/Leoni " + nume_fisier +
                ".xlsx")
        log_file("Creat wire " + nume_fisier + ".xlsx")
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat wirelist " + nume_fisier + ".xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul " + nume_fisier + "este read-only!")
        return None


def prn_excel_compare_ksk_light(sheet1, data):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Comparatie KSK Light"
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))

    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/Comparatie KSK " + data + ".xlsx")
        log_file("Creat wire Comparatie KSK Light.xlsx")
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat Comparatie KSK Light.xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul Comparatie KSK Light este read-only!")
        return None


def prn_excel_variatii(sheet1):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Variatii lungimi"
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=int(sheet1[i][x]))
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Report Files/Variatii lungimi.xlsx")
        log_file("Creat Variatii lungimi.xlsx")
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat Variatii lungimi.xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul Variatii lungimi.xlsx este read-only!")
        return None


def prn_excel_compare_ksk_lhd(sheet1, sheet2):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Comparatie KSK LHD"
    rowno = 1
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            rowno = i + 2
            ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
    for i in range(len(sheet2)):
        for x in range(len(sheet2[i])):
            ws1.cell(column=x + 1, row=i + rowno, value=str(sheet2[i][x]))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Comparatii/LHD.xlsx")
        log_file("Creat wire Comparatie KSK LHD.xlsx")
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat Comparatie KSK LHD.xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul Comparatie KSK LHD este read-only!")
        return None


def prn_excel_compare_ksk_rhd(sheet1, sheet2):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Comparatie KSK RHD"
    rowno = 1
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            rowno = i + 2
            ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
    for i in range(len(sheet2)):
        for x in range(len(sheet2[i])):
            ws1.cell(column=x + 1, row=i + rowno, value=str(sheet2[i][x]))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Comparatii/RHD.xlsx")
        log_file("Creat wire Comparatie KSK RHD.xlsx")
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat Comparatie KSK LHD.xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul Comparatie KSK RHD este read-only!")
        return None


def prn_excel_wires_light(sheet1, sheet2, sheet3, sheet4, sheet5, sheet6, sheet7, sheet8):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    wb = Workbook()
    ws1 = wb.active
    ws1.title = sheet1[1][0]
    ws2 = wb.create_sheet(sheet2[1][0])
    ws3 = wb.create_sheet(sheet3[1][0])
    ws4 = wb.create_sheet("Variatie Lungimi")
    ws5 = wb.create_sheet("Erori")
    ws6 = wb.create_sheet("Bracket")
    ws7 = wb.create_sheet("Same Wire")
    ws8 = wb.create_sheet("X6616X6490")
    ws9 = wb.create_sheet("Variatie Lungimi IST SOLL")
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            if "E-" in sheet1[i][x]:
                try:
                    ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
                except:
                    ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
            else:
                try:
                    ws1.cell(column=x + 1, row=i + 1, value=float(sheet1[i][x]))
                except:
                    ws1.cell(column=x + 1, row=i + 1, value=sheet1[i][x])
    for i in range(len(sheet2)):
        for x in range(len(sheet2[i])):
            if x == 2:
                ws2.cell(column=x + 1, row=i + 1, value=str(sheet2[i][x]))
            else:
                try:
                    ws2.cell(column=x + 1, row=i + 1, value=float(sheet2[i][x]))
                except:
                    ws2.cell(column=x + 1, row=i + 1, value=sheet2[i][x])
    for i in range(len(sheet3)):
        for x in range(len(sheet3[i])):
            if x == 2:
                ws3.cell(column=x + 1, row=i + 1, value=str(sheet3[i][x]))
            else:
                try:
                    ws3.cell(column=x + 1, row=i + 1, value=float(sheet3[i][x]))
                except:
                    ws3.cell(column=x + 1, row=i + 1, value=sheet3[i][x])
    for i in range(len(sheet4)):
        for x in range(len(sheet4[i])):
            try:
                ws4.cell(column=x + 1, row=i + 1, value=float(sheet4[i][x]))
            except:
                ws4.cell(column=x + 1, row=i + 1, value=sheet4[i][x])
    for i in range(len(sheet5)):
        for x in range(len(sheet5[i])):
            if x == 2:
                ws5.cell(column=x + 1, row=i + 1, value=str(sheet5[i][x]))
            else:
                try:
                    ws5.cell(column=x + 1, row=i + 1, value=float(sheet5[i][x]))
                except:
                    ws5.cell(column=x + 1, row=i + 1, value=sheet5[i][x])
    for i in range(len(sheet6)):
        for x in range(len(sheet6[i])):
            try:
                ws6.cell(column=x + 1, row=i + 1, value=float(sheet6[i][x]))
            except:
                ws6.cell(column=x + 1, row=i + 1, value=sheet6[i][x])
    """Cosmetica"""
    colorsgray = PatternFill(start_color='aabbcc', end_color='aabbcc', fill_type='solid')
    colorsdoi = PatternFill(start_color='7EF1EA', end_color='7EF1EA', fill_type='solid')
    for i in range(1, 7):
        for x in range(1, len(ws6['A']) + 1):
            ws6.cell(column=i, row=x).border = thin_border
            ws6.cell(column=i, row=x).alignment = Alignment(horizontal='center')
    for row in ws6['C']:
        if row.value == "RHD":
            for i in range(1, 7):
                ws6.cell(column=i, row=row.row).fill = colorsgray
        elif row.value == "LHD":
            for i in range(1, 7):
                ws6.cell(column=i, row=row.row).fill = colorsdoi
    for cells in ws1['1']:
        ws1.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws2['1']:
        ws2.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws3['1']:
        ws3.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws4['1']:
        ws4.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws5['1']:
        ws5.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws6['1']:
        ws6.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws7['1']:
        ws7.cell(column=cells.column, row=1).font = Font(bold=True)
    """Cosmetica"""
    for i in range(len(sheet7)):
        for x in range(len(sheet7[i])):
            if x == 2:
                ws7.cell(column=x + 1, row=i + 1, value=str(sheet7[i][x]))
            else:
                try:
                    ws7.cell(column=x + 1, row=i + 1, value=float(sheet7[i][x]))
                except:
                    ws7.cell(column=x + 1, row=i + 1, value=sheet7[i][x])
    for x in range(len(sheet8[0])):
        try:
            ws8.cell(column=1, row=x + 1, value=float(sheet8[0][x][0]))
            ws8.cell(column=2, row=x + 1, value=float(sheet8[0][x][1]))
        except:
            ws8.cell(column=1, row=x + 1, value=sheet8[0][x][0])
            ws8.cell(column=2, row=x + 1, value=sheet8[0][x][1])
    for x in range(len(sheet8[1])):
        try:
            ws8.cell(column=4, row=x + 1, value=float(sheet8[1][x][0]))
            ws8.cell(column=5, row=x + 1, value=float(sheet8[1][x][1]))
        except:
            ws8.cell(column=4, row=x + 1, value=sheet8[1][x][0])
            ws8.cell(column=5, row=x + 1, value=sheet8[1][x][1])
    for x in range(len(sheet8[2])):
        try:
            ws8.cell(column=7, row=x + 1, value=float(sheet8[2][x][0]))
            ws8.cell(column=8, row=x + 1, value=float(sheet8[2][x][1]))
        except:
            ws8.cell(column=7, row=x + 1, value=sheet8[2][x][0])
            ws8.cell(column=8, row=x + 1, value=sheet8[2][x][1])
    for x in range(len(sheet8[3])):
        try:
            ws8.cell(column=10, row=x + 1, value=float(sheet8[3][x][0]))
            ws8.cell(column=11, row=x + 1, value=float(sheet8[3][x][1]))
        except:
            ws8.cell(column=10, row=x + 1, value=sheet8[3][x][0])
            ws8.cell(column=11, row=x + 1, value=sheet8[3][x][1])
    for i in range(len(sheet4)):
        ws9.cell(column=1, row=i + 1, value="IST")
        for x in range(len(sheet4[i])):
            try:
                ws9.cell(column=x + 2, row=i + 1, value=float(sheet4[i][x]))
            except:
                ws9.cell(column=x + 2, row=i + 1, value=sheet4[i][x])
    ws9.cell(column=1, row=1, value="IST / SOLL")
    ws9.insert_cols(5)
    ws9.cell(column=5, row=1, value="Description")
    istsoll(ws6, ws9)
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/Beius/Prelucrate/" + sheet1[1][0] + ".xlsx")
        log_file("Creat Light" + sheet1[1][0] + ".xlsx")
    except PermissionError:
        log_file("Eroare scriere. Nu am salvat Light" + sheet1[1][0] + ".xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul " + sheet1[1][0] + "este read-only!")
        return None
    return None


def prn_excel_moduleinksk(sheet1):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "All KSK"
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=int(sheet1[i][x]))
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Lista module in KSK.xlsx")
        log_file("Lista module in KSK.xlsx")
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat Lista module in KSK.xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul Lista module in KSK.xlsx este read-only!")
        return None


def prn_excel_ksk_neprelucrate(sheet, nume_fisier):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = nume_fisier
    for i in range(len(sheet)):
        ws1.cell(column=1, row=i + 1, value=str(sheet[i]))
    if globale.director_salvare_raport == "":
        try:
            wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/Beius/Neprelucrate/" + nume_fisier + ".xlsx")
            log_file("Creat Error file " + nume_fisier + ".xlsx")
        except PermissionError:
            log_file("Eroare scriere. Nu am salvat Error file " + nume_fisier + ".xlsx")
            messagebox.showerror('Eroare scriere', "Fisierul Error file " + nume_fisier + "este read-only!")
            return None
    else:
        try:
            wb.save(globale.director_salvare_raport + "/Error file " + nume_fisier + ".xlsx")
        except PermissionError:
            messagebox.showerror('Eroare scriere', "Fisierul Error file " + nume_fisier + "este read-only!")
            return None


def prn_excel_cutting_module(sheet1, sheet2, sheet3):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Lista fire"
    ws2 = wb.create_sheet("Lista fire unice KSK")
    ws3 = wb.create_sheet("Lista Fire CST")
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=sheet1[i][x])
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))
    for i in range(len(sheet2)):
        for x in range(len(sheet2[i])):
            try:
                ws2.cell(column=x + 1, row=i + 1, value=sheet2[i][x])
            except:
                ws2.cell(column=x + 1, row=i + 1, value=str(float(sheet2[i][x])))
    for i in range(len(sheet3)):
        for x in range(len(sheet3[i])):
            try:
                ws3.cell(column=x + 1, row=i + 1, value=sheet3[i][x])
            except:
                ws3.cell(column=x + 1, row=i + 1, value=str(float(sheet3[i][x])))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Separare KSK/Lista Cutting Forecast.xlsx")
        log_file("Creat wire Lista Cutting Forecast.xlsx")
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat wirelist Lista Cutting Forecast.xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul Lista Cutting Forecast.xlsx este read-only!")
        return None


def prn_databasecontent_ksk(sheet1, ksk):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = ksk
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=sheet1[i][x])
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Database/KSK Export/" + ksk + ".xlsx")
        log_file("Creat wire Lista Cutting Forecast.xlsx")

    except PermissionError:
        log_file("Eroare salvare. Nu am salvat wirelist Lista Cutting Forecast.xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul Lista Cutting Forecast.xlsx este read-only!")
        return None


def prn_excel_export_database(sheet1):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Database"
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=sheet1[i][x])
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Database/Database.xlsx")
        log_file("Creat export database.xlsx")
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat wirelist Lista Cutting Forecast.xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul Lista Cutting Forecast.xlsx este read-only!")
        return None


def prn_excel_diagrame(sheet1, sheet2):
    wb = Workbook()
    ws1 = wb.active

    ws1.title = "Diferente diagrame"
    ws2 = wb.create_sheet("Diagrame noi")
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=sheet1[i][x])
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))
    for i in range(len(sheet2)):
        for x in range(len(sheet2[i])):
            try:
                ws2.cell(column=x + 1, row=i + 1, value=sheet2[i][x])
            except:
                ws2.cell(column=x + 1, row=i + 1, value=str(float(sheet2[i][x])))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Diferente diagrame.xlsx")
        log_file("Creat Diferente diagrame.xlsx")
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat Diferente diagrame.xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul Diferente diagrame.xlsx este read-only!")
        return None


def prn_excel_asocierediagramemodule(sheet1):
    wb = Workbook()
    ws1 = wb.active

    ws1.title = "Acociere"
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=sheet1[i][x])
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Asociere diagrame cu module din matrix.xlsx")
        log_file("Creat Asociere diagrame cu module din matrix.xlsx")
    except PermissionError:
        log_file("Eroare salvare. Nu am salvat Asociere diagrame cu module din matrix.xlsx")
        messagebox.showerror('Eroare scriere', "Fisierul Asociere diagrame cu module din matrix.xlsx este read-only!")
        return None