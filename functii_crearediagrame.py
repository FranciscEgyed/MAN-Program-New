import csv
import math
import os
import time
from tkinter import messagebox, filedialog, Tk, ttk, HORIZONTAL, Label
import json
import xml.etree.ElementTree as ET
from openpyxl.drawing.image import Image
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
from itertools import combinations
import tkinter as tk
from tkinter import ttk


def xmltojson():
    fisier_xml = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                                title="Incarcati fisierul care necesita prelucrare")
    parse_xml(fisier_xml)
    messagebox.showinfo("Finalizat", "Finalizat JSON!")

def parse_xml(file):
    def process_element(element):
        # Create a dictionary for the current tag
        tag_dict = {}

        # Store the attributes of the element in the dictionary
        tag_dict.update(element.attrib)

        # Check if the element has text content
        if element.text and element.text.strip():
            # Store the text content under the "__text__" key
            tag_dict["__text__"] = element.text.strip()

        # Process the child elements recursively
        for child in element:
            child_dict = process_element(child)
            tag_dict.setdefault(child.tag, []).append(child_dict)

        return tag_dict

    # Root elements to process
    root_element_names = ["TitleBlock", "HarnessAttributes", "Sheets", "LengthVariants", "Modules", "Families",
                          "Configurations", "Comments", "Accessories", "Fixings", "Images", "Junctions", "Segments",
                          "MultiCores", "Wires", "CavitySeals", "CavityPlugs", "Terminals", "Connectors", "Tapes",
                          "Dimensions", "Tables", "Associations", "Attributes", "FusePMDs", "RelayPMDs",
                          "AccessoryPMDs", "AssemblyPartPMDs", "WiringGroupPMDs", "HarnessConfigurationPMDs",
                          "ConnectorPMDs", "FixingPMDs", "ComponentBoxPMDs", "SymbolPMDs", "TapePMDs", "TerminalPMDs",
                          "PlugPMDs", "SealPMDs", "PartPMDs", "GeneralWirePMDs", "GeneralSpecialWirePMDs"]

    # Parse the XML file
    tree = ET.parse(file)
    for element in root_element_names:
        # Get the root element based on the given root_element_name
        root = tree.find(".//{}".format(element))

        if root is not None:
            # Process the root element
            root_dict = process_element(root)
            # Save the dictionary to a JSON file
            json_file_path = element + ".json"
            with open(os.path.abspath(os.curdir) +
                      "/MAN/Output/Diagrame/Input files XML/"+ element + ".json", "w") as json_file:
                json.dump(root_dict, json_file)
        else:
            print(f"Root element '{element}' not found in the XML file.")


def prelucrare_json():
    files = ["Modules", "LengthVariants", "Wires", "CavitySeals", "Connectors",
             "Tapes", "Terminals", "CavityPlugs", "Accessories", "Configurations", "AccessoryPMDs", "Segments" ,
             "TapePMDs"]

    def extract_wire_ids(wire_list):
        with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Input files XML/GeneralSpecialWirePMDs.json", "r") as json_file:
            gsw = json.load(json_file)
        with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Input files XML/GeneralWirePMDs.json", "r") as json_file:
            gw = json.load(json_file)
        wires = []
        for wire in wire_list['Wire']:
            cablu = "Empty"
            sectiune = "Empty"
            culoare = "Empty"
            pmd = wire['PMD']
            for fir in gsw['GeneralSpecialWirePMD']:
                for core in fir['CorePMD']:
                    if core["CustomerPartNo"] == pmd:
                        cablu = core["CustomerPartNo"]
                        sectiune = core["CSA"]
                        try:
                            culoare = core["Colour"].split("_")[1]
                        except IndexError:
                            culoare = core["Colour"]
                    else:
                        for firut in gw['GeneralWirePMD']:
                            if firut["ID"] == pmd:
                                cablu = firut["ID"]
                                sectiune = firut["CSA"]
                                try:
                                    culoare = core["Colour"].split("_")[1]
                                except IndexError:
                                    culoare = core["Colour"]
            for moduleid in wire["WireModuleRefs"][0]["WireModuleRef"]:
                wires.append([moduleid["ModuleID"], wire["ID"], wire["WireNo"], wire["RouteLength"], cablu, sectiune,
                              culoare])
        wires.insert(0, ["Module ID", "Wire ID", "Wire No", "Length", "Cablu", "Sectiune", "Culoare"])
        printfile(wires, "Wires list")

    def extract_conector_ids(wire_list):
        conectors = []
        for conector in wire_list['Connector']:
            conid = conector["ID"]
            eleID = conector["ElementID"]
            pmdID = conector["PMD"]
            juncID = conector["JunctionID"]
            txID = conector["TableX"]
            tyID = conector["TableY"]
            for ele in conector["Slots"][0]["Slot"][0]["Cavities"][0]["Cavity"]:
                pinid = ele["PinNo"]
                try:
                    plugid = ele["PlugID"]
                except KeyError:
                    plugid = "Empty"
                try:
                    for conectorwire in ele["ConnectorWire"]:
                        try:
                            wireid = conectorwire["WireID"]
                        except KeyError:
                            wireid = "Empty"
                        try:
                            terminalid = conectorwire["Terminals"][0]["Terminal"][0]["TerminalID"]
                        except KeyError:
                            terminalid = "Empty"
                        try:
                            sealid = conectorwire["Seals"][0]["Seal"][0]["SealID"]
                        except KeyError:
                            sealid = "Empty"
                        conectors.append(
                            [conid, pmdID, eleID, juncID, txID, tyID, pinid, wireid, terminalid, sealid, plugid])
                except KeyError:
                    sealid = "Empty"
                    terminalid = "Empty"
                    wireid = "Empty"

                    conectors.append(
                        [conid, pmdID, eleID, juncID, txID, tyID, pinid, wireid, terminalid, sealid, plugid])
        conectors.insert(0, ["Conector ID", "PMD", "Name", "Junction", "X coord", "Y coord", "Pin", "Wire ID", "Terminal ID",
                             "Seal ID", "Plug ID"])
        printfile(conectors, "Conectors list")

    def extract_module_ids(module_list):
        modules = []
        for module in module_list["Module"]:
            moduleid = module["ID"]
            eleID = module["TitleBlock"][0]["CustomerPartNo"]
            familyref = module["FamilyRef"]
            logisticsID = module["TitleBlock"][0]["Description"]
            modules.append([moduleid, eleID, familyref, logisticsID])
        modules.insert(0, ["Module ID", "MAN ID", "Family", "Description"])
        printfile(modules, "Modules list")

    def extract_variants_ids(variant_list):
        nume_variatii = [["Nume"]]
        for variant in variant_list["VariantNames"][0]["VariantName"]:
            varname = variant["__text__"]
            nume_variatii[0].append(varname)
        for variant in variant_list["VariantParameters"][0]["VariantParameter"]:
            valoare = []
            for value in variant["VariantParameterValue"]:
                valoare.append(value["__text__"])
            valoare.insert(0, variant["Name"])
            nume_variatii.append(valoare)
        printfile(nume_variatii, "Length Variations list")

    def extract_tape_ids(tape_list):
        tapes = []
        for tape in tape_list['Tape']:
            tapeid = tape["ID"]
            tapepmd = tape["PMD"]
            length = tape["TapeLength"]
            try:
                txID = tape["TableX"]
                tyID = tape["TableY"]
            except KeyError:
                txID = "Empty"
                tyID = "Empty"
            tapesegment = tape["TapeSegments"][0]["TapeSegment"][0]["SegmentID"]
            try:
                for ele in tape["TapeModuleRefs"][0]["TapeModuleRef"]:
                    tapes.append([tapeid, tapepmd, length, txID, tyID, ele["ModuleID"], tapesegment])
            except KeyError:
                continue
        tapes.insert(0, ["Tape ID", "PMD", "Length", "X coord", "Y coord", "Module ID", "TapeSegment"])
        printfile(tapes, "Tapes list")

    def extract_terminal_ids(terminal_list):
        terminals = []
        for terminal in terminal_list["Terminal"]:
            terminalid = terminal["ID"]
            terminalpmd = terminal["PMD"]
            for ele in terminal["TerminalModuleRefs"][0]["TerminalModuleRef"]:
                terminals.append([terminalid, terminalpmd, ele["ModuleID"]])
        terminals.insert(0, ["Terminal ID", "PMD", "Module ID"])
        printfile(terminals, "Terminals list")

    def extract_seal_ids(seal_list):
        seals = []
        for seal in seal_list["CavitySeal"]:
            sealid = seal["ID"]
            sealpmd = seal["PMD"]
            for ele in seal["CavitySealModuleRefs"][0]["CavitySealModuleRef"]:
                seals.append([sealid, sealpmd, ele["ModuleID"]])
        seals.insert(0, ["Seal ID", "PMD", "Module ID"])
        printfile(seals, "Seals list")

    def extract_plugs_ids(plug_list):
        plugs = []
        for plug in plug_list["CavityPlug"]:
            plugid = plug["ID"]
            plugpmd = plug["PMD"]
            for ele in plug["CavityPlugModuleRefs"][0]["CavityPlugModuleRef"]:
                plugs.append([plugid, plugpmd, ele["ModuleID"]])
        plugs.insert(0, ["Plug ID", "PMD", "Module ID"])
        printfile(plugs, "Plugs list")

    def extract_accessory_ids(accessory_list):
        accessorys = []
        for accessory in accessory_list["Accessory"]:
            accessoryid = accessory["ID"]
            accessorypmd = accessory["PMD"]
            accessoryconectorID = accessory["ReferencedConnectors"][0]["ReferencedConnector"][0]["ConnectorID"]
            for ele in accessory["AccessoryModuleRefs"][0]["AccessoryModuleRef"]:
                accessorys.append([accessoryid, accessorypmd, accessoryconectorID, ele["ModuleID"]])
        accessorys.insert(0, ["Accessory ID", "PMD", "ConnectorID", "Module ID"])
        printfile(accessorys, "Accessory list")


    def extract_configurations(configurations_list):
        configurati = []
        for configuration in configurations_list["Configuration"]:
            configurationid = configuration["ID"]
            configurationnickname = configuration["NickName"]
            for moduleid in configuration["ConfigurationModule"]:
                configurati.append([configurationid, configurationnickname, moduleid["ModuleID"]])
        configurati.insert(0, ["Configuration ID", "NickName", "Module ID"])
        printfile(configurati, "Configurations")


    def extract_accessorypmds(accesory_list):
        accesorypmd = []
        for accesory in accesory_list["AccessoryPMD"]:
            accesorypmdid = accesory["ID"]
            abbr = accesory["Abbreviation"]
            desc = accesory["Description"]
            accesorypmd.append([accesorypmdid, abbr, desc])
        accesorypmd.insert(0, ["AccessoryPMD ID", "Abbreviation", "Description"])
        printfile(accesorypmd, "AccessoryPMDs")

    def extract_segments(segments_list):
        segments = []
        for segment in segments_list["Segment"]:
            segmentid = segment["ID"]
            junctionid1 = segment["JunctionID1"]
            junctionid2 = segment["JunctionID2"]
            lenght = segment["Length"]
            diameter = segment["Diameter"]
            segments.append([segmentid, junctionid1, junctionid2, lenght, diameter])
        segments.insert(0, ["segmentid", "junctionid1", "junctionid2", "lenght", "diameter"])
        printfile(segments, "Segments")

    def extract_tapepmds(tapepmd_list):
        tapepmds = []
        for tape in tapepmd_list["TapePMD"]:
            tapeid = tape["ID"]
            abbr = tape["Abbreviation"]
            desc = tape["Description"]
            prottype = tape["ProtectionType"]
            tapematerial = tape["TapeMaterial"]
            tapepmds.append([tapeid, abbr, desc, prottype, tapematerial])
        tapepmds.insert(0, ["TapeID", "Abbreviation", "Description", "ProtectionType", "TapeMaterial"])
        printfile(tapepmds, "TapePMDs")

    def printfile(list_to_print, file_name):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = file_name
        for i in range(len(list_to_print)):
            for x in range(len(list_to_print[i])):
                ws1.cell(column=x + 1, row=i + 1, value=str(list_to_print[i][x]))
        try:
            wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output Excel/" + file_name + ".xlsx")
        except PermissionError:
            messagebox.showerror('Eroare scriere', "Fisierul este read-only!")
            return None

        with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/" + file_name + ".txt", 'w', newline='',
                  encoding='utf-8') as myfile:
            wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
            wr.writerows(list_to_print)




    for file in files:
        with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Input files XML/" + file + ".json", "r") as json_file:
            loaded_dictionary = json.load(json_file)

            if file == "Modules":
                extract_module_ids(loaded_dictionary)
            if file == "LengthVariants":
                extract_variants_ids(loaded_dictionary)
            if file == "Wires":
                extract_wire_ids(loaded_dictionary)
            if file == "CavitySeals":
                extract_seal_ids(loaded_dictionary)
            if file == "Connectors":
                extract_conector_ids(loaded_dictionary)
            if file == "Tapes":
                extract_tape_ids(loaded_dictionary)
            if file == "Terminals":
                extract_terminal_ids(loaded_dictionary)
            if file == "CavityPlugs":
                extract_plugs_ids(loaded_dictionary)
            if file == "Accessories":
                extract_accessory_ids(loaded_dictionary)
            if file == "Configurations":
                extract_configurations(loaded_dictionary)
            if file == "AccessoryPMDs":
                extract_accessorypmds(loaded_dictionary)
            if file == "Segments":
                extract_segments(loaded_dictionary)
            if file == "TapePMDs":
                extract_tapepmds(loaded_dictionary)

    # atasare supersleeve la conector
    #Load required data files
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Conectors list.txt", newline='') as csvfile:
        array_conectors  = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Tapes list.txt", newline='') as csvfile:
        array_tapes = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Segments.txt", newline='') as csvfile:
        array_segments = list(csv.reader(csvfile, delimiter=';'))

    conector_ss = {}
    for conector in array_conectors[1:]:
        connector_id = conector[0]
        juction_id = conector[3]
        if connector_id not in conector_ss:
            conector_ss[connector_id] = juction_id

    junctionid1 = [segment[1] for segment in array_segments]
    junctionid2 = [segment[2] for segment in array_segments]

    conector_and_ss = []
    for conector in conector_ss:
        temp_array = [conector]
        seg = conector_ss.get(conector)
        if seg in junctionid2:
            while seg in junctionid2 and temp_array[-1] != seg:
                for segment in array_segments:
                    if seg == segment[2]:
                        temp_array.append(segment[2])
                        if segment[4] != "0":
                            seg = segment[1]
                            break
            temp_array.append(seg)

        elif seg in junctionid1:
            while seg in junctionid1 and temp_array[-1] != seg:
                for segment in array_segments:
                    if seg == segment[1]:
                        temp_array.append(segment[1])
                        if segment[4] != "0":
                            seg = segment[2]
                            break
            temp_array.append(seg)
        for x in range(1, len(temp_array)):
            for segment in array_segments:
                if segment[2] == temp_array[x] or segment[1] == temp_array[x]:
                    leght = segment[3]
                    for tape in array_tapes:
                        if tape[6] == segment[0]:
                            pmd = tape[1]
            conector_and_ss.append([conector, pmd, leght])
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Conector_and_ss.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(conector_and_ss)



    messagebox.showinfo('Finalizat', "Fisierele JSON finalizate!")


def selectie_conectori():

    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Wires list.txt", newline='') as csvfile:
        array_wires = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Conectors list.txt", newline='') as csvfile:
        array_conectors  = list(csv.reader(csvfile, delimiter=';'))

    lista_fire_conector = []
    for i in range(len(array_conectors)):
        for x in range(len(array_wires)):
            if array_conectors[i][7] == array_wires[x][1]:
                lista_fire_conector.append([array_conectors[i][0], array_wires[x][1], array_wires[x][0]])

    def count_module_ids(connections, target_connector_id):
        module_ids = set()
        for connection in connections[1:]:  # Skip the header row
            connector_id2, _, module_id = connection
            if connector_id2 == target_connector_id:
                module_ids.add(module_id)
        return len(module_ids)

    def combinatii(valoare):
        combinatii_posibile = 0
        for t in range(1, valoare + 1):
            combinatii_posibile = combinatii_posibile + math.factorial(valoare) / (
                        math.factorial(t) * math.factorial(valoare - t))
        return int(combinatii_posibile)

    extracted_info = {}
    print(array_conectors)
    for line in array_conectors[1:]:
        module_count = count_module_ids(lista_fire_conector, line[0])
        connector_id = line[0]
        name = line[2]
        pin = int(line[6].replace("S", ""))  # Convert the pin to an integer
        # Check if the combination of Connector ID and Name already exists in the dictionary
        if (connector_id, name) in extracted_info:
            # Update the largest pin if the current pin is larger
            if pin > int(extracted_info[(connector_id, name)][2]):
                extracted_info[(connector_id, name)] = (connector_id, name, pin, module_count,
                                                        combinatii(module_count), int(combinatii(module_count)/180000))
        else:
            extracted_info[(connector_id, name)] = (connector_id, name, pin, module_count,
                                                    combinatii(module_count), int(combinatii(module_count)/180000))

    data = list(sorted(extracted_info.values(), key=lambda q: q[4]))

##########################################################################################################
    hederlist = ['Connector Name', 'Connector ID', 'Pin count', 'Module count', 'Number of combinations',
                 'Durata minute']

    def check_all():
        for var in checkbox_var:
            var.set(True)

    def clear_all():
        for var in checkbox_var:
            var.set(False)

    def print_selected():
        selected_items = []
        for idx, var in enumerate(checkbox_var):
            if var.get():
                selected_items.append(data[idx][0])
        input_text = input_box.get()
        root.destroy()
        creare_diagrame(selected_items, int(input_text))

    root = tk.Tk()
    root.title("Checkbox Example")
    root.geometry("800x800+50+50")

    # Create a canvas to hold the checkboxes_frame and scrollbar
    canvas = tk.Canvas(root)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # Create a frame to hold the checkboxes
    checkboxes_frame = tk.Frame(canvas)

    # Attach the checkboxes_frame to the canvas
    canvas.create_window((0, 0), window=checkboxes_frame, anchor="nw")

    # Create a vertical scrollbar
    scrollbar = ttk.Scrollbar(root, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    # Configure the canvas to work with the scrollbar
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    checkbox_var = []

    for header in hederlist:
        label_text = header
        label = tk.Label(checkboxes_frame, text=label_text)
        label.grid(row=0, column=hederlist.index(header), sticky="w")

    for idx, (connector_id, name, pin, module_count, combinations, divisions) in enumerate(data):
        var = tk.BooleanVar(root)
        var.set(False)
        checkbox = tk.Checkbutton(checkboxes_frame, text=name, variable=var)
        checkbox.grid(row=idx + 1, column=0, sticky="w")
        checkbox_var.append(var)

        label = tk.Label(checkboxes_frame, text=connector_id)
        label.grid(row=idx + 1, column=1, sticky="w")
        label = tk.Label(checkboxes_frame, text=pin)
        label.grid(row=idx + 1, column=2, sticky="w")
        label = tk.Label(checkboxes_frame, text=module_count)
        label.grid(row=idx + 1, column=3, sticky="w")
        label = tk.Label(checkboxes_frame, text=combinations)
        label.grid(row=idx + 1, column=4, sticky="w")
        label = tk.Label(checkboxes_frame, text=divisions)
        label.grid(row=idx + 1, column=5, sticky="w")

    # Configure canvas scrolling region
    canvas.config(scrollregion=canvas.bbox("all"))

    check_all_button = tk.Button(root, text="Check All", command=check_all)
    check_all_button.pack(fill="x", pady=5)

    clear_all_button = tk.Button(root, text="Clear All", command=clear_all)
    clear_all_button.pack(fill="x", pady=5)

    print_selected_button = tk.Button(root, text="Print Selected", command=print_selected)
    print_selected_button.pack(fill="x", pady=5)

    label = tk.Label(root, text="Number of tubes to process")
    label.pack(fill="x", pady=5)

    # Create an input box (Entry widget)
    input_box = tk.Entry(root)
    input_box.pack(fill="x", pady=5)

    root.mainloop()






def creare_diagrame(lista_de_prelucrat, tubeiterations):
    image_folder = filedialog.askdirectory(initialdir=os.path.abspath(os.curdir),
                                                title="Select folder with connector images:")

    #creare GUI
    pbargui = Tk()
    pbargui.title("Creare fisiere pentru diagrame")
    pbargui.geometry("1000x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    timelabel = Label(pbargui, text="Time . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2)
    timelabel.grid(row=2, column=2)
    # clear folder from old files
    test_excel = []
    for file_all in os.listdir(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Diagrame create/Conectori/Compatibili/"):
        try:
            os.remove(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Diagrame create/Conectori/Compatibili/" + file_all)
        except:
            continue
    for file_all in os.listdir(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Diagrame create/Conectori/Incompatibili/"):
        try:
            os.remove(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Diagrame create/Conectori/Incompatibili/" + file_all)
        except:
            continue
    for file_all in os.listdir(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Diagrame create/"):
        try:
            os.remove(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Diagrame create/" + file_all)
        except:
            continue
    #Load required data files
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Wires list.txt", newline='') as csvfile:
        array_wires = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Conectors list.txt", newline='') as csvfile:
        array_conectors  = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Length Variations list.txt", newline='') as csvfile:
        array_variatii = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Modules list.txt", newline='') as csvfile:
        array_modules = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Plugs list.txt", newline='') as csvfile:
        array_plugs = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Seals list.txt", newline='') as csvfile:
        array_seals = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Tapes list.txt", newline='') as csvfile:
        array_tapes = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Terminals list.txt", newline='') as csvfile:
        array_terminals = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Accessory list.txt", newline='') as csvfile:
        array_accessory = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/AccessoryPMDs.txt", newline='') as csvfile:
        array_accessorypmd = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Segments.txt", newline='') as csvfile:
        array_segments = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/TapePMDs.txt", newline='') as csvfile:
        array_tapepmds = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output TXT/Conector_and_ss.txt", newline='') as csvfile:
        array_ssconectori = list(csv.reader(csvfile, delimiter=';'))

# function needed
    output = []
    # Extragere conectori si fire
    statuslabel["text"] = "Extracting conectors and wires . . . "
    pbar['value'] += 2
    pbargui.update_idletasks()
    for i in range(len(array_conectors)):
        pbar['value'] += 2
        pbargui.update_idletasks()
        if array_conectors[i][7] == "Empty":
            temp_list = [item for item in array_conectors[i]]
            temp_list.append("Empty")
            temp_list.append("Empty")
            temp_list.append("Empty")
            output.append(temp_list)
        else:
            for x in range(len(array_wires)):
                if array_conectors[i][7] == array_wires[x][1]:
                    temp_list = [item for item in array_conectors[i]]
                    temp_list.append(array_wires[x][0])
                    temp_list.append(array_wires[x][2])
                    temp_list.append(array_wires[x][3])
                    output.append(temp_list)



        # atasare accesorii la conector
    #for i in range(1, len(output)):
    #    temp_list_acc = []
    #    for x in range(len(array_accessory)):
    #        if output[i][0] == array_accessory[x][2] and output[i][11] == array_accessory[x][3]:
    #            temp_list_acc.append(array_accessory[x][1])
    #    output[i].extend(temp_list_acc)
    # inlocuire ID Modul cu part number MAN
    statuslabel["text"] = "Replacing Module IDs . . . "
    pbar['value'] += 2
    pbargui.update_idletasks()
    for i in range(1, len(output)):
        for x in range(len(array_modules)):
            if output[i][11] == array_modules[x][0]:
                output[i][11] = array_modules[x][1].replace("PM", "81").replace("VM", "81")
                break

    # inlocuire ID terminal cu part number MAN
    for i in range(1, len(output)):
        for x in range(len(array_terminals)):
            if output[i][8] == array_terminals[x][0]:
                output[i][8] = array_terminals[x][1]
                break
    # inlocuire ID seal cu part number MAN
    for i in range(1, len(output)):
        for x in range(len(array_seals)):
            if output[i][9] == array_seals[x][0]:
                output[i][9] = array_seals[x][1]
                break
    # inlocuire ID plug cu part number MAN
    for i in range(1, len(output)):
        for x in range(len(array_plugs)):
            if output[i][10] == array_plugs[x][0]:
                output[i][10] = array_plugs[x][1]
                break
    statuslabel["text"] = "Creating unique conector list . . . "
    pbar['value'] += 2
    pbargui.update_idletasks()
    # list unica conectori
    lista_conectori = []
    lista_conectori_cu_pin = []
    for i in range(1, len(output)):
        if output[i][0] not in lista_conectori:
            lista_conectori.append(output[i][0])
    # creare lista conectori pentru diagrame
    statuslabel["text"] = "Checking conector pinings . . . "
    pbar['value'] += 2
    pbargui.update_idletasks()
    for i in range(len(lista_conectori)):
        max_pin = 0
        for x in range(len(output)):
            if lista_conectori[i] == output[x][0]:
                current_pin = int(output[x][6].replace("S", ""))
                if current_pin > max_pin:
                    max_pin = current_pin
        if (lista_conectori[i], max_pin) not in lista_conectori_cu_pin:
            lista_conectori_cu_pin.append((lista_conectori[i], max_pin))

    statuslabel["text"] = "Saving wire list . . . "
    pbar['value'] += 2
    pbargui.update_idletasks()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "output"
    for i in range(len(output)):
        statuslabel["text"] = "Writing line ... " + str(i) + "                             "
        pbar['value'] += 2
        pbargui.update_idletasks()
        for x in range(len(output[i])):
            ws1.cell(column=x + 1, row=i + 1, value=str(output[i][x]))
    try:
        wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Output Excel/Wirelist.xlsx")
    except:
        print()
    statuslabel["text"] = ""
    timelabel["text"] = ""
    pbar['value'] += 2
    pbargui.update_idletasks()


    lista_conectori_neprelucrati = []
    lista_diagrame_create = []

    #functie pentru a cauta valori identice in seturi
    def has_common_elements(set_list):
        for b in range(len(set_list)):
            for j in range(b + 1, len(set_list)):
                if len(set_list[b].intersection(set_list[j])) > 0:
                    return True
        return False

    for pereche in lista_conectori_cu_pin:
        if pereche[0] in lista_de_prelucrat:
            pbar['value'] += 2
            pbargui.update_idletasks()
            array_temp_wirelist = []
            lista_combinatii_incompatibile = []
            lista_combinatii_compatibile = []

            # variabila de prelucrare
            if 0 < pereche[1] < 100:
                for x in range(len(output)):
                    if pereche[0] == output[x][0]:
                        array_temp_wirelist.append(output[x])
            lista_module_conector = set([x[11] for x in array_temp_wirelist if x[11] != "Empty"])
            module_pins = {}
            for row in output:
                if row[11] != "Empty" and row[0] == pereche[0]:
                    module = row[11]
                    pin = row[6]
                    if module in module_pins:
                        module_pins[module].add(pin)
                    else:
                        module_pins[module] = {pin}

            # variabila de prelucrare
            for t in range(1, len(lista_module_conector) + 1):
                for combination in combinations(lista_module_conector, t):
                    pbar['value'] += 2
                    pbargui.update_idletasks()
                    if len(combination) == 1:
                        common_pins = False
                    else:
                        pins_lists = [module_pins.get(module_id, set()) for module_id in combination]
                        common_pins = has_common_elements(pins_lists)
                    if common_pins:
                        lista_combinatii_incompatibile.append(combination)
                    else:
                        lista_combinatii_compatibile.append(combination)
                    end = time.time()
                    wb = Workbook()
                    ws1 = wb.active
                    ws1.title = "Module compatibile"
                    for i in range(len(lista_combinatii_compatibile)):
                        for x in range(len(lista_combinatii_compatibile[i])):
                            try:
                                ws1.cell(column=x + 1, row=i + 1, value=float(lista_combinatii_compatibile[i][x]))
                            except:
                                ws1.cell(column=x + 1, row=i + 1, value=str(lista_combinatii_compatibile[i][x]))
                    wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Diagrame create/Conectori/Compatibili/" +
                            str(pereche[1]) + " " + pereche[0] + ".xlsx")

                    wb2 = Workbook()
                    ws21 = wb2.active
                    ws21.title = "Module incompatibile"
                    for i in range(len(lista_combinatii_incompatibile)):
                        for x in range(len(lista_combinatii_incompatibile[i])):
                            try:
                                ws21.cell(column=x + 1, row=i + 1, value=float(lista_combinatii_incompatibile[i][x]))
                            except:
                                ws21.cell(column=x + 1, row=i + 1, value=str(lista_combinatii_incompatibile[i][x]))
                    wb2.save(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Diagrame create/Conectori/Incompatibili/" +
                            pereche[0] + ".xlsx")
        else:
            lista_conectori_neprelucrati.append(pereche[0])

    wb3 = Workbook()
    ws31 = wb3.active
    ws31.title = "Conectori neprelucrati"
    for i in range(len(lista_conectori_neprelucrati)):
        try:
            ws31.cell(column=1, row=i + 1, value=float(lista_conectori_neprelucrati[i]))
        except:
            ws31.cell(column=1, row=i + 1, value=str(lista_conectori_neprelucrati[i]))
    wb3.save(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Diagrame create/Conectori neperlucrati.xlsx")

    def sort_by_first_two_characters(name):
        # Extract the first two characters of the filename
        first_two_characters = name[:2]

        # Convert the first two characters to an integer (if possible)
        try:
            return int(first_two_characters)
        except ValueError:
            # If conversion to an integer fails, return a large value
            return float('inf')

    def colorpicker(codculoare):
        if codculoare == "bk":
            cell_fill = PatternFill(start_color="000000", end_color="000000", fill_type='solid')
        elif codculoare == "bl":
            cell_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type='solid')
        elif codculoare == "ws":
            cell_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid')
        elif codculoare == "rs":
            cell_fill = PatternFill(start_color="ff99ff", end_color="ff99ff", fill_type='solid')
        elif codculoare == "bn":
            cell_fill = PatternFill(start_color="996600", end_color="996600", fill_type='solid')
        elif codculoare == "rd":
            cell_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type='solid')
        else:
            cell_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid')
        return cell_fill


    contents = os.listdir(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Diagrame create/Conectori/Compatibili/")
    # Sort the contents based on the first two characters as numbers
    contents.sort(key=sort_by_first_two_characters)
    lista_fire = output[1:]
    counter_print = 0
    for i in range(len(contents)):
        counter_print = counter_print + 1
        timelabel["text"] = "       Printing " + str(counter_print) + " from " + str(len(contents))
        pbar['value'] += 2
        pbargui.update_idletasks()
        counter_diagrama = 1
        wb = load_workbook(os.path.abspath(os.curdir) +
                           "/MAN/Output/Diagrame/Diagrame create/Conectori/Compatibili/" + contents[i])
        ws1 = wb.worksheets[0]
        row_count = 1
        conectoropus = "WIP"
        numeconector = contents[i][contents[i].find(' ') + 1:-5]
        for row in ws1.iter_rows():
            lista_moduleids = []
            output_accesorii = [["MAN Pn", "LEONI Pn", "Descriere"]]
            lista_module_diagrama = []
            output_diagrama = {key: None for key in range(1, int(contents[i][0:contents[i].find(' ')]) + 1)}
            # Iterate through worksheet and print cell contents
            for cell in row:
                if cell.value is not None:
                    lista_module_diagrama.append(cell.value)
            for key in output_diagrama.keys():
                for conector in array_conectors:
                    if conector[0] == numeconector and conector[7] != "Empty":
                        fir = conector[7]
                        for wire in array_wires:
                            if wire[1] == fir:
                                modul = wire[0]
                                for plug in array_plugs:
                                    if plug[2] == modul:
                                        numarfir = plug[1]
                                        break
                modulul = ""
                culoare = "bk"
                codcul1 = "bk"
                codcul2 = "bk"
                sectiune = ""

                for fir in lista_fire:
                    if int(fir[6].replace("S", "")) == key and fir[11] in lista_module_diagrama and\
                            fir[0] == numeconector:
                        modulul = fir[11]
                        for v in range(len(array_modules)):
                            if array_modules[v][1] == modulul:
                                modulid = array_modules[v][0]
                                lista_moduleids.append(modulid)
                                break
                        numarfir = fir[12]
                        for wire in array_wires:
                            if wire[1] == fir[7]:
                                culoare = wire[6]
                                codcul1 = wire[6][:2]
                                if wire[6][2:] != "":
                                    codcul2 = wire[6][2:]
                                else:
                                    codcul2 = wire[6][:2]
                                sectiune = wire[5]

                output_diagrama[key] = modulul, numarfir, culoare, codcul1, codcul2, sectiune, conectoropus
            output_diagrama[0] = "Modul PN", "Imprimare Fir", "Culoare", "", "", "Sectiune", "Conector opus"

            # lista cu accesorii pentru diagrama
            lista_abrevieri_ss = ["G44", "G43", "G42", "G45", "G46", "04.37161-9133", "04.37161-9123", "04.37161-9115",
                                  "04.37161-9144", "04.37161-9149"]
            for q in range(1, len(array_accessory)):
                if array_accessory[q][2] == numeconector:
                    for w in range(1, len(array_accessorypmd)):
                        if array_accessorypmd[w][0] == array_accessory[q][1]:
                            accesory_descr = array_accessorypmd[w][2]
                            break
                        else:
                            accesory_descr = "None found"
                    if [array_accessory[q][1], "leoni", accesory_descr] not in output_accesorii:
                        output_accesorii.append([array_accessory[q][1], "leoni", accesory_descr])
            lista_tuburi_diagrama = []
            for w in range(len(array_ssconectori)):

                if (numeconector == array_ssconectori[w][0] and array_ssconectori[w][1] in lista_abrevieri_ss and
                        int(array_ssconectori[w][2]) > 60):
                    lista_tuburi_diagrama.append([array_ssconectori[w][1], "leoni", array_ssconectori[w][2]])
            for x in range(tubeiterations):
                try:
                    output_accesorii.append(lista_tuburi_diagrama[x])
                except IndexError:
                    continue

            wb = Workbook()
            ws1 = wb.active
            ws1.title = "V" + str(counter_diagrama)
            # Write conector information
            nume_conector = ""
            partman_conector = ""
            descrierecon = ""
            for con in array_conectors:
                if con[0] == numeconector:
                    partman_conector = con[1]
                    nume_conector = con[2]
                    break
            for pmd in array_accessorypmd:
                if pmd[0] == partman_conector:
                    descrierecon = pmd[2]
                    break

            ws1.cell(row=1, column=2, value=nume_conector)
            ws1.cell(row=2, column=2, value=partman_conector)
            ws1.cell(row=3, column=2, value=descrierecon)

            # Write the data from the dictionary to the Excel worksheet
            for row_num, row_data in output_diagrama.items():
                ws1.cell(row=row_num + 4, column=1, value=row_data[0])  # Column 1: '81.25481-7038'
                ws1.cell(row=row_num + 4, column=2, value=row_num)  # Column 2: Key
                ws1.cell(row=row_num + 4, column=3, value=row_data[1])  # Column 3: 'blau_001'
                ws1.cell(row=row_num + 4, column=4, value=row_data[2])  # Column 4: 'bl'
                ws1.cell(row=row_num + 4, column=5, value="").fill = colorpicker(row_data[3])  # Column 5: 'bl'
                ws1.cell(row=row_num + 4, column=6, value="").fill = colorpicker(row_data[4])  # Column 6: 'bl'
                ws1.cell(row=row_num + 4, column=7, value=row_data[5])  # Column 7: '1'
                ws1.cell(row=row_num + 4, column=8, value=row_data[6])  # Column 8: 'WIP'

            for e in range(len(output_accesorii)):
                for r in range(len(output_accesorii[e])):
                    ws1.cell(row=20 + e, column=9+r, value=output_accesorii[e][r])


            # creare diagrama
            # Replace with the image file name you want to insert
            image_filename = partman_conector + "@connector.png"
            image_path = os.path.join(image_folder, image_filename)
            if os.path.exists(image_path):
                img = Image(image_path)
                img.width = 500  # Set the desired width
                img.height = 375  # Set the desired height
                ws1.add_image(img, 'I1')

            for row_num, row_data in output_diagrama.items():
                if row_num > 0:
                    image_filenamefir = partman_conector + "@" + str(row_num) + "@0.png"
                    image_path = os.path.join(image_folder, image_filenamefir)
                    if os.path.exists(image_path):
                        img = Image(image_path)
                        img.width = 500  # Set the desired width
                        img.height = 375  # Set the desired height
                        ws1.add_image(img, 'I1')
            try:
                wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Diagrame create/" +
                        nume_conector + " V" + str(counter_diagrama)  + ".xlsx")
                row_count = row_count + 1
                lista_diagrame_create.append([nume_conector + " V" + str(counter_diagrama),
                                          lista_module_diagrama])
                counter_diagrama = counter_diagrama + 1
            except PermissionError:
                messagebox.showerror('Read Only file', nume_conector + " V" + str(counter_diagrama)  + ".xlsx")

    timelabel["text"] = "     Printing final list in EXCEL"
    pbar['value'] += 2
    pbargui.update_idletasks()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Module compatibile"
    for i in range(len(lista_diagrame_create)):
        for x in range(len(lista_diagrame_create[i])):
            try:
                ws1.cell(column=x + 1, row=i + 1, value=float(lista_diagrame_create[i][x]))
            except:
                ws1.cell(column=x + 1, row=i + 1, value=str(lista_diagrame_create[i][x]))
    wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Diagrame/Diagrame create/Lista diagrame create.xlsx")
    end = time.time()
    print("FINISH")
    messagebox.showinfo('Finalizat', "!!!!!!!!!!!!!!")
    pbargui.destroy()

