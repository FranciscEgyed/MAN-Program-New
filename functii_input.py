import csv
import os
from tkinter import Tk, ttk, HORIZONTAL, Label, filedialog, messagebox
import time
from openpyxl.reader.excel import load_workbook
from diverse import log_file
from functii_print import prn_cm_to_excel


def load_source():
    file_load = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                           title="Incarcati fisierul cu informatiile sursa:")
    wb = load_workbook(file_load)
    ws1 = wb["Sortare module"]
    ws2 = wb["Bracket Side"]
    ws3 = wb["Klappschalle"]
    ws4 = wb["BKK"]
    ws5 = wb["Module implementate"]
    ws6 = wb["Längenmodule"]
    ws7 = wb["Combinatii sectiuni"]
    ws8 = wb["Heckmodule"]
    ws9 = wb["Module excluse"]
    ws10 = wb["MY2023"]
    ws11 = wb["Prufung"]
    ws12 = wb["CKD"]
    ws13 = wb["Module active"]
    ws14 = wb["KSKLight"]
    ws15 = wb["Supersleeve"]
    ws16 = wb["SS Database"]
    ws17 = wb["Component Overview"]
    ws18 = wb["ETE"]
    ws19 = wb["TerminaliSubby"]

    'sortare module'
    array_write = [[], [], []]
    for row in ws1['A']:
        if row.value is not None:
            array_write[0].append(row.value)
    for row in ws1['B']:
        if row.value is not None:
            array_write[1].append(row.value)
    for row in ws1['C']:
        if row.value is not None:
            array_write[2].append(row.value)
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Sortare Module.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel sortare module")

    'moduleimplementate'
    array_write = []
    for row in ws5['A']:
        if row.value != "part" and row.value is not None:
            array_write.append([row.value, ws5.cell(row=row.row, column=2).value])
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Module Implementate.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel module implementate")

    'moduleactive'
    array_write = []
    for row in ws13['B']:
        if row.value != "Moduls PN" and row.value is not None:
            array_write.append([row.value, ws13.cell(row=row.row, column=3).value,
                                ws13.cell(row=row.row, column=4).value, ws13.cell(row=row.row, column=5).value])
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Module Active.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel module active")

    'load_langenmodule'
    array_write = []
    for row in ws6['A']:
        if row.value != "Side" and row.value is not None:
            array_write.append([row.value, ws6.cell(row=row.row, column=2).value, ws6.cell(row=row.row, column=3).value,
                                ws6.cell(row=row.row, column=4).value, ws6.cell(row=row.row, column=5).value,
                                ws6.cell(row=row.row, column=6).value, ws6.cell(row=row.row, column=7).value,
                                ws6.cell(row=row.row, column=8).value, ws6.cell(row=row.row, column=9).value,
                                ws6.cell(row=row.row, column=10).value, ws6.cell(row=row.row, column=11).value,
                                ws6.cell(row=row.row, column=12).value, ws6.cell(row=row.row, column=13).value,
                                ws6.cell(row=row.row, column=14).value, ws6.cell(row=row.row, column=15).value,
                                ws6.cell(row=row.row, column=16).value])
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Langenmodule.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel langenmodule")

    'load_combinatii'
    array_write = []
    for row in ws7['A']:
        if row.value != "Combinatie" and row.value is not None:
            array_write.append([row.value, ws7.cell(row=row.row, column=2).value])
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Combinatii sectiuni.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel combinatii sectiuni")

    'load_heck'
    array_write = []
    for row in ws8['A']:
        if row.value is not None:
            array_write.append(
                [row.value, ws8.cell(row=row.row, column=2).value, ws8.cell(row=row.row, column=3).value])
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Heck Modules.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel heckmodules")

    'incarcare_sidebracket'
    array_write = [[], []]
    for row in ws2['A']:
        if row.value != "LHD" and row.value is not None:
            array_write[0].append(row.value)
    for row in ws2['B']:
        if row.value != "RHD" and row.value is not None:
            array_write[1].append(row.value)
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Bracket Side.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel side bracket")

    'incarcare_klappschale'
    array_write = []
    for row in ws3['A']:
        if row.value != "harnessCC" and row.value is not None:
            array_write.append([row.value, ws3.cell(row=row.row, column=2).value, ws3.cell(row=row.row, column=3).value,
                                ws3.cell(row=row.row, column=4).value, ws3.cell(row=row.row, column=5).value])
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Tabel klappschale.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel klappschalle")

    'incarcare_bkk'
    array_write = []
    for row in ws4['A']:
        if row.value != "COA" and row.value is not None:
            array_write.append([row.value, ws4.cell(row=row.row, column=2).value])
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Tabel BKK.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel module BKK")

    'module_excluse'
    array_write = [[]]
    for row in ws9['A']:
        if row.value != "Module excluse" and row.value is not None:
            array_write[0].append(row.value)
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Module Excluse.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel module excluse")

    'module_my2023'
    array_write = []
    for row in ws10['A']:
        if row.value != "Module" and row.value is not None:
            array_write.append([row.value, ws10.cell(row=row.row, column=2).value])
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Module MY2023.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel module MY2023")

    'prufung'
    array_write = []
    for row in ws11['A']:
        if row.value != "Modul 1" and row.value is not None:
            array_write.append([row.value, ws11.cell(row=row.row, column=2).value,
                                ws11.cell(row=row.row, column=3).value, ws11.cell(row=row.row, column=4).value,
                                ws11.cell(row=row.row, column=5).value])
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Prufung.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel Prufung")

    'load_ckd'
    array_write = [[]]
    for row in ws12['A']:
        if row.value != "CKD" and row.value is not None:
            array_write[0].append(row.value)
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/CKD.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel CKD")

    'load_KSKLight'
    array_write = [[]]
    for row in ws14['A']:
        if row.value != "CKD" and row.value is not None:
            array_write[0].append(row.value)
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/KSKLight.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    array_write = [[]]
    for row in ws14['B']:
        if row.value != "CKD" and row.value is not None:
            array_write[0].append(row.value)
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/KSKLight+.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)

    log_file("Incarcat excel KSKLight")

    'load_Supersleeve'
    array_write = [[]]
    for row in ws15['B']:
        if row.value != "Sachnummer" and row.value is not None:
            array_write[0].append(row.value)
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Supersleeve.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel Supersleeve")

    'load_SS database'
    array_write = []
    array_write2 = []
    for row in ws16['A']:
        if row.value is not None:
            array_write.append([ws16.cell(row=row.row, column=1).value, ws16.cell(row=row.row, column=2).value,
                                ws16.cell(row=row.row, column=3).value, ws16.cell(row=row.row, column=2).value,
                                ws16.cell(row=row.row, column=5).value, ws16.cell(row=row.row, column=10).value,
                                ws16.cell(row=row.row, column=13).value, ws16.cell(row=row.row, column=14).value])
            if ws16.cell(row=row.row, column=4).value == "81.25480-8011" or \
                    ws16.cell(row=row.row, column=4).value == "81.25480-8013":
                array_write2.append([ws16.cell(row=row.row, column=13).value, ws16.cell(row=row.row, column=4).value])
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/SSDatabase.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/SSDrawings.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write2)
    log_file("Incarcat excel SSDatabase")

    'load_component overview'
    array_write = []
    for row in ws17['A']:
        if row.value is not None:
            array_write.append([ws17.cell(row=row.row, column=1).value, ws17.cell(row=row.row, column=2).value,
                                ws17.cell(row=row.row, column=3).value])
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Component Overview.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel Component Overview")

    'load_ETE'
    array_write = [[]]
    for row in ws18['A']:
        if row.value is not None:
            array_write[0].append(row.value)
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/ETE.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel ETE")

    'load_terminali'
    array_write = [[]]
    for row in ws19['A']:
        if row.value is not None:
            array_write[0].append(row.value)
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Terminalisubby.txt", 'w', newline='',
              encoding='utf-8') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
        wr.writerows(array_write)
    log_file("Incarcat excel ETE")
    messagebox.showinfo("Finalizat", "Finalizat!")
    return None


def cmtoexcel():
    pbargui = Tk()
    pbargui.title("Trasfer CM to one EXCEL file")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    control_matrix = []
    try:
        with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Control_Matrix_CSL.txt", newline='',
                  encoding="utf8") as csvfile:
            control_matrix_csl = list(csv.reader(csvfile, delimiter=';'))
        with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Control_Matrix_CSR.txt", newline='',
                  encoding="utf8") as csvfile:
            control_matrix_csr = list(csv.reader(csvfile, delimiter=';'))
        with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Control_Matrix_TGLML.txt", newline='',
                  encoding="utf8") as csvfile:
            control_matrix_tgl = list(csv.reader(csvfile, delimiter=';'))
        with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Control_Matrix_TGLMR.txt", newline='',
                  encoding="utf8") as csvfile:
            control_matrix_tgr = list(csv.reader(csvfile, delimiter=';'))
        with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Control_Matrix_4AXELL.txt", newline='',
                  encoding="utf8") as csvfile:
            control_matrix_4al = list(csv.reader(csvfile, delimiter=';'))
        with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Control_Matrix_4AXELR.txt", newline='',
                  encoding="utf8") as csvfile:
            control_matrix_4ar = list(csv.reader(csvfile, delimiter=';'))
    except FileNotFoundError:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showerror('Eroare fisier', 'Lipsa fisier Control_Matrix din INPUT')
        return None
    for i in range(len(control_matrix_csl)):
        if control_matrix_csl[i][1] != "":
            control_matrix.append(control_matrix_csl[i])
    pbar['value'] += 2
    pbargui.update_idletasks()
    for i in range(len(control_matrix_csr)):
        if control_matrix_csr[i][1] != "":
            control_matrix.append(control_matrix_csr[i])
    pbar['value'] += 2
    pbargui.update_idletasks()
    for i in range(len(control_matrix_tgl)):
        if control_matrix_tgl[i][1] != "":
            control_matrix.append(control_matrix_tgl[i])
    pbar['value'] += 2
    pbargui.update_idletasks()
    for i in range(len(control_matrix_tgr)):
        if control_matrix_tgr[i][1] != "":
            control_matrix.append(control_matrix_tgr[i])
    pbar['value'] += 2
    pbargui.update_idletasks()
    for i in range(len(control_matrix_4al)):
        if control_matrix_4al[i][1] != "":
            control_matrix.append(control_matrix_4al[i])
    pbar['value'] += 2
    pbargui.update_idletasks()
    for i in range(len(control_matrix_4ar)):
        if control_matrix_4ar[i][1] != "":
            control_matrix.append(control_matrix_4ar[i])
    pbar['value'] += 2
    pbargui.update_idletasks()
    pbar['value'] += 2
    statuslabel["text"] = "Printing file                     "
    pbargui.update_idletasks()
    prn_cm_to_excel(control_matrix)
    pbar.destroy()
    pbargui.destroy()
    messagebox.showinfo('Finalizat!', "Fisierul este gata !")


def cmss():
    pbargui = Tk()
    pbargui.title("Control Matrix Super Sleeve")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    timelabel = Label(pbargui, text="Time . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    timelabel.grid(row=2, column=2)

    file_load = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                           title="Incarcati fisierul Control Matrix Super Sleeve:")
    wb = load_workbook(file_load)
    statuslabel["text"] = "Loading excel file . . . "
    pbargui.update_idletasks()
    ws1 = wb["BASICK_Module"]
    ws2 = wb["Matrix SS_location"]
    file_counter = len(ws1['A'])
    file_progres = 0
    array_write = []
    start0 = time.time()
    for row in ws1['A']:
        if row.value is not None:
            array_write.append([ws1.cell(row=row.row, column=1).value, ws1.cell(row=row.row, column=2).value,
                                ws1.cell(row=row.row, column=3).value])
            end0 = time.time()
            file_progres += 1
            pbar['value'] += 2
            statuslabel["text"] = "Inregistrari parcurse " + str(file_progres) + "/" + str(file_counter)
            timelabel["text"] = "Estimated time to complete : " + \
                                str(((file_counter * 0.00095) - (end0 - start0)) / 60)[:5] + " minutes."
            pbargui.update_idletasks()

    array_write[0].extend([ws2.cell(row=2, column=1).value, ws2.cell(row=2, column=6).value,
                           ws2.cell(row=2, column=8).value, ws2.cell(row=2, column=9).value,
                           ws2.cell(row=2, column=11).value, ws2.cell(row=2, column=12).value,
                           ws2.cell(row=2, column=13).value, ws2.cell(row=2, column=14).value,
                           ws2.cell(row=2, column=2).value, "Nume", "Lungime mm"])
    # file_counter = len(ws2['AA']) * len(array_write) + len(ws2['AC']) * len(array_write)
    file_counter = 200
    print(file_counter)
    file_progres = 0
    start1 = time.time()
    for i in range(0, 200):  # len(array_write)):
        start2 = time.time()
        for row in ws2['AA']:
            if row.value is not None and row.value == array_write[i][0]:
                array_write[i].extend([ws2.cell(row=row.row, column=1).value, ws2.cell(row=row.row, column=6).value,
                                       ws2.cell(row=row.row, column=8).value, ws2.cell(row=row.row, column=9).value,
                                       ws2.cell(row=row.row, column=11).value, ws2.cell(row=row.row, column=12).value,
                                       ws2.cell(row=row.row, column=13).value, ws2.cell(row=row.row, column=14).value,
                                       ws2.cell(row=row.row, column=2).value, ws2.cell(row=row.row, column=10).value])
        for row in ws2['AC']:
            if row.value is not None and row.value == array_write[i][0]:
                array_write[i].extend([ws2.cell(row=row.row, column=1).value, ws2.cell(row=row.row, column=6).value,
                                       ws2.cell(row=row.row, column=8).value, ws2.cell(row=row.row, column=9).value,
                                       ws2.cell(row=row.row, column=11).value, ws2.cell(row=row.row, column=12).value,
                                       ws2.cell(row=row.row, column=13).value, ws2.cell(row=row.row, column=14).value,
                                       ws2.cell(row=row.row, column=2).value, ws2.cell(row=row.row, column=10).value])
        end1 = time.time()
        print(end1 - start2)
        file_progres += 2
        pbar['value'] += 2
        statuslabel["text"] = "Inregistrari parcurse " + str(file_progres) + "/" + str(file_counter)
        timelabel["text"] = "Estimated time to complete : " + \
                            str((file_counter * 0.035) - (end1 - start1))[:5] + " minutes."
        pbargui.update_idletasks()
        end3 = time.time()
        print(end3 - start1)


def cmcsrnew():
    pbargui = Tk()
    pbargui.title("Control Matrix CSR")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    listatwist = ["131_002", "131_102", "grau_047"]
    fisier_cm = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                           title="Incarcati fisierul control matrix:")
    array_print = [["Module ID+REAL NAME", "KANBAN-AG", "REAL NAME", "Kanban name", "Module ID", "Ledset",
                    "Type", "Material PN", "Conector 1", "Pin 1", "Conector 2", "Pin 2", "Splice", "Ferc", "X/Y",
                    "Twist", "PB place", "My DWG", "DWG", "APAB_1", "APAB_2", "APAB_3", "APAB_4",
                    "Що змінилось", "(CW)", "ЗМІНА"]]
    idx_conector = []
    idx_pin = []
    idx_module = []
    idx_knname = ""
    idx_realname = ""
    idx_leadset = ""
    idx_kanbanag = ""
    idx_pnmaterial = ""
    idx_splice = ""
    idx_ferc = ""
    idx_xy = ""
    idx_twist = ""
    idx_pbplace = ""
    idx_mydwg = ""
    idx_dwg = ""
    idx_apab1 = ""
    idx_apab2 = ""
    idx_apab3 = ""
    idx_apab4 = ""
    idx_apab5 = ""
    idx_aem1 = ""
    idx_aem2 = ""
    idx_aem3 = ""

    if fisier_cm[-3:] == "csv":
        with open(fisier_cm, newline='') as csvfile:
            array_sortare = list(csv.reader(csvfile, delimiter=','))
        if "8014" in array_sortare[0] or "8O14" in array_sortare[0]:
            for i, j in enumerate(array_sortare[5]):
                if j == 'Kurzname':
                    idx_conector.append(i)
                elif j == 'Pin':
                    idx_pin.append(i)
                elif j == 'KN Name':
                    idx_knname = i
                elif j == 'Real Name':
                    idx_realname = i
                elif j == 'Ledset':
                    idx_leadset = i
                elif j == 'KANBAN-AG':
                    idx_kanbanag = i
                elif j == 'PN':
                    idx_pnmaterial = i
                elif j == 'SPLICE':
                    idx_splice = i
                elif j == 'Ferc':
                    idx_ferc = i
                elif j == 'X/Y/S':
                    idx_xy = i
                elif j == 'Twist':
                    idx_twist = i
                elif 'PB_PLACE' in j:
                    idx_pbplace = i
                elif j == 'MY/ DWG':
                    idx_mydwg = i
                elif j == 'DWG':
                    idx_dwg = i
                elif j == 'APAB_1':
                    idx_apab1 = i
                elif j == 'APAB_2':
                    idx_apab2 = i
                elif j == 'APAB_3':
                    idx_apab3 = i
                elif j == 'APAB_4':
                    idx_apab4 = i
                elif j == 'APAB_5':
                    idx_apab5 = i
                elif j == "????????":
                    idx_aem1 = i
                elif j == 'CW47':
                    idx_aem2 = i
                elif j == '??????':
                    idx_aem3 = i
            for i, j in enumerate(array_sortare[3]):
                if len(j) == 13:
                    idx_module.append(i)

            statuslabel["text"] = "Working on it . . . "
            pbar['value'] += 2
            pbargui.update_idletasks()

            for i in range(6, len(array_sortare)):
                try:
                    for x in range(0, len(idx_module)):
                        if array_sortare[i][idx_module[x]] == "X" or array_sortare[i][idx_module[x]] == "x":
                            array_print.append([array_sortare[3][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname], array_sortare[3][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "FIR",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                        elif array_sortare[i][idx_module[x]] == "Y" or array_sortare[i][idx_module[x]] == "y":
                            array_print.append([array_sortare[3][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname], array_sortare[3][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "OPERATIE",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                        elif (array_sortare[i][idx_module[x]] == "S" or array_sortare[i][idx_module[x]] == "s")\
                                and array_sortare[i][idx_realname] in listatwist:
                            array_print.append([array_sortare[3][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i + 4][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i + 4][idx_knname], array_sortare[3][idx_module[x]],
                                                array_sortare[i + 4][idx_leadset].upper().replace("23U", "23W"), "FIR",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                        elif (array_sortare[i][idx_module[x]] == "S" or array_sortare[i][idx_module[x]] == "s")\
                                and array_sortare[i][idx_realname] not in listatwist:
                            array_print.append([array_sortare[3][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname], array_sortare[3][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "FIRCOMP",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                except:
                    pbar.destroy()
                    pbargui.destroy()
                    messagebox.showerror('Eroare cap de tabel', "Cap de tabel diferit de cel din lista")
            lista_leadset = []
            for i in range(len(array_print)):
                if array_print[i][5] != "":
                    lista_leadset.append([array_print[i][5], array_print[i][4], array_print[i][3],
                                         array_print[i][4] + array_print[i][3]])
            for i in range(1, len(array_print)):
                for x in range(len(lista_leadset)):
                    if array_print[i][4] + array_print[i][3] == lista_leadset[x][3]:
                        array_print[i][1] = lista_leadset[x][0].split("#", 1)[0]
                        statuslabel["text"] = "Replacing leadsets to RO . . . "
                        pbar['value'] += 2
                        pbargui.update_idletasks()
            statuslabel["text"] = "Printing file . . . "
            pbar['value'] += 2
            pbargui.update_idletasks()
            with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Control_Matrix_CSR.txt", 'w', newline='',
                      encoding='utf-8') as myfile:
                wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
                wr.writerows(array_print)
            pbar.destroy()
            pbargui.destroy()
            messagebox.showinfo('Finalizat!', "Finalizat.")
        else:
            pbar.destroy()
            pbargui.destroy()
            messagebox.showerror('Fisier gresit!', "Nu ati incarcat fisierul CSR")
    else:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showerror('Extensie gresita!', "Incarcati fisierul CSV")


def cmcslnew():
    pbargui = Tk()
    pbargui.title("Control Matrix CSL")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    listatwist = ["131_002", "131_102", "grau_047"]
    fisier_cm = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                           title="Incarcati fisierul control matrix:")
    array_print = [["Module ID+REAL NAME", "KANBAN-AG", "REAL NAME", "Kanban name", "Module ID", "Ledset",
                    "Type", "Material PN", "Conector 1", "Pin 1", "Conector 2", "Pin 2", "Splice", "Ferc", "X/Y",
                    "Twist", "Coiling", "PB place", "8011", "8012", "APAB_1", "APAB_2", "APAB_3", "APAB_4",
                    "Що змінилось", "(CW)", "ЗМІНА"]]

    idx_conector = []
    idx_pin = []
    idx_module = []
    idx_knname = ""
    idx_realname = ""
    idx_leadset = ""
    idx_kanbanag = ""
    idx_pnmaterial = ""
    idx_splice = ""
    idx_ferc = ""
    idx_xy = ""
    idx_twist = ""
    idx_coiling = ""
    idx_pbplace = ""
    idx_8011 = ""
    idx_8012 = ""
    idx_apab1 = ""
    idx_apab2 = ""
    idx_apab3 = ""
    idx_apab4 = ""
    idx_aem1 = ""
    idx_aem2 = ""
    idx_aem3 = ""
    if fisier_cm[-3:] == "csv":
        with open(fisier_cm, newline='') as csvfile:
            array_sortare = list(csv.reader(csvfile, delimiter=','))
        print(array_sortare[5][27])
        if "8011" in array_sortare[0]:
            for i, j in enumerate(array_sortare[2]):
                if j == 'von' or j == 'nach':
                    idx_conector.append(i)
                elif j == 'Pin':
                    idx_pin.append(i)
                elif j == 'Kanban name':
                    idx_knname = i
                elif j == 'REAL NAME':
                    idx_realname = i
                elif j == 'Leadset':
                    idx_leadset = i
                elif j == 'Actual KanbanAG':
                    idx_kanbanag = i
                elif j == 'FORS PN':
                    idx_pnmaterial = i
                elif j == 'Splice':
                    idx_splice = i
                elif j == 'Ferc':
                    idx_ferc = i
                elif j == 'X/Y':
                    idx_xy = i
                elif j == 'Twist':
                    idx_twist = i
                elif j == 'Coiling':
                    idx_coiling = i
                elif '2nd Place' in j:
                    idx_pbplace = i
                elif j == '8011':
                    idx_8011 = i
                elif j == '8012':
                    idx_8012 = i
                elif j == 'APAB_1':
                    idx_apab1 = i
                elif j == 'APAB_2':
                    idx_apab2 = i
                elif j == 'APAB_3':
                    idx_apab3 = i
                elif j == 'APAB_4':
                    idx_apab4 = i
                elif j == "?? ?????????":
                    idx_aem1 = i
                elif j == '(CW)':
                    idx_aem2 = i
                elif j == '?????':
                    idx_aem3 = i
            for i, j in enumerate(array_sortare[1]):
                if len(j) == 13:
                    idx_module.append(i)
            statuslabel["text"] = "Working on it . . . "
            pbar['value'] += 2
            pbargui.update_idletasks()
            for i in range(6, len(array_sortare)):
                try:
                    for x in range(0, len(idx_module)):
                        if array_sortare[i][idx_module[x]] == "X" or array_sortare[i][idx_module[x]] == "x":
                            array_print.append([array_sortare[1][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname].upper(), array_sortare[1][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "FIR",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist], array_sortare[i][idx_coiling],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_8011],
                                                array_sortare[i][idx_8012], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_aem1],
                                                array_sortare[i][idx_aem2], array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                        elif array_sortare[i][idx_module[x]] == "Y" or array_sortare[i][idx_module[x]] == "y":
                            array_print.append([array_sortare[1][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname].upper(), array_sortare[1][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "OPERATIE",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist], array_sortare[i][idx_coiling],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_8011],
                                                array_sortare[i][idx_8012], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_aem1],
                                                array_sortare[i][idx_aem2], array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                        elif (array_sortare[i][idx_module[x]] == "S" or array_sortare[i][idx_module[x]] == "s")\
                                and array_sortare[i][idx_realname] in listatwist:
                            array_print.append([array_sortare[1][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i + 4][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i + 4][idx_knname].upper(), array_sortare[1][idx_module[x]],
                                                array_sortare[i + 4][idx_leadset].upper().replace("23U", "23W"), "FIR",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist], array_sortare[i][idx_coiling],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_8011],
                                                array_sortare[i][idx_8012], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_aem1],
                                                array_sortare[i][idx_aem2], array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                        elif (array_sortare[i][idx_module[x]] == "S" or array_sortare[i][idx_module[x]] == "s")\
                                and array_sortare[i][idx_realname] not in listatwist:
                            array_print.append([array_sortare[1][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname].upper(), array_sortare[1][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "FIRCOMP",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist], array_sortare[i][idx_coiling],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_8011],
                                                array_sortare[i][idx_8012], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_aem1],
                                                array_sortare[i][idx_aem2], array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()

                except:
                    pbar.destroy()
                    pbargui.destroy()
                    messagebox.showerror('Eroare cap de tabel', "Cap de tabel diferit de cel din lista")
            lista_leadset = []
            for i in range(len(array_print)):
                if array_print[i][5] != "":
                    lista_leadset.append([array_print[i][5], array_print[i][4], array_print[i][3],
                                         array_print[i][4] + array_print[i][3]])
            for i in range(1, len(array_print)):
                for x in range(len(lista_leadset)):
                    if array_print[i][4] + array_print[i][3] == lista_leadset[x][3]:
                        array_print[i][1] = lista_leadset[x][0].split("#", 1)[0]
                        statuslabel["text"] = "Replacing leadsets to RO . . . "
                        pbar['value'] += 2
                        pbargui.update_idletasks()
            statuslabel["text"] = "Printing file . . . "
            pbar['value'] += 2
            pbargui.update_idletasks()
            with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Control_Matrix_CSL.txt", 'w', newline='',
                      encoding='utf-8') as myfile:
                wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
                wr.writerows(array_print)
            pbar.destroy()
            pbargui.destroy()
            messagebox.showinfo('Finalizat!', "Finalizat.")
        else:
            pbar.destroy()
            pbargui.destroy()
            messagebox.showerror('Fisier gresit!', "Nu ati incarcat fisierul CSL")
    else:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showerror('Extensie gresita!', "Incarcati fisierul CSV")


def cmtglml():
    pbargui = Tk()
    pbargui.title("Control Matrix CSR")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    listatwist = ["131_002", "131_102", "grau_047"]
    fisier_cm = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                           title="Incarcati fisierul control matrix:")
    array_print = [["Module ID+REAL NAME", "KANBAN-AG", "REAL NAME", "Kanban name", "Module ID", "Ledset",
                    "Type", "Material PN", "Conector 1", "Pin 1", "Conector 2", "Pin 2", "Splice", "Ferc", "X/Y",
                    "Twist", "PB place", "My DWG", "DWG", "APAB_1", "APAB_2", "APAB_3", "APAB_4",
                    "Що змінилось", "(CW)", "ЗМІНА"]]
    idx_conector = []
    idx_pin = []
    idx_module = []
    idx_knname = ""
    idx_realname = ""
    idx_leadset = ""
    idx_kanbanag = ""
    idx_pnmaterial = ""
    idx_splice = ""
    idx_ferc = ""
    idx_xy = ""
    idx_twist = ""
    idx_pbplace = ""
    idx_mydwg = ""
    idx_dwg = ""
    idx_apab1 = ""
    idx_apab2 = ""
    idx_apab3 = ""
    idx_apab4 = ""
    idx_apab5 = ""
    idx_aem1 = ""
    idx_aem2 = 3
    idx_aem3 = 3

    if fisier_cm[-3:] == "csv":
        with open(fisier_cm, newline='') as csvfile:
            array_sortare = list(csv.reader(csvfile, delimiter=','))
        if "8023" in array_sortare[1]:
            for i, j in enumerate(array_sortare[3]):
                if 'Kurzname' in j:
                    idx_conector.append(i)
                elif 'Pin' in j:
                    idx_pin.append(i)
                elif j == 'KANBAN Name':
                    idx_knname = i
                elif j == 'Ltg-Nr.':
                    idx_realname = i
                elif j == 'Leadset':
                    idx_leadset = i
                elif j == 'COD':
                    idx_kanbanag = i
                elif 'Wires' in j:
                    idx_pnmaterial = i
                elif j == 'Splice':
                    idx_splice = i
                elif j == 'Ferc.':
                    idx_ferc = i
                elif j == '23UMTGL':
                    idx_xy = i
                elif j == 'Twist2/4 WPPAK':
                    idx_twist = i
                elif 'PB' in j:
                    idx_pbplace = i
                elif j == 'FOR_MY23':
                    idx_mydwg = i
                elif j == 'Drawing':
                    idx_dwg = i
                elif j == 'APAB_1':
                    idx_apab1 = i
                elif j == 'APAB_2':
                    idx_apab2 = i
                elif j == 'APAB_3':
                    idx_apab3 = i
                elif j == 'APAB_4':
                    idx_apab4 = i
                elif j == 'APAB_5':
                    idx_apab5 = i
                elif j == "AEM":
                    idx_aem1 = i
                elif j == 'AEM':
                    idx_aem2 = i
                elif j == 'AEM':
                    idx_aem3 = i
            for i, j in enumerate(array_sortare[2]):
                if len(j) == 13:
                    idx_module.append(i)
            statuslabel["text"] = "Working on it . . . "
            pbar['value'] += 2
            pbargui.update_idletasks()
            for i in range(4, len(array_sortare)):
                try:
                    for x in range(0, len(idx_module)):
                        if array_sortare[i][idx_module[x]] == "X" or array_sortare[i][idx_module[x]] == "x":
                            array_print.append([array_sortare[2][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname], array_sortare[2][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "FIR",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                        elif array_sortare[i][idx_module[x]] == "Y" or array_sortare[i][idx_module[x]] == "y":
                            array_print.append([array_sortare[2][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname], array_sortare[2][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "OPERATIE",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                        elif (array_sortare[i][idx_module[x]] == "S" or array_sortare[i][idx_module[x]] == "s")\
                                and array_sortare[i][idx_realname] in listatwist:
                            array_print.append([array_sortare[2][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i + 4][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i + 4][idx_knname], array_sortare[2][idx_module[x]],
                                                array_sortare[i + 4][idx_leadset].upper().replace("23U", "23W"), "FIR",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                        elif (array_sortare[i][idx_module[x]] == "S" or array_sortare[i][idx_module[x]] == "s")\
                                and array_sortare[i][idx_realname] not in listatwist:
                            array_print.append([array_sortare[2][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname], array_sortare[2][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "FIRCOMP",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                except:
                    pbar.destroy()
                    pbargui.destroy()
                    messagebox.showerror('Eroare cap de tabel', "Cap de tabel diferit de cel din lista")
            lista_leadset = []
            for i in range(len(array_print)):
                if array_print[i][5] != "":
                    lista_leadset.append([array_print[i][5], array_print[i][4], array_print[i][3],
                                         array_print[i][4] + array_print[i][3]])
            for i in range(1, len(array_print)):
                for x in range(len(lista_leadset)):
                    if array_print[i][4] + array_print[i][3] == lista_leadset[x][3]:
                        array_print[i][1] = lista_leadset[x][0].split("#", 1)[0]
                        statuslabel["text"] = "Replacing leadsets to RO . . . "
                        pbar['value'] += 2
                        pbargui.update_idletasks()
            statuslabel["text"] = "Printing file . . . "
            pbar['value'] += 2
            pbargui.update_idletasks()
            with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Control_Matrix_TGLML.txt", 'w', newline='',
                      encoding='utf-8') as myfile:
                wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
                wr.writerows(array_print)
            pbar.destroy()
            pbargui.destroy()
            messagebox.showinfo('Finalizat!', "Finalizat.")
        else:
            pbar.destroy()
            pbargui.destroy()
            messagebox.showerror('Fisier gresit!', "Nu ati incarcat fisierul CSR")
    else:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showerror('Extensie gresita!', "Incarcati fisierul CSV")


def cmtglmr():
    pbargui = Tk()
    pbargui.title("Control Matrix CSR")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    listatwist = ["131_002", "131_102", "grau_047"]
    fisier_cm = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                           title="Incarcati fisierul control matrix:")
    array_print = [["Module ID+REAL NAME", "KANBAN-AG", "REAL NAME", "Kanban name", "Module ID", "Ledset",
                    "Type", "Material PN", "Conector 1", "Pin 1", "Conector 2", "Pin 2", "Splice", "Ferc", "X/Y",
                    "Twist", "PB place", "My DWG", "DWG", "APAB_1", "APAB_2", "APAB_3", "APAB_4",
                    "Що змінилось", "(CW)", "ЗМІНА"]]
    idx_conector = []
    idx_pin = []
    idx_module = []
    idx_knname = ""
    idx_realname = ""
    idx_leadset = ""
    idx_kanbanag = ""
    idx_pnmaterial = ""
    idx_splice = ""
    idx_ferc = ""
    idx_xy = ""
    idx_twist = ""
    idx_pbplace = ""
    idx_mydwg = ""
    idx_dwg = ""
    idx_apab1 = ""
    idx_apab2 = ""
    idx_apab3 = ""
    idx_apab4 = ""
    idx_apab5 = ""
    idx_aem1 = ""
    idx_aem2 = 3
    idx_aem3 = 3

    if fisier_cm[-3:] == "csv":
        with open(fisier_cm, newline='') as csvfile:
            array_sortare = list(csv.reader(csvfile, delimiter=','))
        if "TGLM" in array_sortare[2]:
            for i, j in enumerate(array_sortare[4]):
                if 'Kurzname' in j:
                    idx_conector.append(i)
                elif 'Pin' in j:
                    idx_pin.append(i)
                elif j == 'KANBAN Name':
                    idx_knname = i
                elif j == 'Ltg-Nr.':
                    idx_realname = i
                elif j == 'Leadset':
                    idx_leadset = i
                elif j == 'COD':
                    idx_kanbanag = i
                elif 'Wires' in j:
                    idx_pnmaterial = i
                elif j == 'Splice':
                    idx_splice = i
                elif j == 'Ferc.':
                    idx_ferc = i
                elif j == '23UMTGR':
                    idx_xy = i
                elif j == 'Twist2/4 WPPAK':
                    idx_twist = i
                elif 'PB' in j:
                    idx_pbplace = i
                elif j == 'A4/A5':
                    idx_mydwg = i
                elif j == 'Drawing':
                    idx_dwg = i
                elif j == 'APAB_1':
                    idx_apab1 = i
                elif j == 'APAB_2':
                    idx_apab2 = i
                elif j == 'APAB_3':
                    idx_apab3 = i
                elif j == 'APAB_4':
                    idx_apab4 = i
                elif j == 'APAB_5':
                    idx_apab5 = i
                elif j == "AEM":
                    idx_aem1 = i
                elif j == 'AEM':
                    idx_aem2 = i
                elif j == 'AEM':
                    idx_aem3 = i
            for i, j in enumerate(array_sortare[3]):
                if len(j) == 13:
                    idx_module.append(i)
            statuslabel["text"] = "Working on it . . . "
            pbar['value'] += 2
            pbargui.update_idletasks()
            for i in range(4, len(array_sortare)):
                try:
                    for x in range(0, len(idx_module)):
                        if array_sortare[i][idx_module[x]] == "X" or array_sortare[i][idx_module[x]] == "x":
                            array_print.append([array_sortare[3][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname], array_sortare[3][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "FIR",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                        elif array_sortare[i][idx_module[x]] == "Y" or array_sortare[i][idx_module[x]] == "y":
                            array_print.append([array_sortare[3][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname], array_sortare[3][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "OPERATIE",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                        elif (array_sortare[i][idx_module[x]] == "S" or array_sortare[i][idx_module[x]] == "s")\
                                and array_sortare[i][idx_realname] in listatwist:
                            array_print.append([array_sortare[3][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i + 4][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i + 4][idx_knname], array_sortare[3][idx_module[x]],
                                                array_sortare[i + 4][idx_leadset].upper().replace("23U", "23W"), "FIR",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                        elif (array_sortare[i][idx_module[x]] == "S" or array_sortare[i][idx_module[x]] == "s")\
                                and array_sortare[i][idx_realname] not in listatwist:
                            array_print.append([array_sortare[3][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname], array_sortare[3][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "FIRCOMP",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                except:
                    pbar.destroy()
                    pbargui.destroy()
                    messagebox.showerror('Eroare cap de tabel', "Cap de tabel diferit de cel din lista")
            lista_leadset = []
            for i in range(len(array_print)):
                if array_print[i][5] != "":
                    lista_leadset.append([array_print[i][5], array_print[i][4], array_print[i][3],
                                         array_print[i][4] + array_print[i][3]])
            for i in range(1, len(array_print)):
                for x in range(len(lista_leadset)):
                    if array_print[i][4] + array_print[i][3] == lista_leadset[x][3]:
                        array_print[i][1] = lista_leadset[x][0].split("#", 1)[0]
                        statuslabel["text"] = "Replacing leadsets to RO . . . "
                        pbar['value'] += 2
                        pbargui.update_idletasks()
            statuslabel["text"] = "Printing file . . . "
            pbar['value'] += 2
            pbargui.update_idletasks()
            with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Control_Matrix_TGLMR.txt", 'w', newline='',
                      encoding='utf-8') as myfile:
                wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
                wr.writerows(array_print)
            pbar.destroy()
            pbargui.destroy()
            messagebox.showinfo('Finalizat!', "Finalizat.")
        else:
            pbar.destroy()
            pbargui.destroy()
            messagebox.showerror('Fisier gresit!', "Nu ati incarcat fisierul CSR")
    else:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showerror('Extensie gresita!', "Incarcati fisierul CSV")


def cm4axell():
    pbargui = Tk()
    pbargui.title("Control Matrix CSR")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    listatwist = ["131_002", "131_102", "grau_047"]
    fisier_cm = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                           title="Incarcati fisierul control matrix:")
    array_print = [["Module ID+REAL NAME", "KANBAN-AG", "REAL NAME", "Kanban name", "Module ID", "Ledset",
                    "Type", "Material PN", "Conector 1", "Pin 1", "Conector 2", "Pin 2", "Splice", "Ferc", "X/Y",
                    "Twist", "PB place", "My DWG", "DWG", "APAB_1", "APAB_2", "APAB_3", "APAB_4",
                    "Що змінилось", "(CW)", "ЗМІНА"]]
    idx_conector = []
    idx_pin = []
    idx_module = []
    idx_knname = ""
    idx_realname = ""
    idx_leadset = ""
    idx_kanbanag = ""
    idx_pnmaterial = ""
    idx_splice = ""
    idx_ferc = ""
    idx_xy = ""
    idx_twist = ""
    idx_pbplace = ""
    idx_mydwg = ""
    idx_dwg = 0
    idx_apab1 = ""
    idx_apab2 = ""
    idx_apab3 = ""
    idx_apab4 = ""
    idx_apab5 = ""
    idx_aem1 = ""
    idx_aem2 = 3
    idx_aem3 = 3

    if fisier_cm[-3:] == "csv":
        with open(fisier_cm, newline='') as csvfile:
            array_sortare = list(csv.reader(csvfile, delimiter=','))
        if "AXL" in array_sortare[3]:
            for i, j in enumerate(array_sortare[2]):
                if 'Xcode' in j:
                    idx_conector.append(i)
                elif 'Cavity' in j:
                    idx_pin.append(i)
                elif 'KANBAN ' in j:
                    idx_knname = i
                elif j == 'WireNo':
                    idx_realname = i
                elif j == 'Leadset':
                    idx_leadset = i
                elif j == 'COD':
                    idx_kanbanag = i
                elif j == 'PN':
                    idx_pnmaterial = i
                elif j == 'Splice':
                    idx_splice = i
                elif j == 'Ferc':
                    idx_ferc = i
                elif j == 'X/Y/S':
                    idx_xy = i
                elif j == 'MultiCore':
                    idx_twist = i
                elif 'PB Place' in j:
                    idx_pbplace = i
                elif 'Format' in j:
                    idx_mydwg = i
                elif j == 'Drawing':
                    idx_dwg = i
                elif j == 'APAB_1':
                    idx_apab1 = i
                elif j == 'APAB_2':
                    idx_apab2 = i
                elif j == 'APAB_3':
                    idx_apab3 = i
                elif j == 'APAB_4':
                    idx_apab4 = i
                elif j == 'APAB_5':
                    idx_apab5 = i
                elif '_AEM' in j:
                    idx_aem1 = i
                elif j == 'CW AEM':
                    idx_aem2 = i
                elif j == 'AEM':
                    idx_aem3 = i
            for i, j in enumerate(array_sortare[1]):
                if len(j) == 13:
                    idx_module.append(i)
            statuslabel["text"] = "Working on it . . . "
            pbar['value'] += 2
            pbargui.update_idletasks()
            for i in range(4, len(array_sortare)):
                try:
                    for x in range(0, len(idx_module)):
                        if array_sortare[i][idx_module[x]] == "X" or array_sortare[i][idx_module[x]] == "x":
                            array_print.append([array_sortare[1][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname], array_sortare[1][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "FIR",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                        elif array_sortare[i][idx_module[x]] == "Y" or array_sortare[i][idx_module[x]] == "y":
                            array_print.append([array_sortare[1][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname], array_sortare[1][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "OPERATIE",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                        elif (array_sortare[i][idx_module[x]] == "S" or array_sortare[i][idx_module[x]] == "s")\
                                and array_sortare[i][idx_realname] in listatwist:
                            array_print.append([array_sortare[1][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i + 4][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i + 4][idx_knname], array_sortare[1][idx_module[x]],
                                                array_sortare[i + 4][idx_leadset].upper().replace("23U", "23W"), "FIR",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                        elif (array_sortare[i][idx_module[x]] == "S" or array_sortare[i][idx_module[x]] == "s")\
                                and array_sortare[i][idx_realname] not in listatwist:
                            array_print.append([array_sortare[1][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname], array_sortare[1][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "FIRCOMP",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                except:
                    pbar.destroy()
                    pbargui.destroy()
                    messagebox.showerror('Eroare cap de tabel', "Cap de tabel diferit de cel din lista")
            lista_leadset = []
            for i in range(len(array_print)):
                if array_print[i][5] != "":
                    lista_leadset.append([array_print[i][5], array_print[i][4], array_print[i][3],
                                         array_print[i][4] + array_print[i][3]])
            for i in range(1, len(array_print)):
                for x in range(len(lista_leadset)):
                    if array_print[i][4] + array_print[i][3] == lista_leadset[x][3]:
                        array_print[i][1] = lista_leadset[x][0].split("#", 1)[0]
                        statuslabel["text"] = "Replacing leadsets to RO . . . "
                        pbar['value'] += 2
                        pbargui.update_idletasks()
            statuslabel["text"] = "Printing file . . . "
            pbar['value'] += 2
            pbargui.update_idletasks()
            with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Control_Matrix_4AXELL.txt", 'w', newline='',
                      encoding='utf-8') as myfile:
                wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
                wr.writerows(array_print)
            pbar.destroy()
            pbargui.destroy()
            messagebox.showinfo('Finalizat!', "Finalizat.")
        else:
            pbar.destroy()
            pbargui.destroy()
            messagebox.showerror('Fisier gresit!', "Nu ati incarcat fisierul CSR")
    else:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showerror('Extensie gresita!', "Incarcati fisierul CSV")


def cm4axelr():
    pbargui = Tk()
    pbargui.title("Control Matrix CSR")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    listatwist = ["131_002", "131_102", "grau_047"]
    fisier_cm = filedialog.askopenfilename(initialdir=os.path.abspath(os.curdir),
                                           title="Incarcati fisierul control matrix:")
    array_print = [["Module ID+REAL NAME", "KANBAN-AG", "REAL NAME", "Kanban name", "Module ID", "Ledset",
                    "Type", "Material PN", "Conector 1", "Pin 1", "Conector 2", "Pin 2", "Splice", "Ferc", "X/Y",
                    "Twist", "PB place", "My DWG", "DWG", "APAB_1", "APAB_2", "APAB_3", "APAB_4",
                    "Що змінилось", "(CW)", "ЗМІНА"]]
    idx_conector = []
    idx_pin = []
    idx_module = []
    idx_knname = ""
    idx_realname = ""
    idx_leadset = ""
    idx_kanbanag = ""
    idx_pnmaterial = ""
    idx_splice = ""
    idx_ferc = ""
    idx_xy = ""
    idx_twist = ""
    idx_pbplace = ""
    idx_mydwg = ""
    idx_dwg = 0
    idx_apab1 = ""
    idx_apab2 = ""
    idx_apab3 = ""
    idx_apab4 = ""
    idx_apab5 = ""
    idx_aem1 = 3
    idx_aem2 = 3
    idx_aem3 = 3
    if fisier_cm[-3:] == "csv":
        with open(fisier_cm, newline='') as csvfile:
            array_sortare = list(csv.reader(csvfile, delimiter=','))
        if "23UAXEXTRA" in array_sortare[0]:
            for i, j in enumerate(array_sortare[2]):
                if 'Xcode' in j:
                    idx_conector.append(i)
                elif 'Cavity' in j:
                    idx_pin.append(i)
                elif 'KANBAN ' in j:
                    idx_knname = i
                elif j == 'WireNo':
                    idx_realname = i
                elif j == 'LeadSet':
                    idx_leadset = i
                elif j == 'COD':
                    idx_kanbanag = i
                elif j == 'PN':
                    idx_pnmaterial = i
                elif j == 'Splice':
                    idx_splice = i
                elif j == 'Ferc':
                    idx_ferc = i
                elif j == 'X/Y/S':
                    idx_xy = i
                elif j == 'MultiCore':
                    idx_twist = i
                elif 'PB Place' in j:
                    idx_pbplace = i
                elif 'Format' in j:
                    idx_mydwg = i
                elif j == 'Drawing':
                    idx_dwg = i
                elif j == 'APAB_1':
                    idx_apab1 = i
                elif j == 'APAB_2':
                    idx_apab2 = i
                elif j == 'APAB_3':
                    idx_apab3 = i
                elif j == 'APAB_4':
                    idx_apab4 = i
                elif j == 'APAB_5':
                    idx_apab5 = i
                elif '№АЕМ' in j:
                    idx_aem1 = i
                elif j == 'CW AEM':
                    idx_aem2 = i
                elif j == 'AEM':
                    idx_aem3 = i
            for i, j in enumerate(array_sortare[1]):
                if len(j) == 13:
                    idx_module.append(i)
            statuslabel["text"] = "Working on it . . . "
            pbar['value'] += 2
            pbargui.update_idletasks()
            for i in range(4, len(array_sortare)):
                try:
                    for x in range(0, len(idx_module)):
                        if array_sortare[i][idx_module[x]] == "X" or array_sortare[i][idx_module[x]] == "x":
                            array_print.append([array_sortare[1][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname], array_sortare[1][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "FIR",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                        elif array_sortare[i][idx_module[x]] == "Y" or array_sortare[i][idx_module[x]] == "y":
                            array_print.append([array_sortare[1][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname], array_sortare[1][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "OPERATIE",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                        elif (array_sortare[i][idx_module[x]] == "S" or array_sortare[i][idx_module[x]] == "s")\
                                and array_sortare[i][idx_realname] in listatwist:
                            array_print.append([array_sortare[1][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i + 4][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i + 4][idx_knname], array_sortare[1][idx_module[x]],
                                                array_sortare[i + 4][idx_leadset].upper().replace("23U", "23W"), "FIR",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                        elif (array_sortare[i][idx_module[x]] == "S" or array_sortare[i][idx_module[x]] == "s")\
                                and array_sortare[i][idx_realname] not in listatwist:
                            array_print.append([array_sortare[1][idx_module[x]] + array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_kanbanag].split("-")[0].upper().replace("23U", "23W"),
                                                array_sortare[i][idx_realname].lower(),
                                                array_sortare[i][idx_knname], array_sortare[1][idx_module[x]],
                                                array_sortare[i][idx_leadset].upper().replace("23U", "23W"), "FIRCOMP",
                                                array_sortare[i][idx_pnmaterial], array_sortare[i][idx_conector[0]],
                                                array_sortare[i][idx_pin[0]], array_sortare[i][idx_conector[1]],
                                                array_sortare[i][idx_pin[1]], array_sortare[i][idx_splice],
                                                array_sortare[i][idx_ferc], array_sortare[i][idx_xy],
                                                array_sortare[i][idx_twist],
                                                array_sortare[i][idx_pbplace], array_sortare[i][idx_mydwg],
                                                array_sortare[i][idx_dwg], array_sortare[i][idx_apab1],
                                                array_sortare[i][idx_apab2], array_sortare[i][idx_apab3],
                                                array_sortare[i][idx_apab4], array_sortare[i][idx_apab5],
                                                array_sortare[i][idx_aem1], array_sortare[i][idx_aem2],
                                                array_sortare[i][idx_aem3]])
                            pbar['value'] += 2
                            pbargui.update_idletasks()
                except:
                    pbar.destroy()
                    pbargui.destroy()
                    messagebox.showerror('Eroare cap de tabel', "Cap de tabel diferit de cel din lista")
            lista_leadset = []
            for i in range(len(array_print)):
                if array_print[i][5] != "":
                    lista_leadset.append([array_print[i][5], array_print[i][4], array_print[i][3],
                                         array_print[i][4] + array_print[i][3]])
            for i in range(1, len(array_print)):
                for x in range(len(lista_leadset)):
                    if array_print[i][4] + array_print[i][3] == lista_leadset[x][3]:
                        array_print[i][1] = lista_leadset[x][0].split("#", 1)[0]
                        statuslabel["text"] = "Replacing leadsets to RO . . . "
                        pbar['value'] += 2
                        pbargui.update_idletasks()
            statuslabel["text"] = "Printing file . . . "
            pbar['value'] += 2
            pbargui.update_idletasks()
            with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Control_Matrix_4AXELR.txt", 'w', newline='',
                      encoding='utf-8') as myfile:
                wr = csv.writer(myfile, quoting=csv.QUOTE_ALL, delimiter=';')
                wr.writerows(array_print)
            pbar.destroy()
            pbargui.destroy()
            messagebox.showinfo('Finalizat!', "Finalizat.")
        else:
            pbar.destroy()
            pbargui.destroy()
            messagebox.showerror('Fisier gresit!', "Nu ati incarcat fisierul CSR")
    else:
        pbar.destroy()
        pbargui.destroy()
        messagebox.showerror('Extensie gresita!', "Incarcati fisierul CSV")
