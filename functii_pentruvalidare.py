import csv
import os
import time
from collections import Counter
from tkinter import Tk, ttk, HORIZONTAL, Label, messagebox

from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.workbook import Workbook

import globale
from diverse import pivotare, skip_file, istsoll, log_file


def wirelist_validare():
        pbargui = Tk()
        pbargui.title("Wirelist toate")
        pbargui.geometry("500x50+50+550")
        pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
        statuslabel = Label(pbargui, text="Waiting . . .")
        timelabel = Label(pbargui, text="Time . . .")
        pbar.grid(row=1, column=1, padx=5, pady=5)
        statuslabel.grid(row=1, column=2, padx=5, pady=5)
        timelabel.grid(row=2, column=2)
        dir_selectat = os.path.abspath(os.curdir) + "/MAN/Input/Module Files/8000/"
        dir_output = os.path.abspath(os.curdir) + "/MAN/Output/Excel Files/8000/"
        with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/KSKLight.txt", newline='') as csvfile:
            array_sortare_light = list(csv.reader(csvfile, delimiter=';'))
        counter = 0
        start = time.time()
        file_counter = 0
        file_progres = 0
        for file_all in os.listdir(dir_selectat):
            if file_all.endswith(".csv"):
                file_counter = file_counter + 1
        pbar['value'] = 0
        start0 = time.time()
        for file_all in os.listdir(dir_selectat):
            if file_all.endswith(".csv"):
                globale.is_light_save = "0"
                file_progres = file_progres + 1
                statuslabel["text"] = "8000 = " + str(file_progres) + "/" + str(file_counter) + " : " + file_all
                pbar['value'] += 1
                pbargui.update_idletasks()
                with open(dir_selectat + file_all, newline='') as csvfile:
                    array_modul = list(csv.reader(csvfile, delimiter=';'))
                if array_modul[0][0] != "Harness" and array_modul[0][0] != "Module":
                    messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect. Eroare cap de tabel!')
                    return
                counter = counter + 1
                array_temporar_module = [array_modul[i][1] for i in range(1, len(array_modul))]
                if set(array_temporar_module).issubset(array_sortare_light[0]):
                    globale.is_light_save = "1"
                prelucrare_wirelist_faza1(array_modul)
                end0 = time.time()
                timelabel["text"] = "Estimated time to complete : " + \
                                    str(((file_counter * 2) - (end0 - start0)) / 60)[:5] + " minutes."
                pbargui.update_idletasks()
                continue
            else:
                continue
        dir_selectat = os.path.abspath(os.curdir) + "/MAN/Input/Module Files/8011/"
        file_counter = 0
        file_progres = 0
        for file_all in os.listdir(dir_selectat):
            if file_all.endswith(".csv"):
                file_counter = file_counter + 1
        pbar['value'] = 0
        start1 = time.time()
        for file_all in os.listdir(dir_selectat):
            if file_all.endswith(".csv"):
                globale.is_light_save = "0"
                file_progres = file_progres + 1
                statuslabel["text"] = "8011 = " + str(file_progres) + "/" + str(file_counter) + " : " + file_all
                pbar['value'] += 1
                pbargui.update_idletasks()
                with open(dir_selectat + file_all, newline='') as csvfile:
                    array_modul = list(csv.reader(csvfile, delimiter=';'))
                if array_modul[0][0] != "Harness" and array_modul[0][0] != "Module":
                    messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect.  Eroare cap de tabel11!')
                    return
                counter = counter + 1
                array_temporar_module = [array_modul[i][1] for i in range(1, len(array_modul))]
                if set(array_temporar_module).issubset(array_sortare_light[0]):
                    globale.is_light_save = "1"
                prelucrare_wirelist_faza1(array_modul)
                end1 = time.time()
                timelabel["text"] = "Estimated time to complete : " + \
                                    str(((file_counter * 2) - (end1 - start1)) / 60)[:5] + " minutes."
                pbargui.update_idletasks()
                continue
            else:
                continue
        dir_selectat = os.path.abspath(os.curdir) + "/MAN/Input/Module Files/8023/"
        file_counter = 0
        file_progres = 0
        for file_all in os.listdir(dir_selectat):
            if file_all.endswith(".csv"):
                file_counter = file_counter + 1
        pbar['value'] = 0
        start2 = time.time()
        for file_all in os.listdir(dir_selectat):
            if file_all.endswith(".csv"):
                globale.is_light_save = "0"
                file_progres = file_progres + 1
                statuslabel["text"] = "8023 = " + str(file_progres) + "/" + str(file_counter) + " : " + file_all
                pbar['value'] += 1
                pbargui.update_idletasks()
                with open(dir_selectat + file_all, newline='') as csvfile:
                    array_modul = list(csv.reader(csvfile, delimiter=';'))
                if array_modul[0][0] != "Harness" and array_modul[0][0] != "Module":
                    messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect.  Eroare cap de tabel!')
                    return
                counter = counter + 1
                array_temporar_module = [array_modul[i][1] for i in range(1, len(array_modul))]
                if set(array_temporar_module).issubset(array_sortare_light[0]):
                    globale.is_light_save = "1"
                prelucrare_wirelist_faza1(array_modul)
                end2 = time.time()
                timelabel["text"] = "Estimated time to complete : " + \
                                    str(((file_counter * 2) - (end2 - start2)) / 60)[:5] + " minutes."
                pbargui.update_idletasks()
                continue
            else:
                continue
        dir_selectat = os.path.abspath(os.curdir) + "/MAN/Input/Module Files/Necunoscut/"
        dir_output = os.path.abspath(os.curdir) + "/MAN/Output/Excel Files/Necunoscut/"
        file_counter = 0
        file_progres = 0
        for file_all in os.listdir(dir_selectat):
            if file_all.endswith(".csv"):
                file_counter = file_counter + 1
        pbar['value'] = 0
        start3 = time.time()
        for file_all in os.listdir(dir_selectat):
            if file_all.endswith(".csv"):
                globale.is_light_save = "0"
                file_progres = file_progres + 1
                statuslabel["text"] = "Necunoscut = " + str(file_progres) + "/" + str(file_counter) + " : " + file_all
                pbar['value'] += 1
                pbargui.update_idletasks()
                with open(dir_selectat + file_all, newline='') as csvfile:
                    array_modul = list(csv.reader(csvfile, delimiter=';'))
                if array_modul[0][0] != "Harness" and array_modul[0][0] != "Module":
                    messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect.  Eroare cap de tabel!')
                    return
                counter = counter + 1
                array_temporar_module = [array_modul[i][1] for i in range(1, len(array_modul))]
                if set(array_temporar_module).issubset(array_sortare_light[0]):
                    globale.is_light_save = "1"
                prelucrare_wirelist_faza1(array_modul)
                end3 = time.time()
                timelabel["text"] = "Estimated time to complete : " + \
                                    str(((file_counter * 2) - (end3 - start3)) / 60)[:5] + " minutes."
                pbargui.update_idletasks()
                continue
            else:
                continue
        end = time.time()
        pbar.destroy()
        pbargui.destroy()
        messagebox.showinfo('Finalizat!',
                            'Prelucrate ' + str(counter) + " fisiere in " + str(end - start)[:6] + " secunde.")


def bom_validare():
    pbargui = Tk()
    pbargui.title("BOM toate")
    pbargui.geometry("500x50+50+550")
    pbar = ttk.Progressbar(pbargui, orient=HORIZONTAL, length=200, mode='indeterminate')
    statuslabel = Label(pbargui, text="Waiting . . .")
    pbar.grid(row=1, column=1, padx=5, pady=5)
    statuslabel.grid(row=1, column=2, padx=5, pady=5)
    dir_selectat = os.path.abspath(os.curdir) + "/MAN/Input/Module Files/8000/"
    counter = 0
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".csv"):
            file_counter = file_counter + 1
    pbar['value'] = 0
    start = time.time()
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".csv"):
            with open(dir_selectat + file_all, newline='') as csvfile:
                array_modul = list(csv.reader(csvfile, delimiter=';'))
            if array_modul[0][0] != "Harness" and array_modul[0][0] != "Module":
                messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect')
                return
            file_progres = file_progres + 1
            statuslabel["text"] = "8000 = " + str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 1
            pbargui.update_idletasks()
            counter = counter + 1
            prelucrare_bom_faza1(array_modul)
            continue
        else:
            continue

    dir_selectat = os.path.abspath(os.curdir) + "/MAN/Input/Module Files/8011/"
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".csv"):
            file_counter = file_counter + 1
    pbar['value'] = 0
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".csv"):
            with open(dir_selectat + file_all, newline='') as csvfile:
                array_modul = list(csv.reader(csvfile, delimiter=';'))
            if array_modul[0][0] != "Harness" and array_modul[0][0] != "Module":
                messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect.  Eroare cap de tabel!')
                return
            counter = counter + 1
            file_progres = file_progres + 1
            statuslabel["text"] = "8011 = " + str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 1
            pbargui.update_idletasks()
            prelucrare_bom_faza1(array_modul)
            continue
        else:
            continue
    dir_selectat = os.path.abspath(os.curdir) + "/MAN/Input/Module Files/8023/"
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".csv"):
            file_counter = file_counter + 1
    pbar['value'] = 0
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".csv"):
            with open(dir_selectat + file_all, newline='') as csvfile:
                array_modul = list(csv.reader(csvfile, delimiter=';'))
            if array_modul[0][0] != "Harness" and array_modul[0][0] != "Module":
                messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect.  Eroare cap de tabel!')
                return
            counter = counter + 1
            file_progres = file_progres + 1
            statuslabel["text"] = "8023 = " + str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 1
            pbargui.update_idletasks()
            prelucrare_bom_faza1(array_modul)
            continue
        else:
            continue
    dir_selectat = os.path.abspath(os.curdir) + "/MAN/Input/Module Files/Necunoscut/"
    file_counter = 0
    file_progres = 0
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".csv"):
            file_counter = file_counter + 1
    pbar['value'] = 0
    for file_all in os.listdir(dir_selectat):
        if file_all.endswith(".csv"):
            with open(dir_selectat + file_all, newline='') as csvfile:
                array_modul = list(csv.reader(csvfile, delimiter=';'))
            if array_modul[0][0] != "Harness" and array_modul[0][0] != "Module":
                messagebox.showerror('Eroare fisier', 'Nu ai incarcat fisierul corect')
                return
            counter = counter + 1
            file_progres = file_progres + 1
            statuslabel["text"] = "Necunoscut = " + str(file_progres) + "/" + str(file_counter) + " : " + file_all
            pbar['value'] += 1
            pbargui.update_idletasks()
            prelucrare_bom_faza1(array_modul)
            continue
        else:
            continue
    pbar.destroy()
    pbargui.destroy()
    end = time.time()
    messagebox.showinfo('Finalizat!',
                        'Prelucrate ' + str(counter) + " fisiere in " + str(end - start)[:6] + " secunde.")




def prelucrare_wirelist_faza1(array_prelucrare):
    # arr_module_file = [["Harness", "Module", "Module", "Drawing Number", "Module Check", "Comments", "Implemented",
    #                    "Desen", "Cantitate", "Date"]]
    lista_module_extrageret = []
    module_implementate = []
    modulemy23 = []
    arr_moduleksk = []
    arr_module_file = [[array_prelucrare[i][0], array_prelucrare[i][1]] for i in range(1, len(array_prelucrare))]
    arr_module_file.insert(0, ["Harness", "Module", "Module", "Drawing Number", "Module Check", "Comments",
                               "Implemented", "Desen", "Cantitate", "Date", "Trailer No"])
    # for i in range(1, len(array_prelucrare)):
    #    arr_module_file.append([array_prelucrare[i][0], array_prelucrare[i][1]])
    "Load required data files"
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Module Active.txt", newline='') as csvfile:
        arr_module_active = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Module Implementate.txt", newline='') as csvfile:
        arr_module_implementate = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Module MY2023.txt", newline='') as csvfile:
        modulemy23 = list(csv.reader(csvfile, delimiter=';'))
    "Prelucrare efectiva"
    for x in range(len(arr_module_active)):
        for y in range(1, len(arr_module_file)):
            if arr_module_file[y][1] == arr_module_active[x][0]:
                arr_module_file[y].extend(arr_module_active[x])
    for i in range(1, len(arr_module_file)):
        if len(arr_module_file[i]) == 2:
            arr_module_file[i].extend(["XXXX", "XXXX", "XXXX", "XXXX"])
    for x in range(len(arr_module_implementate)):
        module_implementate.append(arr_module_implementate[x][0])
    for y in range(1, len(arr_module_file)):
        if arr_module_file[y][1] in module_implementate:
            arr_module_file[y].append("OK")
        elif arr_module_file[y][5] == "XXXX":
            arr_module_file[y].append("XXXX")
        else:
            arr_module_file[y].append("Not OK")

    for i in range(1, len(arr_module_file)):
        if arr_module_file[i][5] != "XXXX":
            lista_module_extrageret.append(arr_module_file[i][5])
    lista_module_extrageret = Counter(lista_module_extrageret).most_common(2)
    lista_module_extragere = [lista_module_extrageret[i][0] for i in range(0, 2)]
    for i in range(1, len(array_prelucrare)):
        if len(array_prelucrare[i][3]) > 10:
            arr_module_file[i].extend([array_prelucrare[i][2], "0"])
        elif array_prelucrare[i][3] == "1000":
            arr_module_file[i].extend([array_prelucrare[i][2], "1"])
        elif array_prelucrare[i][3] == "2000":
            arr_module_file[i].extend([array_prelucrare[i][2], "2"])
        elif array_prelucrare[i][3] == "3000":
            arr_module_file[i].extend([array_prelucrare[i][2], "3"])
        else:
            arr_module_file[i].extend([array_prelucrare[i][2], array_prelucrare[i][3]])
    for i in range(len(array_prelucrare)):
        if len(array_prelucrare[i]) == 4:
            array_prelucrare[i].append("PC")
    for i in range(1, len(array_prelucrare)):
        if len(array_prelucrare[i]) == 5:
            arr_module_file[i].append("No date")
        else:
            arr_module_file[i].append(array_prelucrare[i][5])
    for i in range(1, len(arr_module_file)):
        try:
            for x in range(1, len(array_prelucrare)):
                if array_prelucrare[x][0] == arr_module_file[i][0]:
                    arr_module_file[i].append(array_prelucrare[x][7])
                    break
        except IndexError:
            arr_module_file[i].append("No info")
    """conditie my23"""
    my23list = []
    for i in range(len(modulemy23)):
        my23list.append(modulemy23[i][0])
    my23 = 0
    for i in range(len(arr_module_file)):
        if arr_module_file[i][1] in my23list:
            my23 = my23 + 1
    if my23 > 0:
        "ceva de my23"
    prelucrare_wirelist_faza2(arr_module_file, lista_module_extragere)


def prelucrare_wirelist_faza2(arr_module_file2, listas):
    # "Selectie fisiere wirelist"
    lista_selectie = (["SATTEL LHD", "8011"], ["SATTEL RHD", "8013"], ["CHASSIS LHD", "8012"], ["CHASSIS RHD", "8014"],
                      ["TGLM LHD", "8023"], ["TGLM RHD", "8024"], ["4AXEL LHD", "8025"], ["4AXEL RHD", "8026"],
                      ["4AXEL MIL LHD", "8000"], ["4AXEL MIL RHD", "8001"], ["CHASSIS MIL RHD", "8030"],
                      ["CHASSIS MIL LHD", "8031"], ["MIL_SAT RHD", "8052"], ["MIL_SAT LHD", "8053"])
    lista_fisiere = []
    array_scriere_sheet2 = [["Harness", "Module", "Ltg-Nr.", "Leitung", "Farbe", "Quer.", "Kurzname", "Pin",
                             "Kurzname", "Pin", "Lange",
                             "Kurzname/Pin", "K/P Count", "One Wire Error", "DC Error", "Cross Sec Error",
                             "Combination Error", "Sonderltg."]]
    array_scriere_sheet3 = [["Harness", "Module", "Ltg-Nr.", "Leitung", "Farbe", "Quer.", "Kurzname", "Pin",
                             "Kurzname", "Pin", "Lange",
                             "Kurzname/Pin", "K/P Count", "One Wire Error", "DC Error", "Cross Sec Error",
                             "Combination Error", "Sonderltg."]]
    concatenare_2 = []
    concatenare_3 = []
    for i in range(len(listas)):
        for x in range(len(lista_selectie)):
            if listas[i] in lista_selectie[x]:
                lista_fisiere.append(lista_selectie[x][1])
    lista_fisiere = sorted(lista_fisiere)
    "Load required data files"
    if len(lista_fisiere) == 0:
        messagebox.showerror('Eroare fisier sursa', 'Verificati fisierul sursa Module Active')
        quit()
    if len(lista_fisiere) != 2:
        lista_fisiere.append("9999")
        skip_file(arr_module_file2[1][0] + " Am gasit fisiere cu un singur wirelist!")
    try:
        with open(os.path.abspath(os.curdir) + "/MAN/Input/Wire Lists/" + lista_fisiere[0] + ".Wirelist.csv",
                  newline='') as csvfile:
            array_wires_1 = list(csv.reader(csvfile, delimiter=';'))
    except FileNotFoundError:
        messagebox.showerror('Eroare fisier', 'Lipsa fisierul ' + lista_fisiere[0])
        quit()
    try:
        with open(os.path.abspath(os.curdir) + "/MAN/Input/Wire Lists/" + lista_fisiere[1] + ".Wirelist.csv",
                  newline='') as csvfile:
            array_wires_2 = list(csv.reader(csvfile, delimiter=';'))
    except FileNotFoundError:
        messagebox.showerror('Eroare fisier', 'Lipsa fisierul ' + lista_fisiere[1])
        quit()
    print (array_wires_1[0:25])
    "Prelucrare efectiva"
    for x in range(1, len(arr_module_file2)):
        for y in range(1, len(array_wires_1)):
            if arr_module_file2[x][1] in array_wires_1[y]:
                array_scriere_sheet2.append([lista_fisiere[0], arr_module_file2[x][1], array_wires_1[y][1],
                                             array_wires_1[y][2], array_wires_1[y][3], array_wires_1[y][4],
                                             array_wires_1[y][5], array_wires_1[y][6],
                                             array_wires_1[y][7], array_wires_1[y][8], array_wires_1[y][10]])

    for x in range(1, len(arr_module_file2)):
        for y in range(1, len(array_wires_1)):
            if arr_module_file2[x][1] in array_wires_1[y]:
                array_scriere_sheet2.append([lista_fisiere[0], arr_module_file2[x][1], array_wires_1[y][1],
                                             array_wires_1[y][2], array_wires_1[y][3], array_wires_1[y][4],
                                             array_wires_1[y][5], array_wires_1[y][6],
                                             array_wires_1[y][7], array_wires_1[y][8], array_wires_1[y][10]])
    for x in range(1, len(arr_module_file2)):
        for y in range(1, len(array_wires_2)):
            if arr_module_file2[x][1] in array_wires_2[y]:
                array_scriere_sheet3.append([lista_fisiere[1], arr_module_file2[x][1], array_wires_2[y][1],
                                             array_wires_2[y][2], array_wires_2[y][3], array_wires_2[y][4],
                                             array_wires_2[y][5], array_wires_2[y][6],
                                             array_wires_2[y][7], array_wires_2[y][8], array_wires_2[y][10]])
    for x in range(1, len(arr_module_file2)):
        for y in range(1, len(array_wires_2)):
            if arr_module_file2[x][1] in array_wires_2[y]:
                array_scriere_sheet3.append([lista_fisiere[1], arr_module_file2[x][1], array_wires_2[y][1],
                                             array_wires_2[y][2], array_wires_2[y][3], array_wires_2[y][4],
                                             array_wires_2[y][7], array_wires_2[y][8], array_wires_2[y][7],
                                             array_wires_2[y][8], array_wires_2[y][10]])
    """creare sheet fantoma pentru cablaje cu un singur side"""
    if lista_fisiere[1] == "9999":
        array_scriere_sheet3.append([lista_fisiere[1], "1", "2", "3", "4", "0.5", "6", "7", "8"])
        array_scriere_sheet3.append([lista_fisiere[1], "11", "22", "33", "44", "0.5", "66", "77", "88"])
        array_scriere_sheet3.append([lista_fisiere[1], "111", "222", "333", "444", "0.5", "666", "777", "888"])
        array_scriere_sheet3.append([lista_fisiere[1], "1111", "2222", "3333", "4444", "0.5", "6666", "7777", "8888"])
    "Concatenare si verificare spl"
    for x in range(1, len(array_scriere_sheet2)):
        array_scriere_sheet2[x].extend([array_scriere_sheet2[x][6] + "/" + array_scriere_sheet2[x][7]])
        concatenare_2.append(array_scriere_sheet2[x][6] + "/" + array_scriere_sheet2[x][7])
    for x in range(1, len(array_scriere_sheet2)):
        cc_count2 = (concatenare_2.count(array_scriere_sheet2[x][9]))
        array_scriere_sheet2[x].append(cc_count2)
    for x in range(1, len(array_scriere_sheet3)):
        array_scriere_sheet3[x].extend([array_scriere_sheet3[x][6] + "/" + array_scriere_sheet3[x][7]])
        concatenare_3.append(array_scriere_sheet3[x][6] + "/" + array_scriere_sheet3[x][7])
    for x in range(1, len(array_scriere_sheet3)):
        cc_count3 = (concatenare_3.count(array_scriere_sheet3[x][9]))
        array_scriere_sheet3[x].append(cc_count3)
    "erorare single"

    for x in range(1, len(array_scriere_sheet2)):
        if int(array_scriere_sheet2[x][12]) == 1 and (array_scriere_sheet2[x][9][:2] == "X9" or
                                                      array_scriere_sheet2[x][9][:3] == "X10" or
                                                      array_scriere_sheet2[x][9][:2] == "SP" or
                                                      array_scriere_sheet2[x][9][:3] == "X11" or
                                                      array_scriere_sheet2[x][9][:3] == "X12"):

            array_scriere_sheet2[x].append("Error")
        elif int(array_scriere_sheet2[x][12]) >= 3 and (array_scriere_sheet2[x][9][:2] != "X9" and
                                                        array_scriere_sheet2[x][9][:3] != "X10" and
                                                        array_scriere_sheet2[x][9][:2] != "SP" and
                                                        array_scriere_sheet2[x][9][:3] != "X11" and
                                                        array_scriere_sheet2[x][9][:3] != "X12"):
            array_scriere_sheet2[x].append("Error")
        else:
            array_scriere_sheet2[x].append("OK")
    for x in range(1, len(array_scriere_sheet3)):
        if int(array_scriere_sheet3[x][12]) == 1 and (array_scriere_sheet3[x][9][:2] == "X9" or
                                                      array_scriere_sheet3[x][9][:3] == "X10" or
                                                      array_scriere_sheet3[x][9][:2] == "SP" or
                                                      array_scriere_sheet3[x][9][:3] == "X11" or
                                                      array_scriere_sheet3[x][9][:3] == "X12"):
            array_scriere_sheet3[x].append("Error")
        elif int(array_scriere_sheet3[x][12]) >= 3 and (array_scriere_sheet3[x][9][:2] != "X9" and
                                                        array_scriere_sheet3[x][9][:3] != "X10" and
                                                        array_scriere_sheet3[x][9][:2] != "SP" and
                                                        array_scriere_sheet3[x][9][:3] != "X11" and
                                                        array_scriere_sheet3[x][9][:3] != "X12"):

            array_scriere_sheet3[x].append("Error")
        else:
            array_scriere_sheet3[x].append("OK")
    "eroare double"
    for x in range(1, len(array_scriere_sheet2)):
        if int(array_scriere_sheet2[x][12]) == 2 and (array_scriere_sheet2[x][9][:2] != "X9" and
                                                      array_scriere_sheet2[x][9][:3] != "X10" and
                                                      array_scriere_sheet2[x][9][:2] != "SP" and
                                                      array_scriere_sheet2[x][9][:3] != "X11" and
                                                      array_scriere_sheet2[x][9][:3] != "X12") and \
                not array_scriere_sheet2[x][1] == "81.25480-5848" and array_scriere_sheet2[x][2] != "591003_1":
            array_scriere_sheet2[x].append("Error")
        else:
            array_scriere_sheet2[x].append("OK")
    for x in range(1, len(array_scriere_sheet3)):
        if int(array_scriere_sheet3[x][12]) == 2 and (array_scriere_sheet3[x][9][:2] != "X9" and
                                                      array_scriere_sheet3[x][9][:3] != "X10" and
                                                      array_scriere_sheet3[x][9][:2] != "SP" and
                                                      array_scriere_sheet3[x][9][:3] != "X11" and
                                                      array_scriere_sheet3[x][9][:3] != "X12") and not \
                array_scriere_sheet3[x][1] == "81.25480-5848" and array_scriere_sheet3[x][2] != "591003_1":
            array_scriere_sheet3[x].append("Error")
        else:
            array_scriere_sheet3[x].append("OK")
    "eroare cross section"
    cross_section = {}
    for i in range(1, len(array_scriere_sheet2)):
        if array_scriere_sheet2[i][9] not in cross_section:
            cross_section[array_scriere_sheet2[i][9]] = float(array_scriere_sheet2[i][5])
        else:
            cross_section[array_scriere_sheet2[i][9]] += float(array_scriere_sheet2[i][5])
    for x in range(1, len(array_scriere_sheet2)):
        if cross_section[array_scriere_sheet2[x][9]] > 48:
            array_scriere_sheet2[x].append("Error")
        else:
            array_scriere_sheet2[x].append("OK")
    cross_section2 = {}
    for i in range(1, len(array_scriere_sheet3)):
        if array_scriere_sheet3[i][9] not in cross_section2:
            cross_section2[array_scriere_sheet3[i][9]] = float(array_scriere_sheet3[i][5])
        else:
            cross_section2[array_scriere_sheet3[i][9]] += float(array_scriere_sheet3[i][5])
    for x in range(1, len(array_scriere_sheet3)):
        if cross_section2[array_scriere_sheet3[x][9]] > 48:
            array_scriere_sheet3[x].append("Error")
        else:
            array_scriere_sheet3[x].append("OK")
    "erorare combinatii"
    for i in range(1, len(array_scriere_sheet2)):
        if array_scriere_sheet2[i][12] > 2:
            array_scriere_sheet2[i].append(pivotare(array_scriere_sheet2, array_scriere_sheet2[i][9]))
        else:
            array_scriere_sheet2[i].append("OK")
    for i in range(1, len(array_scriere_sheet3)):
        if array_scriere_sheet3[i][12] > 2:
            array_scriere_sheet3[i].append(pivotare(array_scriere_sheet3, array_scriere_sheet3[i][9]))
        else:
            array_scriere_sheet3[i].append("OK")
    # adaugareSonderltg
    for i in range(1, len(array_scriere_sheet2)):
        for x in range(1, len(array_wires_1)):
            if array_scriere_sheet2[i][1] == array_wires_1[x][0] and array_scriere_sheet2[i][2] == array_wires_1[x][1]:
                array_scriere_sheet2[i].append(array_wires_1[x][9])
    for i in range(1, len(array_scriere_sheet3)):
        for x in range(1, len(array_wires_2)):
            if array_scriere_sheet3[i][1] == array_wires_2[x][0] and array_scriere_sheet3[i][2] == array_wires_2[x][1]:
                array_scriere_sheet3[i].append(array_wires_2[x][9])
    variatie_lungimi(arr_module_file2, array_scriere_sheet2, array_scriere_sheet3)


def variatie_lungimi(sheet1, sheet2, sheet3):
    array_scriere_sheet4 = [["Type", "Abbreviation", "Part_Number_2", "Side", "q-VLA", "r-VLA/RQT",
                             "s-Radstand", "t-NLA", "u-Aoeberhang", "Heck module", "Identic"]]
    "Load required data files"
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Langenmodule.txt", newline='') as csvfile:
        langenmodule = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Heck Modules.txt", newline='') as csvfile:
        heckmodule = list(csv.reader(csvfile, delimiter=';'))
    for i in range(1, len(sheet1)):
        for x in range(len(langenmodule)):
            if sheet1[i][1] == langenmodule[x][4]:
                array_scriere_sheet4.append(
                    [langenmodule[x][0], langenmodule[x][10], langenmodule[x][4], langenmodule[x][0],
                     langenmodule[x][11]])
    for i in range(1, len(array_scriere_sheet4)):
        for x in range(len(langenmodule)):
            if array_scriere_sheet4[i][2] == langenmodule[x][5]:
                array_scriere_sheet4[i].append(langenmodule[x][12])
    for i in range(1, len(array_scriere_sheet4)):
        for x in range(len(langenmodule)):
            if array_scriere_sheet4[i][2] == langenmodule[x][6]:
                array_scriere_sheet4[i].append(langenmodule[x][13])
    for i in range(1, len(array_scriere_sheet4)):
        for x in range(len(langenmodule)):
            if array_scriere_sheet4[i][2] == langenmodule[x][7]:
                array_scriere_sheet4[i].append(langenmodule[x][14])
    for i in range(1, len(array_scriere_sheet4)):
        for x in range(len(langenmodule)):
            if array_scriere_sheet4[i][2] == langenmodule[x][8]:
                array_scriere_sheet4[i].append(langenmodule[x][15])
    for i in range(1, len(array_scriere_sheet4)):
        for x in range(9 - len(array_scriere_sheet4[i])):
            array_scriere_sheet4[i].append("")
    for i in range(1, len(array_scriere_sheet4)):
        for x in range(len(heckmodule)):
            if array_scriere_sheet4[i][2] == heckmodule[x][0]:
                array_scriere_sheet4[i].append(heckmodule[x][0])
    for i in range(1, len(array_scriere_sheet4)):
        for x in range(len(array_scriere_sheet4[i])):
            if array_scriere_sheet4[i][0] == "Sat Li":
                array_scriere_sheet4[i][0] = "8011"
                array_scriere_sheet4[i][3] = "LEFT"
            elif array_scriere_sheet4[i][0] == "Sat Re":
                array_scriere_sheet4[i][0] = "8013"
                array_scriere_sheet4[i][3] = "RIGHT"
            elif array_scriere_sheet4[i][0] == "Cha Li":
                array_scriere_sheet4[i][0] = "8012"
                array_scriere_sheet4[i][3] = "LEFT"
            elif array_scriere_sheet4[i][0] == "Cha Re":
                array_scriere_sheet4[i][0] = "8014"
                array_scriere_sheet4[i][3] = "RIGHT"
            elif array_scriere_sheet4[i][0] == "TGLM Li":
                array_scriere_sheet4[i][0] = "8023"
                array_scriere_sheet4[i][3] = "LEFT"
            elif array_scriere_sheet4[i][0] == "TGLM Re":
                array_scriere_sheet4[i][0] = "8024"
                array_scriere_sheet4[i][3] = "RIGHT"
            elif array_scriere_sheet4[i][0] == "4 Ach Li":
                array_scriere_sheet4[i][0] = "8025"
                array_scriere_sheet4[i][3] = "LEFT"
            elif array_scriere_sheet4[i][0] == "4 Ach Re":
                array_scriere_sheet4[i][0] = "8026"
                array_scriere_sheet4[i][3] = "RIGHT"
            elif array_scriere_sheet4[i][0] == "4 Ach Mil Li":
                array_scriere_sheet4[i][0] = "8000"
                array_scriere_sheet4[i][3] = "LEFT"
            elif array_scriere_sheet4[i][0] == "4 Ach Mil Re":
                array_scriere_sheet4[i][0] = "8001"
                array_scriere_sheet4[i][3] = "RIGHT"
    copiere_erori(sheet1, sheet2, sheet3, array_scriere_sheet4)


def copiere_erori(sheet1, sheet2, sheet3, sheet4):
    array_scriere_sheet5 = [["Harness", 'Module', 'Ltg-Nr.', 'Leitung', 'Farbe', 'Quer.', 'Kurzname', 'Pin', 'Lange',
                             'Kurzname/Pin', 'K/P Count', "Duplicate"]]
    arr_count = []
    for x in range(11, 14):
        for i in range(1, len(sheet2)):
            if sheet2[i][x] == "Error":
                array_scriere_sheet5.append(
                    [sheet2[i][0], sheet2[i][1], sheet2[i][2], sheet2[i][3], sheet2[i][4], sheet2[i][5],
                     sheet2[i][6], sheet2[i][7], sheet2[i][8], sheet2[i][9], sheet2[i][10]])
    for x in range(11, 14):
        for i in range(1, len(sheet3)):
            if sheet3[i][x] == "Error":
                array_scriere_sheet5.append(
                    [sheet3[i][0], sheet3[i][1], sheet3[i][2], sheet3[i][3], sheet3[i][4], sheet3[i][5],
                     sheet3[i][6], sheet3[i][7], sheet3[i][8], sheet3[i][9], sheet3[i][10]])
    for i in range(1, len(array_scriere_sheet5)):
        arr_count.append(array_scriere_sheet5[i][2])
    for i in range(1, len(array_scriere_sheet5)):
        if arr_count.count(array_scriere_sheet5[i][2]) > 1 and array_scriere_sheet5[i][10] == 2:
            array_scriere_sheet5[i].append("Duplicat")
        else:
            array_scriere_sheet5[i].append("OK")
    klappschale(sheet1, sheet2, sheet3, sheet4, array_scriere_sheet5)


def klappschale(sheet1, sheet2, sheet3, sheet4, sheet5):
    # Load required data files
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Tabel Klappschale.txt", newline='') as csvfile:
        arr_tabel_klappschale = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Bracket Side.txt", newline='') as csvfile:
        arr_bracket_side = list(csv.reader(csvfile, delimiter=';'))
    try:
        array_scriere_sheet6 = [["Conector", "Klapppschale", "Steering Side", "Harness", sheet2[1][0], sheet3[1][0]]]
    except IndexError:
        messagebox.showinfo("Eroare klappschale!",
                            "Pentru " + str(sheet1[1][0]) + " nu am gasit modulele in nici un wirelist!")
        return None
    for i in range(len(arr_tabel_klappschale)):
        if sheet2[1][0].replace(".MY23", "") in arr_tabel_klappschale[i][0]:
            array_scriere_sheet6.append([arr_tabel_klappschale[i][1], arr_tabel_klappschale[i][2],
                                         arr_tabel_klappschale[i][3], arr_tabel_klappschale[i][4]])
    if len(array_scriere_sheet6) == 1:
        messagebox.showinfo("Eroare klappschale!",
                            "Pentru " + str(sheet1[1][0]) + " nu exista informatii in tabelul de klappschale!" +
                            "Nu se va salva nimic")
        return None

    for i in range(1, len(array_scriere_sheet6)):
        counter = 0
        for x in range(len(sheet2)):
            if array_scriere_sheet6[i][0] in sheet2[x][6]:
                counter = counter + 1
        if counter > 0:
            array_scriere_sheet6[i].append("X")
        else:
            array_scriere_sheet6[i].append("0")
    for i in range(1, len(array_scriere_sheet6)):
        counter1 = 0
        for x in range(len(sheet3)):
            if array_scriere_sheet6[i][0] in sheet3[x][6]:
                counter1 = counter1 + 1
        if counter1 > 0:
            array_scriere_sheet6[i].append("X")
        else:
            array_scriere_sheet6[i].append("0")
    for i in range(len(array_scriere_sheet6)):
        array_scriere_sheet6[i].append("")
    arr_module_existente = [["Module", "Drawing", "Quantity"]]
    lista_klappschale = []
    for i in range(1, len(sheet1)):
        if "Klapp" in sheet1[i][4]:
            arr_module_existente.append([sheet1[i][1], sheet1[i][7], sheet1[i][8]])
            lista_klappschale.append([sheet1[i][1], sheet1[i][7], sheet1[i][8]])
    # Verificare bracket side
    sidebracket = "Error"
    for i in range(len(arr_bracket_side[0])):
        for x in range(1, len(sheet1)):
            if sheet1[x][1] == arr_bracket_side[0][i]:
                sidebracket = "LHD"
                break
    for y in range(len(arr_bracket_side[1])):
        for u in range(1, len(sheet1)):
            if sheet1[u][1] == arr_bracket_side[1][y]:
                sidebracket = "RHD"
                break
    # Verificare module klappschale lipsa
    arr_module_absente = []
    arr_module_existente_ver = []
    for x in range(1, len(arr_module_existente)):
        arr_module_existente_ver.append(arr_module_existente[x][0])
    for i in range(1, len(array_scriere_sheet6)):
        if array_scriere_sheet6[i][2] == sidebracket and array_scriere_sheet6[0][4] in array_scriere_sheet6[i][3]:
            if array_scriere_sheet6[i][5] == "X" and not array_scriere_sheet6[i][1] in arr_module_absente:
                if array_scriere_sheet6[i][1] not in arr_module_existente_ver:
                    arr_module_absente.append(array_scriere_sheet6[i][1])
    for i in range(1, len(array_scriere_sheet6)):
        if array_scriere_sheet6[i][2] == sidebracket and array_scriere_sheet6[0][5] in array_scriere_sheet6[i][3]:
            if array_scriere_sheet6[i][5] == "X" and not array_scriere_sheet6[i][1] in arr_module_absente:
                if array_scriere_sheet6[i][1] not in arr_module_existente_ver:
                    arr_module_absente.append(array_scriere_sheet6[i][1])
    array_scriere_sheet6[0].append("Side bracket")
    array_scriere_sheet6[1].append(sidebracket)
    for i in range(2, len(array_scriere_sheet6)):
        array_scriere_sheet6[i].append("")
    for i in range(len(array_scriere_sheet6)):
        array_scriere_sheet6[i].append("")
    for i in range(len(arr_module_existente)):
        array_scriere_sheet6[i].extend(arr_module_existente[i])
    for i in range(len(arr_module_existente), len(array_scriere_sheet6)):
        array_scriere_sheet6[i].append("")
        array_scriere_sheet6[i].append("")
        array_scriere_sheet6[i].append("")
    for i in range(len(array_scriere_sheet6)):
        array_scriere_sheet6[i].append("")
    array_scriere_sheet6[0].append("Module absente")
    for i in range(len(arr_module_absente)):
        array_scriere_sheet6[i + 1].append(arr_module_absente[i])
    for i in range(len(arr_module_absente) + 1, len(array_scriere_sheet6)):
        array_scriere_sheet6[i].append("")
    # Verificare desen klappschale (lhd sau rhd)
    for x in range(len(lista_klappschale)):
        for i in range(len(array_scriere_sheet6)):
            if lista_klappschale[x][0] == array_scriere_sheet6[i][1] and array_scriere_sheet6[i][2] == sidebracket \
                    and array_scriere_sheet6[i][4] == "X":
                array_scriere_sheet6[x + 1][12] = array_scriere_sheet6[0][4]
            elif lista_klappschale[x][0] == array_scriere_sheet6[i][1] and array_scriere_sheet6[i][2] == sidebracket \
                    and array_scriere_sheet6[i][5] == "X":
                array_scriere_sheet6[x + 1][12] = array_scriere_sheet6[0][5]

    lista_rl_klappschale = [["8011", "BODYL"], ["8012", "BODYL"], ["8013", "BODYR"], ["8014", "BODYR"],
                            ["8014", "BODYR"], ["8023", "BODYL"], ["8024", "BODYR"], ["8025", "BODYL"],
                            ["8026", "BODYR"], ["8030", "BODYR"], ["8001", "BODYR"], ["8000", "BODYL"],
                            ["8001", "BODYR"], ["8022", "BODYR"], ["8023", "BODYL"], ["8026", "BODYL"],
                            ["8027", "BODYR"], ["8027", "BODYR"], ["8030", "BODYR"], ["8031", "BODYL"],
                            ["8032", "BODYR"], ["8032", "BODYR"], ["8033", "BODYL"], ["8044", "BODYR"],
                            ["8057", "BODYL"], ["8000", "BODYL"], ["8052", "BODYL"], ["8053", "BODYR"]]

    for i in range(len(array_scriere_sheet6)):
        for x in range(len(lista_rl_klappschale)):
            if array_scriere_sheet6[i][12] == lista_rl_klappschale[x][0]:
                array_scriere_sheet6[i][12] = lista_rl_klappschale[x][1]

    # Verificare integritate(sa nu fie platforme combinate)
    ch = 0
    st = 0
    tg = 0
    a4 = 0
    mil = 0
    for i in range(1, len(sheet1)):
        if "CHASSIS" in sheet1[i][5]:
            ch = ch + 1
        elif "SATTEL" in sheet1[i][5]:
            st = st + 1
        elif "TGLM" in sheet1[i][5]:
            tg = tg + 1
        elif "4AXEL" in sheet1[i][5]:
            a4 = a4 + 1
        elif "Mil" in sheet1[i][5] or "MIL" in sheet1[i][5]:
            mil = mil + 1
    intergitate = "NOT OK"
    if ch > 0 and st == 0 and tg == 0 and a4 == 0 and mil == 0:
        intergitate = "OK"
    if ch == 0 and st > 0 and tg == 0 and a4 == 0 and mil == 0:
        intergitate = "OK"
    if ch == 0 and st == 0 and tg > 0 and a4 == 0 and mil == 0:
        intergitate = "OK"
    if ch == 0 and st == 0 and tg == 0 and a4 > 0 and mil == 0:
        intergitate = "OK"
    if ch == 0 and st == 0 and tg == 0 and a4 == 0 and mil > 0:
        intergitate = "OK"
    for i in range(len(array_scriere_sheet6)):
        array_scriere_sheet6[i].append("")
    array_scriere_sheet6[0].append("Integritate")
    array_scriere_sheet6[0].append(intergitate)
    array_scriere_sheet6[1].append("CHASSIS")
    array_scriere_sheet6[1].append(ch)
    array_scriere_sheet6[2].append("SATTEL")
    array_scriere_sheet6[2].append(st)
    array_scriere_sheet6[3].append("TGLM")
    array_scriere_sheet6[3].append(tg)
    array_scriere_sheet6[4].append("4AXEL")
    array_scriere_sheet6[4].append(a4)
    array_scriere_sheet6[5].append("Military")
    array_scriere_sheet6[5].append(mil)
    for i in range(6, len(array_scriere_sheet6)):
        array_scriere_sheet6[i].append("")
        array_scriere_sheet6[i].append("")
    bkk(sheet1, sheet2, sheet3, sheet4, sheet5, array_scriere_sheet6)


def bkk(sheet1, sheet2, sheet3, sheet4, sheet5, sheet6):
    arr_bkk_left = []
    arr_bkk_right = []
    arr_erori_bkk = []
    "Load required data files"
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Tabel BKK.txt", newline='') as csvfile:
        arr_tabel_bkk = list(csv.reader(csvfile, delimiter=';'))
    for i in range(len(arr_tabel_bkk)):
        if arr_tabel_bkk[i][1] == "BKK-Left":
            arr_bkk_left.append(arr_tabel_bkk[i][0])
        elif arr_tabel_bkk[i][1] == "BKK-Right":
            arr_bkk_right.append(arr_tabel_bkk[i][0])
    left_val = 0
    right_val = 0
    if "8013" in sheet3[1][0] or "8014" in sheet3[1][0] or "8024" in sheet3[1][0] or "8026" in sheet3[1][0]:
        for i in range(1, len(sheet3)):
            if sheet3[i][6] in arr_bkk_left:
                left_val = left_val + 1
                arr_erori_bkk.append(sheet3[i][6])
        for i in range(1, len(sheet3)):
            if sheet3[i][6] in arr_bkk_right:
                right_val = right_val + 1
                arr_erori_bkk.append(sheet3[i][6])

    if "8013" in sheet2[1][0] or "8014" in sheet2[1][0] or "8024" in sheet2[1][0] or "8026" in sheet2[1][0]:
        for i in range(1, len(sheet2)):
            if sheet2[i][6] in arr_bkk_left:
                left_val = left_val + 1
                arr_erori_bkk.append(sheet2[i][6])
        for i in range(1, len(sheet2)):
            if sheet2[i][6] in arr_bkk_right:
                right_val = right_val + 1
                arr_erori_bkk.append(sheet2[i][6])
    arr_erori_bkk = list(dict.fromkeys(arr_erori_bkk))
    bkkstatus = "NOT OK"
    if left_val == 0 and right_val == 0:
        bkkstatus = "OK"
    elif left_val == 0 and right_val > 0:
        bkkstatus = "OK"
    elif left_val > 0 and right_val == 0:
        bkkstatus = "OK"
    for i in range(len(sheet6)):
        sheet6[i].append("")
    sheet6[0].append("Status BKK")
    sheet6[1].append(bkkstatus)
    for i in range(len(arr_erori_bkk)):
        sheet6[i + 2].append(arr_erori_bkk[i])
    for i in range(len(arr_erori_bkk) + 2, len(sheet6)):
        sheet6[i].append("")
    samewire(sheet1, sheet2, sheet3, sheet4, sheet5, sheet6)


def samewire(sheet1, sheet2, sheet3, sheet4, sheet5, sheet6):
    arr_sheet7 = [["Harness", "Module", "Ltg-Nr.", "Leitung", "Farbe", "Quer.", "Kurzname", "Pin", "Lange",
                   "Kurzname/Pin", "K/P Count", "Verificare"]]
    arr_resistor = []
    arr_verificare = []
    x66161a1 = []
    x66162a1 = []
    x64901a1 = []
    x64902a1 = []
    "module in configuratie"
    moduleinconfig = []
    for i in range(1, len(sheet1)):
        moduleinconfig.append(sheet1[i][1])
    for i in range(len(sheet2)):
        if "RESISTOR" in sheet2[i][2]:
            arr_resistor.append(sheet2[i][1])
    for i in range(len(sheet3)):
        if "RESISTOR" in sheet3[i][2]:
            arr_resistor.append(sheet3[i][1])
    for i in range(1, len(sheet2)):
        if sheet2[i][10] == 2 and not sheet2[i][1] in arr_resistor and sheet2[i][2] != "591003_1":
            arr_sheet7.append([sheet2[i][0], sheet2[i][1], sheet2[i][2], sheet2[i][3], sheet2[i][4], sheet2[i][5],
                               sheet2[i][6], sheet2[i][7], sheet2[i][8], sheet2[i][9], sheet2[i][10]])
    for i in range(1, len(sheet3)):
        if sheet3[i][10] == 2 and not sheet3[i][1] in arr_resistor and sheet3[i][2] != "591003_1":
            arr_sheet7.append([sheet3[i][0], sheet3[i][1], sheet3[i][2], sheet3[i][3], sheet3[i][4], sheet3[i][5],
                               sheet3[i][6], sheet3[i][7], sheet3[i][8], sheet3[i][9], sheet3[i][10]])
    for i in range(1, len(arr_sheet7)):
        arr_verificare.append(arr_sheet7[i][2])
    for i in range(1, len(arr_sheet7)):
        arr_sheet7[i].append(arr_verificare.count(arr_sheet7[i][2]))
    "x6616/x6490"
    x6616sheet = [[], [], [], []]
    for i in range(1, len(sheet2)):
        if sheet2[i][6] == "X6616.1A1":
            x66161a1.append(int(sheet2[i][7]))
            x6616sheet[0].append([sheet2[i][6], int(sheet2[i][7])])
        elif sheet2[i][6] == "X6616.2A1":
            x66162a1.append(int(sheet2[i][7]))
            x6616sheet[1].append([sheet2[i][6], int(sheet2[i][7])])
        elif sheet2[i][6] == "X6490.1A1":
            x64901a1.append(int(sheet2[i][7]))
            x6616sheet[2].append([sheet2[i][6], int(sheet2[i][7])])
        elif sheet2[i][6] == "X6490.2A1":
            x64902a1.append(int(sheet2[i][7]))
            x6616sheet[3].append([sheet2[i][6], int(sheet2[i][7])])
    for i in range(1, len(sheet3)):
        if sheet3[i][6] == "X6616.1A1":
            x66161a1.append(int(sheet3[i][7]))
            x6616sheet[0].append([sheet3[i][6], int(sheet3[i][7])])
        elif sheet3[i][6] == "X6616.2A1":
            x66162a1.append(int(sheet3[i][7]))
            x6616sheet[1].append([sheet3[i][6], int(sheet3[i][7])])
        elif sheet3[i][6] == "X6490.1A1":
            x64901a1.append(int(sheet3[i][7]))
            x6616sheet[2].append([sheet3[i][6], int(sheet3[i][7])])
        elif sheet3[i][6] == "X6490.2A1":
            x64902a1.append(int(sheet3[i][7]))
            x6616sheet[3].append([sheet3[i][6], int(sheet3[i][7])])
    x66161a1.sort()
    x66162a1.sort()
    x64901a1.sort()
    x64902a1.sort()
    x6616sheetsortat = [[], [], [], []]
    for item in x66161a1:
        x6616sheetsortat[0].append(["X6616.1A1", item])
    for item in x66162a1:
        x6616sheetsortat[1].append(["X6616.2A1", item])
    for item in x64901a1:
        x6616sheetsortat[2].append(["X6490.1A1", item])
    for item in x64902a1:
        x6616sheetsortat[3].append(["X6490.2A1", item])
    if sheet6[4][16] > 0:
        if x66161a1 == x66162a1:
            x6616 = "OK"
        else:
            x6616 = "NOT OK"
        if x64901a1 == x64902a1:
            x6490 = "OK"
        else:
            try:
                x64902a1.remove(12)
                if x64902a1 == x64901a1 and "81.25481-7608" in moduleinconfig and "81.25481-5580" in moduleinconfig:
                    x6490 = "OK"
                else:
                    x6490 = "NOT OK"
            except:
                if x64902a1 == x64901a1:
                    x6490 = "OK"
                else:
                    x6490 = "NOT OK"
    elif sheet6[2][16] > 0:
        if x66161a1 == x66162a1:
            x6616 = "OK"
        else:
            try:
                x66162a1.remove(36)
                if x66162a1 == x66161a1:
                    x6616 = "OK"
                else:
                    x6616 = "NOT OK"
            except:
                if x66162a1 == x66161a1:
                    x6616 = "OK"
                else:
                    x6616 = "NOT OK"
        if x64901a1 == x64902a1:
            x6490 = "OK"
        else:
            x6490 = "NOT OK"
    else:
        if x66161a1 == x66162a1:
            x6616 = "OK"
        else:
            x6616 = "NOT OK"
        if x64901a1 == x64902a1:
            x6490 = "OK"
        else:
            x6490 = "NOT OK"
    if len(x64901a1) == 0 and len(x64902a1) == 0:
        x6490 = "OK"
    if len(x66161a1) == 0 and len(x66162a1) == 0:
        x6616 = "OK"

    for i in range(len(sheet6)):
        sheet6[i].append("")
    sheet6[0].append("Status X6616")
    sheet6[0].append(x6616)
    sheet6[2].append("Status X6490")
    sheet6[2].append(x6490)
    prn_excel_wires_validare(sheet1, sheet2, sheet3, sheet4, sheet5, sheet6, arr_sheet7, x6616sheetsortat)


def prn_excel_wires_validare(sheet1, sheet2, sheet3, sheet4, sheet5, sheet6, sheet7, sheet8):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    wb = Workbook()
    ws1 = wb.active
    ws1.title = sheet1[1][0]
    ws2 = wb.create_sheet(sheet2[1][0])
    ws3 = wb.create_sheet(sheet3[1][0])
    ws4 = wb.create_sheet("Variatie Lungimi")
    ws5 = wb.create_sheet("Erori")
    ws6 = wb.create_sheet("Bracket")
    ws7 = wb.create_sheet("Same Wire")
    ws8 = wb.create_sheet("X6616X6490")
    ws9 = wb.create_sheet("Variatie Lungimi IST SOLL")
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            if "E-" in sheet1[i][x]:
                try:
                    ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
                except:
                    ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
            else:
                try:
                    ws1.cell(column=x + 1, row=i + 1, value=float(sheet1[i][x]))
                except:
                    ws1.cell(column=x + 1, row=i + 1, value=sheet1[i][x])
    for i in range(len(sheet2)):
        for x in range(len(sheet2[i])):
            if x == 2:
                ws2.cell(column=x + 1, row=i + 1, value=str(sheet2[i][x]))
            else:
                try:
                    ws2.cell(column=x + 1, row=i + 1, value=float(sheet2[i][x]))
                except:
                    ws2.cell(column=x + 1, row=i + 1, value=sheet2[i][x])
    for i in range(len(sheet3)):
        for x in range(len(sheet3[i])):
            if x == 2:
                ws3.cell(column=x + 1, row=i + 1, value=str(sheet3[i][x]))
            else:
                try:
                    ws3.cell(column=x + 1, row=i + 1, value=float(sheet3[i][x]))
                except:
                    ws3.cell(column=x + 1, row=i + 1, value=sheet3[i][x])
    for i in range(len(sheet4)):
        for x in range(len(sheet4[i])):
            try:
                ws4.cell(column=x + 1, row=i + 1, value=float(sheet4[i][x]))
            except:
                ws4.cell(column=x + 1, row=i + 1, value=sheet4[i][x])
    for i in range(len(sheet5)):
        for x in range(len(sheet5[i])):
            if x == 2:
                ws5.cell(column=x + 1, row=i + 1, value=str(sheet5[i][x]))
            else:
                try:
                    ws5.cell(column=x + 1, row=i + 1, value=float(sheet5[i][x]))
                except:
                    ws5.cell(column=x + 1, row=i + 1, value=sheet5[i][x])
    for i in range(len(sheet6)):
        for x in range(len(sheet6[i])):
            try:
                ws6.cell(column=x + 1, row=i + 1, value=float(sheet6[i][x]))
            except:
                ws6.cell(column=x + 1, row=i + 1, value=sheet6[i][x])
    """Cosmetica"""
    colorsgray = PatternFill(start_color='aabbcc', end_color='aabbcc', fill_type='solid')
    colorsdoi = PatternFill(start_color='7EF1EA', end_color='7EF1EA', fill_type='solid')
    for i in range(1, 7):
        for x in range(1, len(ws6['A']) + 1):
            ws6.cell(column=i, row=x).border = thin_border
            ws6.cell(column=i, row=x).alignment = Alignment(horizontal='center')
    for row in ws6['C']:
        if row.value == "RHD":
            for i in range(1, 7):
                ws6.cell(column=i, row=row.row).fill = colorsgray
        elif row.value == "LHD":
            for i in range(1, 7):
                ws6.cell(column=i, row=row.row).fill = colorsdoi
    for cells in ws1['1']:
        ws1.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws2['1']:
        ws2.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws3['1']:
        ws3.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws4['1']:
        ws4.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws5['1']:
        ws5.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws6['1']:
        ws6.cell(column=cells.column, row=1).font = Font(bold=True)
    for cells in ws7['1']:
        ws7.cell(column=cells.column, row=1).font = Font(bold=True)
    """Cosmetica"""
    for i in range(len(sheet7)):
        for x in range(len(sheet7[i])):
            if x == 2:
                ws7.cell(column=x + 1, row=i + 1, value=str(sheet7[i][x]))
            else:
                try:
                    ws7.cell(column=x + 1, row=i + 1, value=float(sheet7[i][x]))
                except:
                    ws7.cell(column=x + 1, row=i + 1, value=sheet7[i][x])
    for x in range(len(sheet8[0])):
        try:
            ws8.cell(column=1, row=x + 1, value=float(sheet8[0][x][0]))
            ws8.cell(column=2, row=x + 1, value=float(sheet8[0][x][1]))
        except:
            ws8.cell(column=1, row=x + 1, value=sheet8[0][x][0])
            ws8.cell(column=2, row=x + 1, value=sheet8[0][x][1])
    for x in range(len(sheet8[1])):
        try:
            ws8.cell(column=4, row=x + 1, value=float(sheet8[1][x][0]))
            ws8.cell(column=5, row=x + 1, value=float(sheet8[1][x][1]))
        except:
            ws8.cell(column=4, row=x + 1, value=sheet8[1][x][0])
            ws8.cell(column=5, row=x + 1, value=sheet8[1][x][1])
    for x in range(len(sheet8[2])):
        try:
            ws8.cell(column=7, row=x + 1, value=float(sheet8[2][x][0]))
            ws8.cell(column=8, row=x + 1, value=float(sheet8[2][x][1]))
        except:
            ws8.cell(column=7, row=x + 1, value=sheet8[2][x][0])
            ws8.cell(column=8, row=x + 1, value=sheet8[2][x][1])
    for x in range(len(sheet8[3])):
        try:
            ws8.cell(column=10, row=x + 1, value=float(sheet8[3][x][0]))
            ws8.cell(column=11, row=x + 1, value=float(sheet8[3][x][1]))
        except:
            ws8.cell(column=10, row=x + 1, value=sheet8[3][x][0])
            ws8.cell(column=11, row=x + 1, value=sheet8[3][x][1])
    for i in range(len(sheet4)):
        ws9.cell(column=1, row=i + 1, value="IST")
        for x in range(len(sheet4[i])):
            try:
                ws9.cell(column=x + 2, row=i + 1, value=float(sheet4[i][x]))
            except:
                ws9.cell(column=x + 2, row=i + 1, value=sheet4[i][x])
    ws9.cell(column=1, row=1, value="IST / SOLL")
    ws9.insert_cols(5)
    ws9.cell(column=5, row=1, value="Description")
    istsoll(ws6, ws9)

    if globale.director_salvare == "":
        try:
            wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Validare/Wirelist/" +
                    sheet1[1][0] + ".xlsx")
            log_file("Creat " + sheet1[1][0] + ".xlsx")
        except PermissionError:
            log_file("Eroare scriere. Nu am salvat " + sheet1[1][0] + ".xlsx")
            messagebox.showerror('Eroare scriere', "Fisierul " + sheet1[1][0] + "este read-only!")
            return None
        except FileNotFoundError:
            messagebox.showerror('Eroare scriere', "Directorul /MAN/Output/Validare/Wirelist/ nu exista!")
            return None
    else:
        try:
            wb.save(globale.director_salvare + "/" + sheet1[1][0] + ".xlsx")
            log_file("Creat " + sheet1[1][0] + ".xlsx")
        except PermissionError:
            log_file("Eroare scriere. Nu am salvat " + sheet1[1][0] + ".xlsx")
            messagebox.showerror('Eroare scriere', "Fisierul " + sheet1[1][0] + "este read-only!")
            return None

    return None


def prelucrare_bom_faza1(array_prelucrare):
    # arr_sheet1 = [["Harness", "Module", "Module", "Drawing Number", "Module Check", "Comments", "Implemented", "Desen",
    #               "Cantitate"]]
    lista_module_extragere = []
    arr_sheet1 = [[array_prelucrare[i][0], array_prelucrare[i][1]] for i in range(1, len(array_prelucrare))]
    arr_sheet1.insert(0, ["Harness", "Module", "Module", "Drawing Number", "Module Check", "Comments", "Implemented",
                          "Desen", "Cantitate"])

    # for i in range(1, len(array_prelucrare)):
    #    arr_sheet1.append([array_prelucrare[i][0], array_prelucrare[i][1]])
    "Load required data files"
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Module Active.txt", newline='') as csvfile:
        arr_module_active = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Module Implementate.txt", newline='') as csvfile:
        arr_module_implementate = list(csv.reader(csvfile, delimiter=';'))

    "Prelucrare efectiva"

    for x in range(len(arr_module_active)):
        for y in range(1, len(arr_sheet1)):
            if arr_sheet1[y][1] in arr_module_active[x]:
                arr_sheet1[y].extend(arr_module_active[x])
    for i in range(1, len(arr_sheet1)):
        if len(arr_sheet1[i]) == 2:
            arr_sheet1[i].extend(["XXXX", "XXXX", "XXXX", "XXXX"])
    for x in range(len(arr_module_implementate)):
        for y in range(1, len(arr_sheet1)):
            if arr_sheet1[y][1] in arr_module_implementate[x]:
                arr_sheet1[y].append("OK")
    for i in range(1, len(arr_sheet1)):
        if len(arr_sheet1[i]) == 6:
            arr_sheet1[i].append("XXXX")
    for i in range(1, len(arr_sheet1)):
        if not arr_sheet1[i][5] in lista_module_extragere and arr_sheet1[i][5] != "XXXX":
            lista_module_extragere.append(arr_sheet1[i][5])
    for i in range(1, len(array_prelucrare)):
        if len(array_prelucrare[i][3]) > 10:
            arr_sheet1[i].extend([array_prelucrare[i][2], "0"])
        elif array_prelucrare[i][3] == "1000":
            arr_sheet1[i].extend([array_prelucrare[i][2], "1"])
        elif array_prelucrare[i][3] == "2000":
            arr_sheet1[i].extend([array_prelucrare[i][2], "2"])
        elif array_prelucrare[i][3] == "3000":
            arr_sheet1[i].extend([array_prelucrare[i][2], "3"])
        else:
            arr_sheet1[i].extend([array_prelucrare[i][2], array_prelucrare[i][3]])
    prelucrare_bom_faza2(arr_sheet1, lista_module_extragere)


def prelucrare_bom_faza2(arr_sheet1, listas):
    ##"Selectie fisiere wirelist"
    with open(os.path.abspath(os.curdir) + "/MAN/Input/Others/Component Overview.txt", newline='') as csvfile:
        arr_component_overview = list(csv.reader(csvfile, delimiter=';'))
    lista_selectie = (["SATTEL LHD", "8011"], ["SATTEL RHD", "8013"], ["CHASSIS LHD", "8012"], ["CHASSIS RHD", "8014"],
                      ["TGLM LHD", "8023"], ["TGLM RHD", "8024"], ["4AXEL LHD", "8025"], ["4AXEL RHD", "8026"],
                      ["4AXEL MIL LHD", "8000"], ["4AXEL MIL RHD", "8001"], ["CHASSIS MIL RHD", "8030"],
                      ["CHASSIS MIL LHD", "8031"], ["MIL_SAT RHD", "8052"], ["MIL_SAT LHD", "8053"])
    lista_fisiere = []
    arr_sheet2 = [["Module", "Quantity", "Bezei", "VOBES-ID", "Benennung", "Verwendung", "Verwendung", "Kurzname", "xy",
                   "Teilenummer", "Vorzugsteil", "TAB-Nummer", "Referenzteil", "Farbe", "E-Komponente",
                   "E-Komponente Part-Nr.", "Einh.", "Leoni PN 1", "Leoni PN 2", "Supplier No"]]
    for i in range(len(listas)):
        for x in range(len(lista_selectie)):
            if listas[i] in lista_selectie[x]:
                lista_fisiere.append(lista_selectie[x][1])
    "Load required data files"
    if len(lista_fisiere) != 2:
        lista_fisiere.append("9999")
        skip_file(arr_sheet1[1][0] + " Am gasit fisiere cu un singur wirelist!")
    with open(os.path.abspath(os.curdir) + "/MAN/Input/BOMs/" + lista_fisiere[0] + ".BOM.csv", newline='') as csvfile:
        arr_bom_1 = list(csv.reader(csvfile, delimiter=';'))
    with open(os.path.abspath(os.curdir) + "/MAN/Input/BOMs/" + lista_fisiere[1] + ".BOM.csv", newline='') as csvfile:
        arr_bom_2 = list(csv.reader(csvfile, delimiter=';'))
    "Prelucrare efectiva"
    for x in range(1, len(arr_sheet1)):
        for y in range(1, len(arr_bom_1)):
            if arr_sheet1[x][1] in arr_bom_1[y]:
                arr_sheet2.append(arr_bom_1[y])
    for x in range(1, len(arr_sheet1)):
        for y in range(1, len(arr_bom_2)):
            if arr_sheet1[x][1] in arr_bom_2[y]:
                arr_sheet2.append(arr_bom_2[y])
    for x in range(1, len(arr_sheet2)):
        for y in range(len(arr_component_overview)):

            if arr_sheet2[x][9] == arr_component_overview[y][0]:
                arr_sheet2[x].append(arr_component_overview[y][1])
                arr_sheet2[x].append(arr_component_overview[y][2])
                arr_sheet2[x].append(arr_component_overview[y][3])
    prn_excel_bom_validare(arr_sheet1, arr_sheet2)
    return None


def prn_excel_bom_validare(sheet1, sheet2):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = sheet1[1][0]
    ws2 = wb.create_sheet("BOM")
    for i in range(len(sheet1)):
        for x in range(len(sheet1[i])):
            if "E-" in sheet1[i][x]:
                try:
                    ws1.cell(column=x + 1, row=i + 1, value=str(sheet1[i][x]))
                except:
                    ws1.cell(column=x + 1, row=i + 1, value=str(float(sheet1[i][x])))
            else:
                try:
                    ws1.cell(column=x + 1, row=i + 1, value=float(sheet1[i][x]))
                except:
                    ws1.cell(column=x + 1, row=i + 1, value=sheet1[i][x])
    for i in range(len(sheet2)):
        for x in range(len(sheet2[i])):
            try:
                ws2.cell(column=x + 1, row=i + 1, value=float(sheet2[i][x]))
            except:
                ws2.cell(column=x + 1, row=i + 1, value=sheet2[i][x])
    if globale.director_salvare == "":
        try:
            wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Validare/BOM" + "/BOM " + sheet1[1][0] + ".xlsx")
            log_file("Creat BOM " + sheet1[1][0] + ".xlsx")
        except PermissionError:
            log_file("Eroare salvare. Nu am salvat BOM " + sheet1[1][0] + ".xlsx")
            messagebox.showerror('Eroare scriere', "Fisierul " + sheet1[1][0] + "este read-only!")
            return None
    else:
        try:
            wb.save(os.path.abspath(os.curdir) + "/MAN/Output/Validare/BOM" + sheet1[1][0] + ".xlsx")
            log_file("Creat BOM " + sheet1[1][0] + ".xlsx")
        except PermissionError:
            log_file("Eroare salvare. Nu am salvat BOM " + sheet1[1][0] + ".xlsx")
            messagebox.showerror('Eroare scriere', "Fisierul " + sheet1[1][0] + "este read-only!")
            return None
